"""
ZH: MYAI 兩段式送出（批次註冊／批次轉點）與初始點數發放的測試。

ZH: 這一組測試守的是三件會造成**不可逆損失**的事：
      1. 第一段的預覽回應被當成「做完了」（原本的缺陷：register_batch 只送預覽，
         呼叫端卻寫進資料庫、寄了開通信）。
      2. 轉點端點被誤用成回收／刪除（方向相反，會扣光學生的點數）。
      3. 同一個帳號重複發放初始點數（背景任務重入 → 發兩倍，收不回來）。

ZH: ⚠️ 這裡**不對廠商送出任何請求** —— 網路層一律用假的 client 攔下來。
"""
import sys
import os
import asyncio
from datetime import datetime, timezone

import pytest

sys.path.insert(0, os.path.join(os.path.dirname(os.path.dirname(
    os.path.abspath(__file__))), "job-scheduler"))

from app.services import myai_sync as M  # noqa: E402


# ══════════════════════════════════════════════════════════════════════════
# ZH: 廠商預覽頁的真實形狀（2026-08-29 實際抓下來的，只留必要的部分）
# ══════════════════════════════════════════════════════════════════════════
REGISTER_PREVIEW = """
<form action="register_batch_info" method="post" autocomplete="off">
<input type="text" readonly name="emails[]" value="a@example.com" title="a" />
<input type="text" readonly name="name_displays[]" value="&#28204;&#35430;" title="x" />
<input type="text" readonly name="passwords[]" value="Abc12345" title="x" />
<input type="text" readonly name="remarks[]" value="auto-provision" title="x" />
<input type="submit" value="確認" class="joinBt" />
</form>
"""

TRANSFER_PREVIEW = """
<form action="transfer_credit_batch_result" method="post" autocomplete="off">
<input type="text" readonly name="emails[]" value="a@example.com" title="a" />
<input type="text" readonly name="transferPoints[]" value="1" title="1" />
<input type="hidden" name="remarks[]" value="auto-provision" />
<input type="submit" value="確認" class="joinBt" />
</form>
"""

REG_FIELDS = ("emails", "name_displays", "passwords", "remarks")
TRF_FIELDS = ("emails", "transferPoints", "remarks")


class FakeResponse:
    """ZH: 假的 httpx 回應。"""

    def __init__(self, text, status_code=200):
        self.text = text
        self.status_code = status_code


def _stub_session(monkeypatch, responses):
    """ZH: 把 _session_request 換掉，依序回傳 responses，並記下被呼叫幾次。

    ZH: 回傳的 list 會被塞入每一次呼叫 —— 用它斷言「第二段到底有沒有送出」。
    """
    calls = []

    async def fake(do_fetch, is_valid):
        calls.append(len(calls))
        return responses[min(len(calls) - 1, len(responses) - 1)]

    monkeypatch.setattr(M, "_session_request", fake)
    return calls


# ══════════════════════════════════════════════════════════════════════════
# ZH: 一、預覽回應的解析與對帳
# ══════════════════════════════════════════════════════════════════════════

def test_echoed_rows_reads_the_vendor_preview():
    """ZH: 讀得出廠商回吐的每一欄（含 HTML entity 還原）。"""
    rows = M._echoed_rows(REGISTER_PREVIEW, REG_FIELDS)
    assert len(rows) == 1
    assert rows[0]["emails"] == "a@example.com"
    assert rows[0]["name_displays"] == "測試"     # ZH: &#28204;&#35430; 要還原
    assert rows[0]["remarks"] == "auto-provision"


def test_echoed_rows_reads_hidden_inputs_too():
    """ZH: 轉點的 remarks 是 hidden —— 一樣要讀得到，不然對帳會少一欄。"""
    rows = M._echoed_rows(TRANSFER_PREVIEW, TRF_FIELDS)
    assert rows == [{"emails": "a@example.com", "transferPoints": "1",
                     "remarks": "auto-provision"}]


def test_echoed_rows_rejects_misaligned_columns():
    """ZH: 欄位數量對不齊要拋錯，不能靜靜地回一個短掉的清單。"""
    broken = REGISTER_PREVIEW.replace('name="passwords[]"', 'name="zzz[]"')
    with pytest.raises(M.MyaiSyncError):
        M._echoed_rows(broken, REG_FIELDS)


# ══════════════════════════════════════════════════════════════════════════
# ZH: 二、第二段（確認送出）真的有送
# ══════════════════════════════════════════════════════════════════════════

def test_register_batch_actually_confirms(monkeypatch):
    """
    ZH: 🔴 這是整組測試的核心 —— 原本的缺陷就是**第二段從來沒送**。
        送出後必須有兩次請求（預覽 + 確認），且 created 為 True。
    """
    calls = _stub_session(monkeypatch, [FakeResponse(REGISTER_PREVIEW),
                                        FakeResponse("<html>完成</html>")])
    res = asyncio.run(M.register_batch(
        [{"email": "a@example.com", "nickname": "測試",
          "password": "Abc12345", "remark": "auto-provision"}]))
    assert len(calls) == 2, "只送了一次 = 第二段沒送出"
    assert res["created"] is True


def test_register_batch_aborts_when_vendor_dropped_a_row(monkeypatch):
    """
    ZH: 廠商只解析出 1 列、我們送了 2 列 → 中止，且**第二段不得送出**。
        （少解析通常是 email 重複；這時按確認會建出跟我們以為的不一樣的結果。）
    """
    calls = _stub_session(monkeypatch, [FakeResponse(REGISTER_PREVIEW),
                                        FakeResponse("<html>完成</html>")])
    rows = [{"email": "a@example.com", "nickname": "A", "password": "Abc12345"},
            {"email": "b@example.com", "nickname": "B", "password": "Abc12345"}]
    with pytest.raises(M.MyaiSyncError):
        asyncio.run(M.register_batch(rows))
    assert len(calls) == 1, "對帳失敗卻還是送出了確認"


def test_transfer_aborts_when_points_differ(monkeypatch):
    """
    ZH: 預覽回吐的點數與送出的不符 → 中止。**多發的點數收不回來。**
    """
    calls = _stub_session(monkeypatch, [FakeResponse(TRANSFER_PREVIEW),
                                        FakeResponse("<html>完成</html>")])
    with pytest.raises(M.MyaiSyncError):
        asyncio.run(M.transfer_credit_batch(
            [{"email": "a@example.com", "points": 999, "remark": "auto-provision"}],
            confirm_grant=True))
    assert len(calls) == 1


def test_transfer_confirms_when_reconciled(monkeypatch):
    """ZH: 對得上就送第二段。"""
    calls = _stub_session(monkeypatch, [FakeResponse(TRANSFER_PREVIEW),
                                        FakeResponse("<html>完成</html>")])
    res = asyncio.run(M.transfer_credit_batch(
        [{"email": "a@example.com", "points": 1, "remark": "auto-provision"}],
        confirm_grant=True))
    assert len(calls) == 2
    assert res["granted"] is True and res["points"] == 1


# ══════════════════════════════════════════════════════════════════════════
# ZH: 三、防呆 —— 不要誤用成回收／刪除
# ══════════════════════════════════════════════════════════════════════════

def test_transfer_refuses_without_explicit_confirm():
    """ZH: 沒有明確傳 confirm_grant 就不准送 —— 手滑呼叫不該會扣點。"""
    with pytest.raises(M.MyaiSyncError):
        asyncio.run(M.transfer_credit_batch(
            [{"email": "a@example.com", "points": 1}]))


@pytest.mark.parametrize("path", [
    "/mcu/gt_sdk/admin_168/user/get_credit_batch",      # ZH: 回收（名單＝被扣點的人）
    "/mcu/gt_sdk/admin_168/user/delete_batch",
    "/mcu/gt_sdk/admin_168/user/register_batch",
])
def test_transfer_guard_rejects_other_batch_endpoints(path):
    """ZH: 白名單只開 transfer_credit_batch 一族，其餘一律擋。"""
    with pytest.raises(M.MyaiSyncError):
        M._assert_transfer_endpoint(path, confirm_grant=True)


def test_transfer_guard_accepts_the_real_endpoints():
    """ZH: 陽性對照 —— 真正要用的兩條路徑必須過，否則上面那些測試等於沒測。"""
    M._assert_transfer_endpoint(M.TRANSFER_BATCH_CHECK_PATH, confirm_grant=True)
    M._assert_transfer_endpoint(M.TRANSFER_CONFIRM_PATH, confirm_grant=True)


@pytest.mark.parametrize("points", [0, -5])
def test_transfer_xlsx_rejects_non_positive_points(points):
    """ZH: 0 是白跑一趟；負數在廠商端的行為未知（可能變成扣點）。"""
    with pytest.raises(M.MyaiSyncError):
        M.build_transfer_xlsx([{"email": "a@example.com", "points": points}])


def test_transfer_xlsx_matches_the_vendor_template():
    """ZH: 三欄、無標題列、A=email B=點數 C=備註（與註冊的四欄不同）。"""
    from openpyxl import load_workbook
    import io as _io
    data = M.build_transfer_xlsx(
        [{"email": "a@example.com", "points": 7, "remark": "r"}])
    ws = load_workbook(_io.BytesIO(data)).active
    assert [c.value for c in ws[1]] == ["a@example.com", 7, "r"]
    assert ws.max_row == 1, "多了標題列"


# ══════════════════════════════════════════════════════════════════════════
# ZH: 四、初始點數發放 —— 冪等
# ══════════════════════════════════════════════════════════════════════════

class FakeAcc:
    def __init__(self, granted_at=None):
        self.credit_granted_at = granted_at
        self.credit_granted_pts = None
        self.credit_grant_note = None


class FakeDb:
    def commit(self):
        pass


def test_grant_is_skipped_when_already_granted(monkeypatch):
    """
    ZH: 🔴 發過就永不再發。自動開通是背景任務、SSO 可能重複觸發，
        只要重入一次就會發兩倍，而點數收不回來。
    """
    called = []

    async def boom(*a, **k):
        called.append(1)
        raise AssertionError("已發放過卻還是呼叫了廠商")

    monkeypatch.setattr(M, "transfer_credit_batch", boom)
    acc = FakeAcc(granted_at=datetime.now(timezone.utc))
    res = asyncio.run(M.grant_initial_credit(FakeDb(), acc, "a@example.com"))
    assert res["granted"] is False and res["reason"] == "already_granted"
    assert not called


def test_grant_is_skipped_when_setting_is_zero(monkeypatch):
    """ZH: myai_initial_credit = 0 就是不發（預設值，且必須真的不送出）。"""
    called = []

    async def boom(*a, **k):
        called.append(1)
        raise AssertionError("設定為 0 卻還是呼叫了廠商")

    monkeypatch.setattr(M, "transfer_credit_batch", boom)
    monkeypatch.setattr(M.crud, "get_setting", lambda db, k: 0)
    res = asyncio.run(M.grant_initial_credit(FakeDb(), FakeAcc(), "a@example.com"))
    assert res["granted"] is False and res["reason"] == "disabled"
    assert not called


def test_grant_records_the_points_on_success(monkeypatch):
    """ZH: 成功要留下紀錄（同時是下一次的冪等鍵）。"""
    async def ok(rows, confirm_grant=False):
        assert confirm_grant is True
        return {"ok": True, "granted": True, "count": 1, "points": 3}

    monkeypatch.setattr(M, "transfer_credit_batch", ok)
    monkeypatch.setattr(M.crud, "get_setting", lambda db, k: 3)
    acc = FakeAcc()
    res = asyncio.run(M.grant_initial_credit(FakeDb(), acc, "a@example.com"))
    assert res["granted"] is True
    assert acc.credit_granted_pts == 3
    assert acc.credit_granted_at is not None


def test_grant_marks_unknown_without_retrying(monkeypatch):
    """
    ZH: 🔴 第二段失敗時點數**可能已經轉出** —— 要標成 unknown 並照樣寫
        granted_at（擋住重入），而不是回失敗讓下一次再送一遍。
    """
    calls = []

    async def half(rows, confirm_grant=False):
        calls.append(1)
        return {"ok": False, "granted": False, "count": 1, "points": 2}

    monkeypatch.setattr(M, "transfer_credit_batch", half)
    monkeypatch.setattr(M.crud, "get_setting", lambda db, k: 2)
    acc = FakeAcc()
    res = asyncio.run(M.grant_initial_credit(FakeDb(), acc, "a@example.com"))
    assert res["reason"] == "unknown"
    assert acc.credit_granted_at is not None, "沒擋住重入 → 下次會再發一次"
    assert acc.credit_granted_pts == 0
    assert "對帳" in acc.credit_grant_note
    assert len(calls) == 1, "重試了"


# ══════════════════════════════════════════════════════════════════════════
# ZH: 五、開通通知信的開關（myai_provision_email）
# ══════════════════════════════════════════════════════════════════════════
# ZH: 這一組用真的 db fixture 跑完整條 provision_user —— 只把「對廠商送出」
#     和「真的寄信」這兩件事換掉。開關測試如果只驗設定讀得到，等於沒驗。

def _provision_env(monkeypatch, db, send_email_flag, sent):
    """ZH: 把 provision_user 需要的外部相依換掉，回傳建好的 user。

    ZH: 廠商呼叫一律換成「成功」，寄信換成記錄到 sent，其餘走真的程式碼。
    """
    from app.services import email_service

    async def fake_register(rows):
        return {"ok": True, "created": True, "status": 200,
                "rows": [{"emails": rows[0]["email"]}], "html": ""}

    async def fake_transfer(rows, confirm_grant=False):
        return {"ok": True, "granted": True, "count": 1, "points": 1}

    monkeypatch.setattr(M, "register_batch", fake_register)
    monkeypatch.setattr(M, "transfer_credit_batch", fake_transfer)
    monkeypatch.setattr(email_service, "send_myai_provisioned",
                        lambda *a, **k: sent.append(a[0]))

    real_get_setting = M.crud.get_setting

    def patched(_db, key):
        if key == "myai_autoprovision":
            return 1
        if key == "myai_provision_email":
            return send_email_flag
        if key == "myai_initial_credit":
            return 0        # ZH: 這組測的是寄信，點數另有測試
        return real_get_setting(_db, key)

    monkeypatch.setattr(M.crud, "get_setting", patched)

    from conftest import make_user
    return make_user(db, username="12360013", email="12360013@example.com")


def test_provision_email_is_sent_when_enabled(monkeypatch, db):
    """ZH: 陽性對照 —— 開著的時候真的會寄，否則下面那個測試等於沒測。"""
    sent = []
    user = _provision_env(monkeypatch, db, 1, sent)
    res = asyncio.run(M.provision_user(db, user))
    assert res["status"] == "created"
    assert sent == ["12360013@example.com"]


def test_provision_email_is_suppressed_when_disabled(monkeypatch, db):
    """ZH: 關掉就完全不寄 —— 而且開通本身仍要成功（關信不等於關功能）。"""
    sent = []
    user = _provision_env(monkeypatch, db, 0, sent)
    res = asyncio.run(M.provision_user(db, user))
    assert res["status"] == "created", "關掉通知信不應影響開通"
    assert sent == [], "設定為 0 卻還是寄了"
