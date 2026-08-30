"""
==============================================================================
Router: 外部 AI 分流路由 (External AI Routing) — v2.5
==============================================================================
ZH: 用途：自家 AI 助手成熟前，以合作廠商 (myai168) 暫時替代給非 admin 使用者。
    廠商無 API/SSO，僅能導流 + 帳號後台造冊，故：
      - 使用者端 GET /me：回傳廠商網址 + 該使用者被指派的廠商帳號名（導流用）。
      - 管理端 /admin/*：admin 維護「平台帳號 ↔ 廠商帳號」對應表 + 設定廠商網址。
    安全原則：只存廠商帳號名，絕不存廠商密碼。
EN: Purpose: Temporarily route non-admin users to a partner vendor (myai168)
    until the in-house AI matures. Vendor has no API/SSO — only redirect +
    back-office provisioning. User endpoint returns URL + assigned vendor
    username; admin endpoints manage the mapping table and the vendor URL.
    Security: store vendor username only, never the vendor password.
==============================================================================
"""

from fastapi import APIRouter, Depends, HTTPException
from pydantic import BaseModel
from sqlalchemy.orm import Session
from typing import Any, Optional
import csv
import io
import re

from .. import crud, schemas, models
from ..auth import get_current_user
from ..database import get_db
from ..services import myai_sync

router = APIRouter(tags=["外部 AI External-AI"])

EXTERNAL_AI_URL_KEY = "external_ai_url"
EXTERNAL_AI_LOGOUT_KEY = "external_ai_logout_url"
# ZH: 已實測的 myai 登出端點（GET 導向即清 session）| EN: verified myai logout endpoint
DEFAULT_MYAI_LOGOUT_URL = "https://www.myai168.com/mcu/ai/user/logout_info"

# ZH: v2.8 低點數提醒設定（存 SystemConfig，admin 可調）
# ZH: 🔴 v3.8 —— 設定鍵與預設值的**正本在 crud**,這裡只是別名。
#     寫入端（存設定）與讀取端（畫面／排程寄信）曾經各拿一份字串常數,
#     改其中一邊不會有任何錯誤,只會變成「存進 A 鍵、讀 B 鍵」→ 設定看起來沒生效。
MYAI_LOW_BALANCE_KEY = crud.MYAI_LOW_BALANCE_KEY
MYAI_APPLY_GUIDE_KEY = "myai_apply_guide_url"          # 申請教學連結（可空，之後再設定）
DEFAULT_LOW_BALANCE = crud.DEFAULT_LOW_BALANCE


class AlertConfig(BaseModel):
    low_balance_threshold: int | None = None
    apply_guide_url: str | None = None


def _low_balance_threshold(db: Session) -> int:
    """ZH: 門檻的唯一定義在 crud（排程寄信也要用同一份）,此處只是舊呼叫端的入口。

    @node job-scheduler/app/routers/external_ai.py::_low_balance_threshold
    """
    return crud.myai_low_balance_threshold(db)


def _current_myai_account(db: Session, current_user: models.User):
    """ZH: 取登入者自己的 myai 帳號列（綁定 sn 優先，退 email）；無則 None。
            ⚠ 一律由 JWT 的 current_user 推導，絕不接受前端傳來的身分參數
              —— 這是「使用者只能看自己」的唯一防線。
       EN: resolve the caller's OWN myai row from the JWT user only (never from
           client-supplied identity). This is the sole guard for self-only access.

    @node job-scheduler/app/routers/external_ai.py::_current_myai_account
    """
    acc = crud.get_external_account_by_user_id(db, current_user.id)
    row = None
    if acc:
        if acc.myai_vendor_sn:
            row = db.query(models.MyaiAccount).filter(
                models.MyaiAccount.vendor_sn == acc.myai_vendor_sn).first()
        if not row and acc.vendor_username:
            row = db.query(models.MyaiAccount).filter(
                models.MyaiAccount.email.ilike(acc.vendor_username)).first()
    if not row and current_user.email:
        row = db.query(models.MyaiAccount).filter(
            models.MyaiAccount.email.ilike(current_user.email)).first()
    return row


def _current_myai_points(db: Session, current_user: models.User):
    """ZH: 取登入者當前的 myai 剩餘點數（綁定 sn 優先，退 email）；無則 None。

    @node job-scheduler/app/routers/external_ai.py::_current_myai_points
    """
    row = _current_myai_account(db, current_user)
    return row.points if row else None


# ZH: 學生端對照的「最小樣本數」。人均 + 活躍人數會反推出個體：
#     若期間內只有 2 個活躍帳號（我＋另一人），總量＝人均×2，對方＝總量−我的 → 精準洩漏。
#     3 人以上就無法反推「特定個人」（只能得到其餘人的總和），這裡取 5 更保守。
#     樣本不足時不給對照（前端顯示說明），只顯示使用者自己的用量。
# EN: minimum cohort for showing the peer baseline to students. avg × active_accounts
#     lets a student solve for another individual when active==2. 3 is the mathematical
#     floor; 5 is the conservative small-cell suppression threshold used here.
MIN_PEER_COHORT = 5


def _peer_baseline(db: Session, since):
    """ZH: 同期間「全體基準」—— 只回聚合數字，**不含任何個人身分**（無姓名/email/sn 對外）。
            人均分母＝期間內真的有用量的帳號數（沒用的人不該稀釋平均）。
            管理端個人查詢與學生端「我的使用量」共用這裡，避免兩份人均邏輯各自漂移。
       EN: same-window aggregate baseline. Returns aggregates only — no per-person
           identity ever leaves this function. Shared by admin lookup and the
           student-facing endpoint so the two can't drift apart.

    @node job-scheduler/app/routers/external_ai.py::_peer_baseline
    """
    q = db.query(models.MyaiTransaction).filter(
        models.MyaiTransaction.event_type == "ai_usage",
        models.MyaiTransaction.occurred_at.isnot(None),
    )
    if since:
        q = q.filter(models.MyaiTransaction.occurred_at >= since)
    daily: dict = {}
    per_sn: dict = {}
    model_points: dict = {}
    total = uses = 0
    for t in q.all():
        c = -(t.points_delta or 0)
        if c <= 0:
            continue
        total += c
        uses += 1
        per_sn[t.vendor_sn] = per_sn.get(t.vendor_sn, 0) + c
        code = t.model or "unknown"
        model_points[code] = model_points.get(code, 0) + c
        d = t.occurred_at.date().isoformat()
        daily[d] = daily.get(d, 0) + c
    return {"daily": daily, "per_sn": per_sn, "model_points": model_points,
            "total": total, "uses": uses, "active": len(per_sn)}


def _own_usage(db: Session, sns: set, since, mmap: dict):
    """ZH: 某組 vendor_sn 的自身用量（消耗/次數/登入、模型別、每日）。
       EN: own usage for a set of vendor_sn: totals, per-model, per-day.

    @node job-scheduler/app/routers/external_ai.py::_own_usage
    """
    tq = db.query(models.MyaiTransaction).filter(models.MyaiTransaction.vendor_sn.in_(sns))
    if since:
        tq = tq.filter(models.MyaiTransaction.occurred_at >= since)
    txs = tq.order_by(models.MyaiTransaction.occurred_at.desc()).all()
    consumed = uses = logins = 0
    model_agg: dict = {}
    daily: dict = {}
    for t in txs:
        if t.event_type == "ai_usage":
            c = -(t.points_delta or 0)
            if c > 0:
                consumed += c
                uses += 1
                code = t.model or "unknown"
                e = mmap.get(code)
                mm = model_agg.setdefault(code, {
                    "model": code,
                    "display_name": (e.display_name if e and e.display_name else code),
                    "provider": (e.provider if e and e.provider else "未對應"),
                    "category": (e.category if e and e.category else "未對應"),
                    "mapped": bool(e), "count": 0, "points": 0,
                })
                mm["count"] += 1
                mm["points"] += c
                if t.occurred_at:
                    d = t.occurred_at.date().isoformat()
                    daily[d] = daily.get(d, 0) + c
        elif t.event_type == "login":
            logins += 1
    return {"txs": txs, "consumed": consumed, "uses": uses, "logins": logins,
            "model_agg": model_agg, "daily": daily}


def _aligned_series(daily_me: dict, peer: dict):
    """ZH: 趨勢軸用「全體有活動的日子」，自己沒用的日子補 0
            —— 這樣才看得出自己在整體節奏中的位置，而不是只看到自己的孤島。
       EN: align own series onto the all-accounts activity axis; zero-fill own gaps.

    @node job-scheduler/app/routers/external_ai.py::_aligned_series
    """
    active = peer["active"] or 1
    return [{
        "date": d,
        "consumed": daily_me.get(d, 0),
        "peer_avg": round(peer["daily"][d] / active, 1),
    } for d in sorted(peer["daily"].keys())]


def _ranked_models(model_agg: dict, consumed: int, peer: dict):
    """ZH: 模型別排序 + 佔比。share=自己佔比、peer_share=全體佔比
            → 並排比得出「我特別吃哪種模型」（絕對點數量級差太多，比不動）。
       EN: per-model list with own vs all-accounts share (%), sorted by points.

    @node job-scheduler/app/routers/external_ai.py::_ranked_models
    """
    rows = sorted(model_agg.values(), key=lambda x: x["points"], reverse=True)
    total_all = peer["total"]
    for m in rows:
        m["share"] = round(100 * m["points"] / consumed, 1) if consumed else 0
        m["peer_share"] = (round(100 * peer["model_points"].get(m["model"], 0) / total_all, 1)
                           if total_all else 0)
    return rows


# ZH: v3.8 這裡原本是**全站第三份** require_admin 的複製實作,
#     三份都寫著 `role != "admin"` —— 身分與權限拆開時要改三個地方才算改完。
#     改成重新匯出共用那支（auth.require_admin，看 is_admin 旗標）。
from ..auth import require_admin  # noqa: E402  ZH: 位置貼著原本的定義,讓 diff 看得出取代關係


# ==============================================================================
# ZH: 使用者端 | EN: User-facing
# ==============================================================================


@router.post("/visit", status_code=204)
def record_visit(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
) -> None:
    """ZH: 記一次「從平台跳去 MYAI」| EN: Record one platform → MYAI redirect

    ZH: 這支存在的理由：`goMyai()` 只是開一個新分頁，在這之前**不留任何痕跡**，
        所以「哪個系在用 MYAI」查不到入口流量這一面。

    ZH: 🔴 **不存 IP**。要回答的是「哪個系在用」，而那由 user_id 推得出來
        （users.department → org_departments）。存 IP 對這個問題沒有貢獻，
        只是多留一份可以反推位置的資料。與問題回報同一條原則。

    ZH: ⚠ 記不起來**不要讓使用者的動作失敗**。前端是「先記、再開分頁」，
        這支回 500 的話他就去不了 MYAI 了 —— 統計比不上那件事。
        所以吞掉例外，回 204。

    @node job-scheduler/app/routers/external_ai.py::record_visit
    """
    try:
        db.add(models.MyaiVisit(user_id=current_user.id))
        db.commit()
    except Exception:                                     # pragma: no cover
        db.rollback()
        logger.warning("ZH: MYAI 跳轉紀錄寫入失敗（不影響使用者）")

@router.get("/me", response_model=schemas.ExternalAiMe)
def get_my_external_ai(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
) -> Any:
    """ZH: 取得自己的外部 AI 導流資訊（網址 + 指派帳號 + 狀態）
       EN: Get my external-AI redirect info (url + assigned account + status)

    @node job-scheduler/app/routers/external_ai.py::get_my_external_ai
    """
    url = crud.get_system_config(db, EXTERNAL_AI_URL_KEY, "")

    acc = crud.get_external_account_by_user_id(db, current_user.id)
    if not acc:
        vendor, status = None, "not_provisioned"
    elif (acc.status or "active") != "active":
        vendor, status = acc.vendor_username, "disabled"
    else:
        vendor, status = acc.vendor_username, "active"

    # v2.8 廠商 Token 餘額：顯式串接 綁定 → myai_accounts（穩定 sn 優先，退而 email）。
    # 尚無綁定者退回直接以 email 比對，確保 auto-match 跑之前也不會空白。
    myai_points = myai_expiry = myai_status = None
    myai_row = None
    if acc:
        if acc.myai_vendor_sn:
            myai_row = (
                db.query(models.MyaiAccount)
                .filter(models.MyaiAccount.vendor_sn == acc.myai_vendor_sn)
                .first()
            )
        if not myai_row and acc.vendor_username:
            myai_row = (
                db.query(models.MyaiAccount)
                .filter(models.MyaiAccount.email.ilike(acc.vendor_username))
                .first()
            )
    if not myai_row and current_user.email:
        myai_row = (
            db.query(models.MyaiAccount)
            .filter(models.MyaiAccount.email.ilike(current_user.email))
            .first()
        )
    if myai_row:
        myai_points, myai_expiry, myai_status = myai_row.points, myai_row.expiry, myai_row.status

    logout_url = crud.get_system_config(db, EXTERNAL_AI_LOGOUT_KEY, DEFAULT_MYAI_LOGOUT_URL)
    return schemas.ExternalAiMe(
        url=url, vendor_username=vendor, status=status,
        myai_points=myai_points, myai_expiry=myai_expiry, myai_status=myai_status,
        logout_url=(logout_url or None),
    )


# ==============================================================================
# ZH: v3.3 自動開通 —— 學生端查詢自己的 MYAI 帳號狀態與初始密碼
# EN: v3.3 auto-provision — per-user account status & initial password
# ==============================================================================
@router.get("/my-provision")
def get_my_provision(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
) -> Any:
    """
    ZH: 回自己的 MYAI 開通狀態。初始密碼**只在保留期內且未確認修改**時才回傳，
        且身分一律由 JWT 推導（不吃任何身分參數）→ 查不到別人的。
    EN: Own provisioning status; initial password only within retention & unacknowledged.

    @node job-scheduler/app/routers/external_ai.py::get_my_provision
    """
    from ..services import myai_sync
    return myai_sync.provision_status(db, current_user)


@router.post("/my-provision/ack")
def ack_my_provision(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
) -> Any:
    """ZH: 學生按「我已修改密碼」→ 立即銷毀暫存的初始密碼（不等保留期到）。

    @node job-scheduler/app/routers/external_ai.py::ack_my_provision
    """
    from ..services import myai_sync
    ok = myai_sync.acknowledge_initial_password(db, current_user)
    if not ok:
        raise HTTPException(status_code=404, detail="尚無開通紀錄")
    return {"message": "已清除暫存的初始密碼"}


@router.get("/my-balance")
def get_my_balance(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
) -> Any:
    """ZH: 登入者的外部 AI 剩餘點數 + 是否低於門檻（供前端低點數彈窗判斷）。
       EN: current user's remaining external-AI points + low-balance flag.

    @node job-scheduler/app/routers/external_ai.py::get_my_balance
    """
    points = _current_myai_points(db, current_user)
    threshold = _low_balance_threshold(db)
    guide = crud.get_system_config(db, MYAI_APPLY_GUIDE_KEY, "")
    return {
        "points": points,
        "threshold": threshold,
        # ZH: v3.8 #9 —— 兩段式。`below` 分不出「快用完」與「已經用完」,
        #     而那是兩件不同的事：前者要提醒他去申請,後者是他現在就用不了。
        #     判定集中在 crud.myai_balance_state,寄信與畫面共用同一份規則 ——
        #     兩邊各判一次的話,信裡說「已用完」而畫面說「偏低」是遲早的事。
        "state": crud.myai_balance_state(points, threshold),
        "below": (points is not None and points < threshold),   # ZH: 保留給既有呼叫端
        "apply_guide_url": (guide or None),
    }


@router.get("/my-consumption")
def get_my_consumption(
    days: int = 30,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
) -> Any:
    """ZH: v3.0 學生端「我的使用量」—— 只給**自己的**用量 + **全體聚合**趨勢對照。
       EN: v3.0 student-facing usage: own usage only, plus aggregate-only baseline.

    ⚠ 隱私邊界（刻意設計，不要放寬）：
      1. 身分只從 JWT 的 current_user 推導，**不吃任何查詢身分的參數** → 無法查別人。
      2. 回傳裡**沒有**其他人的姓名/email/序號、沒有 Top 消耗者、沒有逐帳號清單。
      3. **不給排名**：排名會變成比賽/公審，且隱含他人相對位置。全體只以「人均」出現。
      4. 對照數字全是聚合值（人均、佔比），單一使用者無法從中反推特定他人。

    @node job-scheduler/app/routers/external_ai.py::get_my_consumption
    """
    from datetime import datetime, timedelta
    row = _current_myai_account(db, current_user)
    if not row:
        # ZH: 沒綁定廠商帳號（或還沒同步到）→ 前端顯示友善說明，不是錯誤
        return {"bound": False, "days": (30 if days is None else int(days)), "summary": {}, "series": [],
                "models": [], "peer": {}, "account": {}}

    # ZH: 不能用 `days or 30` —— 0 在 Python 是 falsy，前端「全部」送的正是 0，
    #     會被悄悄換成 30，導致「全部」實際只看近 30 天（同頁的「個人查詢」
    #     用 `days or 0` 反而是對的，兩個面板同一個詞卻不同行為）。
    days = 30 if days is None else int(days)
    since = datetime.now() - timedelta(days=min(days, 3650)) if days > 0 else None
    mmap = {m.code: m for m in db.query(models.MyaiModelMap).all()}
    own = _own_usage(db, {row.vendor_sn}, since, mmap)
    peer = _peer_baseline(db, since)
    active = peer["active"]
    # ZH: 5.「樣本太少就不給對照」—— 見 MIN_PEER_COHORT：人均會反推出特定個人。
    #     不足時仍完整顯示「自己的」用量，只是拿掉全體對照。
    show_peer = active >= MIN_PEER_COHORT
    series = _aligned_series(own["daily"], peer) if show_peer else [
        {"date": d, "consumed": own["daily"][d]} for d in sorted(own["daily"].keys())
    ]
    model_list = _ranked_models(own["model_agg"], own["consumed"], peer)
    if not show_peer:
        for m in model_list:
            m.pop("peer_share", None)     # 全體佔比同樣是對照資料，一併拿掉
    return {
        "bound": True,
        "days": days,
        # ZH: 只有自己的帳號資訊（餘額/有效期），這本來就是他自己的
        "account": {"points": row.points, "expiry": row.expiry},
        "summary": {"consumed": own["consumed"], "uses": own["uses"], "logins": own["logins"]},
        "series": series,
        "models": model_list,
        # ZH: 聚合值 only —— 刻意沒有 rank、沒有逐帳號資訊；樣本不足時連人均都不給。
        "peer": {
            "show": show_peer,
            "active_accounts": active,
            "min_cohort": MIN_PEER_COHORT,
            **({"avg_consumed": round(peer["total"] / (active or 1), 1),
                "avg_uses": round(peer["uses"] / (active or 1), 1)} if show_peer else {}),
        },
    }


# ==============================================================================
# ZH: 管理端 — 廠商網址設定 | EN: Admin — vendor URL setting
# ==============================================================================

@router.get("/admin/url", response_model=schemas.ExternalAiUrl)
def get_external_url(
    db: Session = Depends(get_db),
    _: models.User = Depends(require_admin),
) -> Any:
    """@node job-scheduler/app/routers/external_ai.py::get_external_url"""
    return schemas.ExternalAiUrl(
        url=crud.get_system_config(db, EXTERNAL_AI_URL_KEY, ""),
        logout_url=crud.get_system_config(db, EXTERNAL_AI_LOGOUT_KEY, DEFAULT_MYAI_LOGOUT_URL),
    )


@router.put("/admin/url", response_model=schemas.ExternalAiUrl)
def set_external_url(
    payload: schemas.ExternalAiUrl,
    db: Session = Depends(get_db),
    _: models.User = Depends(require_admin),
) -> Any:
    """@node job-scheduler/app/routers/external_ai.py::set_external_url"""
    crud.set_system_config(
        db, EXTERNAL_AI_URL_KEY, payload.url.strip(),
        description="外部 AI 平台網址（空=未啟用，退回即將開放）",
    )
    if payload.logout_url is not None:
        crud.set_system_config(
            db, EXTERNAL_AI_LOGOUT_KEY, payload.logout_url.strip(),
            description="外部 AI 廠商登出網址（共用機台『結束使用』會開它殺掉廠商 session）",
        )
    return schemas.ExternalAiUrl(
        url=crud.get_system_config(db, EXTERNAL_AI_URL_KEY, ""),
        logout_url=crud.get_system_config(db, EXTERNAL_AI_LOGOUT_KEY, DEFAULT_MYAI_LOGOUT_URL),
    )


@router.get("/admin/alert-config")
def get_alert_config(
    db: Session = Depends(get_db),
    _: models.User = Depends(require_admin),
) -> Any:
    """ZH: 低點數提醒設定（門檻 + 申請教學連結）。

    @node job-scheduler/app/routers/external_ai.py::get_alert_config
    """
    return {
        "low_balance_threshold": _low_balance_threshold(db),
        "apply_guide_url": crud.get_system_config(db, MYAI_APPLY_GUIDE_KEY, ""),
    }


@router.put("/admin/alert-config")
def set_alert_config(
    payload: AlertConfig,
    db: Session = Depends(get_db),
    _: models.User = Depends(require_admin),
) -> Any:
    """@node job-scheduler/app/routers/external_ai.py::set_alert_config"""
    if payload.low_balance_threshold is not None:
        crud.set_system_config(
            db, MYAI_LOW_BALANCE_KEY, str(max(0, int(payload.low_balance_threshold))),
            description="外部 AI 低點數提醒門檻（低於此絕對點數 → 提醒學生）",
        )
    if payload.apply_guide_url is not None:
        crud.set_system_config(
            db, MYAI_APPLY_GUIDE_KEY, payload.apply_guide_url.strip(),
            description="外部 AI 點數申請教學連結（顯示在低點數彈窗）",
        )
    return {
        "low_balance_threshold": _low_balance_threshold(db),
        "apply_guide_url": crud.get_system_config(db, MYAI_APPLY_GUIDE_KEY, ""),
    }


# ==============================================================================
# ZH: 管理端 — 對應表 CRUD | EN: Admin — mapping CRUD
# ==============================================================================

@router.get("/admin/accounts", response_model=list[schemas.ExternalAiAccountResponse])
def list_accounts(
    db: Session = Depends(get_db),
    _: models.User = Depends(require_admin),
) -> Any:
    """@node job-scheduler/app/routers/external_ai.py::list_accounts"""
    return crud.list_external_accounts(db)


@router.post("/admin/accounts", response_model=schemas.ExternalAiAccountResponse)
def create_account(
    payload: schemas.ExternalAiAccountCreate,
    db: Session = Depends(get_db),
    _: models.User = Depends(require_admin),
) -> Any:
    """@node job-scheduler/app/routers/external_ai.py::create_account"""
    try:
        acc = crud.create_external_account(
            db, payload.platform_username.strip(), payload.vendor_username.strip(),
            payload.status or "active", payload.note,
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    return schemas.ExternalAiAccountResponse(
        id=acc.id, user_id=acc.user_id, platform_username=payload.platform_username.strip(),
        vendor_username=acc.vendor_username, status=acc.status, note=acc.note,
        updated_at=acc.updated_at,
    )


@router.put("/admin/accounts/{account_id}", response_model=schemas.ExternalAiAccountResponse)
def update_account(
    account_id: str,
    payload: schemas.ExternalAiAccountUpdate,
    db: Session = Depends(get_db),
    _: models.User = Depends(require_admin),
) -> Any:
    """@node job-scheduler/app/routers/external_ai.py::update_account"""
    acc = crud.update_external_account(
        db, account_id,
        vendor_username=(payload.vendor_username.strip() if payload.vendor_username else None),
        status=payload.status, note=payload.note,
    )
    if not acc:
        raise HTTPException(status_code=404, detail="ZH: 找不到這筆對應設定 | EN: mapping not found")
    user = db.query(models.User).filter(models.User.id == acc.user_id).first()
    return schemas.ExternalAiAccountResponse(
        id=acc.id, user_id=acc.user_id,
        platform_username=(user.username if user else None),
        vendor_username=acc.vendor_username, status=acc.status, note=acc.note,
        updated_at=acc.updated_at,
    )


@router.delete("/admin/accounts/{account_id}")
def delete_account(
    account_id: str,
    db: Session = Depends(get_db),
    _: models.User = Depends(require_admin),
) -> Any:
    """@node job-scheduler/app/routers/external_ai.py::delete_account"""
    if not crud.delete_external_account(db, account_id):
        raise HTTPException(status_code=404, detail="ZH: 找不到這筆對應設定 | EN: mapping not found")
    return {"ok": True}


# ==============================================================================
# ZH: 管理端 — CSV 批次匯入造冊結果 | EN: Admin — CSV bulk import
# ==============================================================================

@router.post("/admin/import", response_model=schemas.ExternalAiImportResult)
def import_accounts_csv(
    payload: dict,
    db: Session = Depends(get_db),
    _: models.User = Depends(require_admin),
) -> Any:
    """ZH: 接收 CSV 文字 (欄位: platform_username,vendor_username)，逐行 upsert。
       EN: Accept CSV text (cols: platform_username,vendor_username), upsert per row.

    @node job-scheduler/app/routers/external_ai.py::import_accounts_csv
    """
    text = (payload or {}).get("csv", "")
    result = schemas.ExternalAiImportResult()
    if not text or not text.strip():
        result.errors.append("empty CSV")
        return result
    reader = csv.reader(io.StringIO(text))
    for idx, row in enumerate(reader, start=1):
        if not row or all(not c.strip() for c in row):
            continue
        # 跳過表頭 | skip header
        if idx == 1 and row[0].strip().lower() in ("platform_username", "username", "平台帳號"):
            continue
        if len(row) < 2 or not row[0].strip() or not row[1].strip():
            result.errors.append(f"第 {idx} 行格式錯誤 (需 platform_username,vendor_username)")
            continue
        try:
            outcome = crud.upsert_external_account_by_username(db, row[0].strip(), row[1].strip())
            if outcome == "created":
                result.created += 1
            elif outcome == "updated":
                result.updated += 1
            else:
                result.skipped += 1
        except ValueError as e:
            result.errors.append(f"第 {idx} 行: {e}")
    return result


# ==============================================================================
# ZH: v2.8 MYAI 廠商平台 headless 同步（唯讀）| EN: v2.8 MYAI headless sync (read-only)
# ==============================================================================

@router.post("/admin/sync-myai")
async def sync_myai(
    db: Session = Depends(get_db),
    _: models.User = Depends(require_admin),
) -> Any:
    """ZH: 立即 headless 登入廠商 → 匯出使用者(含 Token 點數) → 同步進 myai_accounts。
       EN: Trigger headless login → export → upsert into myai_accounts.

    @node job-scheduler/app/routers/external_ai.py::sync_myai
    """
    try:
        return await myai_sync.sync(db)
    except myai_sync.MyaiSyncError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:  # noqa: BLE001
        raise HTTPException(status_code=500, detail=f"同步失敗：{e}")


@router.get("/admin/myai-accounts")
def list_myai_accounts(
    db: Session = Depends(get_db),
    _: models.User = Depends(require_admin),
) -> Any:
    """ZH: 列出已同步的廠商帳號/Token（供 admin 顯示）| EN: list synced MYAI accounts.

    @node job-scheduler/app/routers/external_ai.py::list_myai_accounts
    """
    rows = (
        db.query(models.MyaiAccount)
        .order_by(models.MyaiAccount.points.desc())
        .all()
    )
    last = max((r.synced_at for r in rows), default=None)
    return {
        "synced_at": last.isoformat() if last else None,
        "count": len(rows),
        "accounts": [
            {
                "vendor_sn": r.vendor_sn, "email": r.email, "name": r.name,
                "user_type": r.user_type, "points": r.points, "expiry": r.expiry,
                "status": r.status, "newsletter": r.newsletter, "note": r.note,
            }
            for r in rows
        ],
    }


# ==============================================================================
# ZH: v2.8 MYAI email 綁定管理（自動配對 / 綁定清單 / 未配對）— 只寫本平台 DB
# EN: v2.8 MYAI email-binding management (auto-match / bindings / unmatched)
# ==============================================================================

@router.post("/admin/auto-match")
def auto_match_bindings(
    db: Session = Depends(get_db),
    _: models.User = Depends(require_admin),
) -> Any:
    """ZH: 以 email 自動配對 myai 帳號 ↔ 平台使用者，建立/回填綁定（不碰廠商）。
       EN: Auto-bind myai accounts to platform users by email (our DB only).

    @node job-scheduler/app/routers/external_ai.py::auto_match_bindings
    """
    return myai_sync.auto_match(db)


@router.get("/admin/bindings")
def list_bindings(
    db: Session = Depends(get_db),
    _: models.User = Depends(require_admin),
) -> Any:
    """ZH: 綁定清單：平台帳號 ↔ myai email ↔ 點數/狀態（join 同步快取）。
       EN: Binding list: platform user ↔ myai email ↔ points/status.

    @node job-scheduler/app/routers/external_ai.py::list_bindings
    """
    rows = (
        db.query(models.ExternalAiAccount, models.User.username, models.User.email)
        .join(models.User, models.User.id == models.ExternalAiAccount.user_id)
        .order_by(models.User.username.asc())
        .all()
    )
    # ZH: 預載 myai 快取，避免逐筆查 | preload myai cache
    myai_by_sn = {m.vendor_sn: m for m in db.query(models.MyaiAccount).all()}
    myai_by_email = {(m.email or "").strip().lower(): m for m in myai_by_sn.values() if m.email}
    out = []
    for acc, username, user_email in rows:
        m = None
        if acc.myai_vendor_sn:
            m = myai_by_sn.get(acc.myai_vendor_sn)
        if not m and acc.vendor_username:
            m = myai_by_email.get(acc.vendor_username.strip().lower())
        out.append({
            "id": acc.id,
            "user_id": acc.user_id,
            "platform_username": username,
            "platform_email": user_email,
            "myai_email": acc.vendor_username,
            "myai_vendor_sn": acc.myai_vendor_sn,
            "status": acc.status,
            "note": acc.note,
            "points": (m.points if m else None),
            "myai_status": (m.status if m else None),
            "synced": bool(m),  # ZH: 是否對得上同步快取 | matched to sync cache
            "updated_at": acc.updated_at.isoformat() if acc.updated_at else None,
        })
    return {"count": len(out), "bindings": out}


@router.get("/admin/unmatched")
def list_unmatched(
    db: Session = Depends(get_db),
    _: models.User = Depends(require_admin),
) -> Any:
    """ZH: 兩邊未配對：① 平台未綁定使用者(標註是否有同 email 的 myai 帳號)
            ② myai 帳號未被任何綁定指向(標註是否有同 email 的平台使用者)。
       EN: Unmatched on both sides for admin follow-up.

    @node job-scheduler/app/routers/external_ai.py::list_unmatched
    """
    accs = db.query(models.ExternalAiAccount).all()
    bound_user_ids = {a.user_id for a in accs}
    bound_sns = {a.myai_vendor_sn for a in accs if a.myai_vendor_sn}
    bound_emails = {(a.vendor_username or "").strip().lower() for a in accs if a.vendor_username}

    myai_rows = db.query(models.MyaiAccount).all()
    myai_emails = {(m.email or "").strip().lower() for m in myai_rows if m.email}

    # ① 平台未綁定使用者 | platform users without a binding
    users = db.query(models.User).all()
    unmatched_users = [
        {
            "user_id": u.id, "username": u.username, "email": u.email,
            "has_myai_match": bool(u.email and u.email.strip().lower() in myai_emails),
        }
        for u in users if u.id not in bound_user_ids
    ]

    # ② myai 帳號未被綁定指向 | myai accounts not targeted by any binding
    platform_emails = {(u.email or "").strip().lower() for u in users if u.email}
    unmatched_myai = [
        {
            "vendor_sn": m.vendor_sn, "email": m.email, "name": m.name,
            "user_type": m.user_type, "points": m.points, "status": m.status,
            "has_platform_user": bool(m.email and m.email.strip().lower() in platform_emails),
        }
        for m in myai_rows
        if m.vendor_sn not in bound_sns
        and (m.email or "").strip().lower() not in bound_emails
    ]
    return {
        "unmatched_users": unmatched_users,
        "unmatched_myai": unmatched_myai,
        "unmatched_user_count": len(unmatched_users),
        "unmatched_myai_count": len(unmatched_myai),
    }


@router.post("/admin/sync-transactions")
async def sync_transactions_endpoint(
    days: int = 90,
    db: Session = Depends(get_db),
    _: models.User = Depends(require_admin),
) -> Any:
    """ZH: 立即同步廠商交易日誌（逐筆、含模型；不存 IP）| EN: sync tx log now.

    @node job-scheduler/app/routers/external_ai.py::sync_transactions_endpoint
    """
    try:
        return await myai_sync.sync_transactions(db, days=days)
    except myai_sync.MyaiSyncError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:  # noqa: BLE001
        raise HTTPException(status_code=500, detail=f"交易同步失敗：{e}")


@router.get("/admin/provision-candidates")
def admin_provision_candidates(
    db: Session = Depends(get_db),
    _admin: models.User = Depends(require_admin),
) -> Any:
    """
    ZH: v3.4 待開通清單。依**事實**分類，不做信心度預測：
          ready         有 email → 可自動開通（信箱真假不預判，寄出後看退件紀錄）
          no_email      完全沒有 email → 無從建號，需人工補
          staff_pending 信箱網域屬教職員域但角色仍是學生 → 只提示，不自動升權
    EN: Unbound SSO users split by a fact (has an address or not), never by confidence.

    @node job-scheduler/app/routers/external_ai.py::admin_provision_candidates
    """
    from ..services import myai_sync
    out = myai_sync.provision_candidates(db)
    out["staff_pending"] = myai_sync.staff_pending(db)
    return out


@router.get("/admin/live-usage")
def admin_live_usage(
    db: Session = Depends(get_db),
    _admin: models.User = Depends(require_admin),
) -> Any:
    """
    ZH: v3.4 MYAI 即時使用四象限（監控 + 稽核）。
        時間窗由系統設定 myai_usage_window_min 控制；⚠ 資料新鮮度受輪詢限制，
        回傳的 last_tx_sync 即上次與廠商同步的時間，前端須顯示以免誤解為即時。
    EN: v3.4 live usage quadrants (monitoring + audit).

    @node job-scheduler/app/routers/external_ai.py::admin_live_usage
    """
    from ..services import myai_sync
    return myai_sync.live_usage_quadrants(
        db, usage_minutes=crud.get_setting(db, "myai_usage_window_min")
    )


@router.get("/admin/consumption")
def consumption_analytics(
    days: int = 30,
    start: Optional[str] = None,
    end: Optional[str] = None,
    db: Session = Depends(get_db),
    _: models.User = Depends(require_admin),
) -> Any:
    """ZH: 消耗分析 —— 以廠商「交易日誌」逐筆(含模型)計算：期間總消耗、每生消耗、
            Top 消耗者、模型/工具別用量、每日趨勢、登入數。
       EN: Consumption analytics from the per-event transaction log (real model).

    @node job-scheduler/app/routers/external_ai.py::consumption_analytics
    """
    from datetime import datetime, timedelta, time as dtime
    # ZH: 不能用 `days or 30` —— 0 在 Python 是 falsy，前端「全部」送的正是 0，
    #     會被悄悄換成 30，導致「全部」實際只看近 30 天（同頁的「個人查詢」
    #     用 `days or 0` 反而是對的，兩個面板同一個詞卻不同行為）。
    #     days<=0 代表「全部」(不設下界，即從最早一筆起)
    days = 30 if days is None else int(days)
    q = db.query(models.MyaiTransaction).filter(models.MyaiTransaction.occurred_at.isnot(None))

    # ==================================================================
    # ZH: v3.8 起訖日期。理由：想看「整個 7 月」的話，
    #     「近 N 天」只有在 8/1 當天操作才對得上。
    #
    # ZH: 🔴 `occurred_at` 是**廠商的當地時間（naive）**，不是 UTC。
    #     所以這裡直接用 naive datetime 比對，**絕對不要轉成 UTC** ——
    #     轉了會整整差八小時，而那種錯位在月初月末最明顯：
    #     7/1 凌晨的交易會被算進 6 月。（見本檔上方同一個注記。）
    #
    # ZH: 🔴 `end` 是**含當天**：選 7/31 意思是「包含 7/31 一整天」。
    #     用 `<= end 00:00` 的話會掉掉最後一天的所有資料，
    #     而使用者看不出來（只會覺得數字「好像少一點」）。
    #
    # ZH: 起訖日期**優先於 days** —— 兩者都給時以起訖為準。
    # ==================================================================
    def _day(v):
        """ZH: 'YYYY-MM-DD' → date；格式不對就回 None（當作沒給）。"""
        try:
            return datetime.strptime((v or "").strip(), "%Y-%m-%d").date()
        except (ValueError, TypeError):
            return None

    d_start, d_end = _day(start), _day(end)
    if d_start and d_end and d_start > d_end:
        d_start, d_end = d_end, d_start     # ZH: 選反了就幫他換過來，不要回一張空表

    if d_start or d_end:
        if d_start:
            q = q.filter(models.MyaiTransaction.occurred_at
                         >= datetime.combine(d_start, dtime.min))
        if d_end:
            q = q.filter(models.MyaiTransaction.occurred_at
                         <= datetime.combine(d_end, dtime.max))
    elif days > 0:
        days = min(days, 3650)
        since = datetime.now() - timedelta(days=days)   # ZH: occurred_at 為廠商當地時間(naive)
        q = q.filter(models.MyaiTransaction.occurred_at >= since)
    txs = q.all()
    # ZH: email → 平台身分(role) / 學系(department)，用來分群 | email → role / department
    role_map: dict = {}
    dept_map: dict = {}
    # ZH: v3.8 #13 —— 也依**學院**與**行政單位**分群。
    #     學院不是 users 的欄位,由 department 經 org_departments 推 ——
    #     查一次做成 dict,不要在下面的交易迴圈裡逐筆查（那是幾萬筆）。
    college_of_dept = {d.name: d.college
                       for d in db.query(models.OrgDepartment).all()}
    college_map: dict = {}
    unit_map: dict = {}
    for u in db.query(models.User).all():
        if u.email:
            k = u.email.strip().lower()
            role_map[k] = u.role or "unknown"
            dept_map[k] = u.department or None
            # ZH: 對不到對照表就回 None（舊系名、打錯字）—— **不猜**,
            #     下面會歸到「未設定」而不是硬塞一個學院。
            college_map[k] = college_of_dept.get((u.department or "").strip()) or None
            unit_map[k] = getattr(u, "unit", None) or None
    # ZH: v2.9 模型對應表（顯示時套用，不改寫原始交易）| EN: display-time model map
    mmap = {m.code: m for m in db.query(models.MyaiModelMap).all()}
    per: dict = {}      # ZH: sn → 每生統計
    model_agg: dict = {}
    cat_agg: dict = {}
    prov_agg: dict = {}
    role_agg: dict = {}
    dept_agg: dict = {}
    college_agg: dict = {}
    unit_agg: dict = {}
    daily: dict = {}
    total = total_uses = total_logins = 0
    for t in txs:
        p = per.setdefault(t.vendor_sn, {
            "vendor_sn": t.vendor_sn, "name": t.name, "email": t.email,
            "consumed": 0, "uses": 0, "logins": 0,
        })
        if t.event_type == "ai_usage":
            c = -(t.points_delta or 0)
            if c > 0:
                p["consumed"] += c; p["uses"] += 1
                total += c; total_uses += 1
                m = t.model or "unknown"
                e = mmap.get(m)
                # ZH: 沒對應到就退回原始代碼並標 mapped=false（前端可提示去補對應表）
                mm = model_agg.setdefault(m, {
                    "model": m,
                    "display_name": (e.display_name if e and e.display_name else m),
                    "provider": (e.provider if e and e.provider else "未對應"),
                    "category": (e.category if e and e.category else "未對應"),
                    "mapped": bool(e), "count": 0, "points": 0,
                })
                mm["count"] += 1; mm["points"] += c
                ca = cat_agg.setdefault(mm["category"], {"category": mm["category"], "consumed": 0, "uses": 0})
                ca["consumed"] += c; ca["uses"] += 1
                pa = prov_agg.setdefault(mm["provider"], {"provider": mm["provider"], "consumed": 0, "uses": 0})
                pa["consumed"] += c; pa["uses"] += 1
                d = t.occurred_at.date().isoformat() if t.occurred_at else "unknown"
                daily[d] = daily.get(d, 0) + c
                ek = (t.email or "").strip().lower()
                # 師生用量：email → 平台 role；對不到 = 未綁定
                role = role_map.get(ek, "unbound")
                ra = role_agg.setdefault(role, {"role": role, "consumed": 0, "uses": 0})
                ra["consumed"] += c; ra["uses"] += 1
                # 學系用量：email → 平台 department；無則「未綁定」
                # ZH: 三個維度的「對不到」分兩種,不要混在一起：
                #       未綁定 —— 這個 email 在平台上根本沒有帳號
                #       未設定 —— 有帳號,但那個欄位是空的（或對不到對照表）
                #     混成一種的話,「廠商那邊有人我們不認識」與
                #     「我們的人資料沒填完」會長得一樣,而那是兩件要做的事。
                known = ek in dept_map
                dept = dept_map.get(ek) or ("未設定" if known else "未綁定")
                da = dept_agg.setdefault(dept, {"department": dept, "consumed": 0, "uses": 0})
                da["consumed"] += c; da["uses"] += 1
                college = college_map.get(ek) or ("未設定" if known else "未綁定")
                ca = college_agg.setdefault(college, {"college": college, "consumed": 0, "uses": 0})
                ca["consumed"] += c; ca["uses"] += 1
                unit = unit_map.get(ek) or ("未設定" if known else "未綁定")
                ua = unit_agg.setdefault(unit, {"unit": unit, "consumed": 0, "uses": 0})
                ua["consumed"] += c; ua["uses"] += 1
        elif t.event_type == "login":
            p["logins"] += 1; total_logins += 1
    accounts = sorted(per.values(), key=lambda x: x["consumed"], reverse=True)
    model_list = sorted(model_agg.values(), key=lambda x: x["points"], reverse=True)
    by_category = sorted(cat_agg.values(), key=lambda x: x["consumed"], reverse=True)
    by_provider = sorted(prov_agg.values(), key=lambda x: x["consumed"], reverse=True)
    by_role = sorted(role_agg.values(), key=lambda x: x["consumed"], reverse=True)
    by_department = sorted(dept_agg.values(), key=lambda x: x["consumed"], reverse=True)
    by_college = sorted(college_agg.values(), key=lambda x: x["consumed"], reverse=True)
    by_unit = sorted(unit_agg.values(), key=lambda x: x["consumed"], reverse=True)
    series = [{"date": d, "consumed": daily[d]} for d in sorted(daily.keys())]
    return {
        "days": days,
        "tx_count": len(txs),
        "total_consumed": total,
        "total_uses": total_uses,
        "total_logins": total_logins,
        "accounts_with_data": sum(1 for a in accounts if a["uses"] > 0),
        "top": accounts[:10],
        "accounts": accounts,
        "models": model_list,
        "by_category": by_category,
        "by_provider": by_provider,
        "unmapped_models": sum(1 for m in model_list if not m["mapped"]),
        "by_role": by_role,
        "by_college": by_college,
        "by_unit": by_unit,
        "by_department": by_department,
        "series": series,
        # ZH: 回傳**實際生效**的區間，不是前端送來的——
        #     前後端對日期的解讀若有出入（例如選反被換過來），
        #     這裡會看得出來；只回送來的值等於自己跟自己對帳。
        "range_start": d_start.isoformat() if d_start else None,
        "range_end": d_end.isoformat() if d_end else None,
    }


# ==============================================================================
# ZH: v3.8 數據匯出
# ------------------------------------------------------------------------------
# ZH: 直接**呼叫** consumption_analytics，不把那 90 行聚合邏輯再寫一份。
#     拄一份的話兩邊遲早會分岔，而分岔的症狀是「匯出的數字跟畫面不一樣」——
#     那種錯沒有人會馬上發現，發現了也不知道該信哪一邊。
#
# ZH: CSV 只有「帳號明細」一張表 —— CSV 本來就沒有分頁。
#     把好幾張表用空行串在同一個檔裡的話，試算表匯入時欄位會全部對不齊。
#     要完整分頁就選 Excel（介面上有寫）。
#
# ZH: CSV 加 UTF-8 BOM —— 沒加的話 Excel 開起來中文全是亂碼。
#     （與 admin.py 的使用者匯出同一個做法。）
# ==============================================================================

# ZH: 分頁定義：(分頁名, 資料鍵, 欄位[(標題, 資料鍵)])
#     第一張是「帳號明細」—— CSV 只匯出它。
_EXPORT_SHEETS = [
    ("\u5e33號明細", "accounts", [
        ("名稱", "name"), ("Email", "email"), ("廠商帳號", "vendor_sn"),
        ("消耗點數", "consumed"), ("使用次數", "uses"), ("登入次數", "logins"),
    ]),
    ("每日趨勢", "series", [("日期", "date"), ("消耗點數", "consumed")]),
    ("依模型", "models", [
        ("模型", "display_name"), ("代碼", "code"),
        ("消耗點數", "consumed"), ("使用次數", "count"),
    ]),
    ("依類別", "by_category", [
        ("類別", "category"), ("消耗點數", "consumed"), ("使用次數", "uses")]),
    ("依供應者", "by_provider", [
        ("供應者", "provider"), ("消耗點數", "consumed"), ("使用次數", "uses")]),
    ("依身分", "by_role", [
        ("身分", "role"), ("消耗點數", "consumed"), ("使用次數", "uses")]),
    ("依學系", "by_department", [
        ("學系", "department"), ("消耗點數", "consumed"), ("使用次數", "uses")]),
    # ZH: v3.8 #13 —— 學院由學系推導,行政單位只有職員有值。
    ("依學院", "by_college", [
        ("學院", "college"), ("消耗點數", "consumed"), ("使用次數", "uses")]),
    ("依行政單位", "by_unit", [
        ("行政單位", "unit"), ("消耗點數", "consumed"), ("使用次數", "uses")]),
]


def _sheet_rows(data: dict, key: str, cols: list) -> list:
    """ZH: 把一張分頁的資料持平成 list of list。

    @node job-scheduler/app/routers/external_ai.py::_sheet_rows
    """
    out = []
    for row in (data.get(key) or []):
        out.append([row.get(c[1], "") for c in cols])
    return out


@router.get("/admin/consumption/export", summary="匯出消耗分析 (CSV / Excel)")
def export_consumption(
    fmt: str = "xlsx",
    days: int = 30,
    start: Optional[str] = None,
    end: Optional[str] = None,
    db: Session = Depends(get_db),
    admin: models.User = Depends(require_admin),
) -> Any:
    """ZH: 匯出目前期間的消耗分析。篩選條件與畫面完全一致。

    @node job-scheduler/app/routers/external_ai.py::export_consumption
    """
    from fastapi.responses import StreamingResponse
    from datetime import datetime as _dt

    if fmt not in ("csv", "xlsx"):
        raise HTTPException(status_code=400, detail="ZH: fmt 只接受 csv 或 xlsx | EN: fmt must be csv or xlsx")

    # ZH: 同一支函式、同一組參數 —— 匯出與畫面不可能分岔。
    data = consumption_analytics(days=days, start=start, end=end, db=db, _=admin)

    # ZH: 檔名帶上實際區間，下載了三個月還看得出來那份是哪一段。
    if data.get("range_start") or data.get("range_end"):
        span = "%s_%s" % (data.get("range_start") or "start", data.get("range_end") or "end")
    else:
        span = "last%dd" % days if days > 0 else "all"
    stamp = _dt.now().strftime("%Y%m%d")
    filename = "consumption-%s-%s.%s" % (span, stamp, fmt)

    if fmt == "csv":
        name, key, cols = _EXPORT_SHEETS[0]
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow([c[0] for c in cols])
        w.writerows(_sheet_rows(data, key, cols))
        # ZH: BOM —— 沒它 Excel 開起來中文是亂碼。
        content = ("\ufeff" + buf.getvalue()).encode("utf-8")
        return StreamingResponse(
            io.BytesIO(content),
            # ZH: 只寫 text/csv —— Starlette 會自己補 charset。
            #     這裡再寫一次的話標頭會變成 charset 出現兩次（實測）。
            media_type="text/csv",
            headers={"Content-Disposition": 'attachment; filename="%s"' % filename},
        )

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        raise HTTPException(status_code=500,
                            detail="ZH: openpyxl 未安裝 | EN: openpyxl not installed")

    wb = Workbook()
    wb.remove(wb.active)
    for name, key, cols in _EXPORT_SHEETS:
        ws = wb.create_sheet(title=name[:31])       # ZH: Excel 分頁名上限 31 字
        ws.append([c[0] for c in cols])
        for row in _sheet_rows(data, key, cols):
            ws.append(row)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="DDEBF7")
        ws.freeze_panes = "A2"
        for i, c in enumerate(cols, start=1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = \
                min(max(len(str(c[0])) + 2, 12), 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="%s"' % filename},
    )


@router.get("/admin/user-consumption")
def user_consumption(
    q: str = "",
    days: int = 0,
    db: Session = Depends(get_db),
    _: models.User = Depends(require_admin),
) -> Any:
    """ZH: 個人查詢 —— 以 email/名稱/序號 搜尋 myai 帳號，回傳該人交易明細
            (消耗/使用/登入、模型別、每日趨勢、近期逐筆) + 同期間「全體人均」對照基準。
            days<=0 = 全部。
       EN: Per-person lookup from the transaction log (by email/name/sn),
            with a same-window all-accounts average as the comparison baseline.

    @node job-scheduler/app/routers/external_ai.py::user_consumption
    """
    from datetime import datetime, timedelta
    q = (q or "").strip()
    empty = {"q": q, "matches": [], "summary": {}, "models": [], "recent": [],
             "series": [], "peer": {}}
    if not q:
        return empty
    ql = q.lower()
    accts = db.query(models.MyaiAccount).all()
    matched = [m for m in accts
               if ql in (m.email or "").lower() or ql in (m.name or "").lower() or q == (m.vendor_sn or "")]
    if not matched:
        return empty
    sns = {m.vendor_sn for m in matched}
    umap = {(u.email or "").strip().lower(): u for u in db.query(models.User).all() if u.email}

    def _meta(m):
        """@node job-scheduler/app/routers/external_ai.py::user_consumption.<nested@852>._meta"""
        u = umap.get((m.email or "").strip().lower())
        return {
            "vendor_sn": m.vendor_sn, "name": m.name, "email": m.email,
            "current_points": m.points, "expiry": m.expiry, "status": m.status,
            "role": (u.role if u else None), "department": (u.department if u else None),
        }

    matches = [_meta(m) for m in matched]
    days = int(days or 0)
    since = datetime.now() - timedelta(days=min(days, 3650)) if days > 0 else None

    # ZH: v2.9 模型對應表 —— 個人查詢也要套，否則這裡顯示 gpt_5_6_sol、上面圖表顯示 GPT-5
    # EN: apply the display-time model map here too, so names match the global charts
    mmap = {m.code: m for m in db.query(models.MyaiModelMap).all()}
    own = _own_usage(db, sns, since, mmap)
    consumed, txs = own["consumed"], own["txs"]

    # ZH: 對照基準 —— 同期間「全體人均」。個人數字沒有基準就沒有意義：看到「消耗 1200」
    #     不知道算兇還是正常。
    peer = _peer_baseline(db, since)
    active = peer["active"] or 1
    # ZH: 排名以「單一帳號」為單位；查詢字串命中多個帳號時排名沒有意義 → 不給，別誤導。
    rank = None
    if len(sns) == 1 and consumed > 0:
        rank = 1 + sum(1 for v in peer["per_sn"].values() if v > consumed)
    series = _aligned_series(own["daily"], peer)

    model_list = _ranked_models(own["model_agg"], consumed, peer)

    recent = [{
        "time": (t.occurred_at.isoformat() if t.occurred_at else None),
        "event": t.event_type, "model": t.model, "points": t.points_delta,
        "balance": t.balance, "note": t.note,
    } for t in txs[:40]]
    return {
        "q": q, "days": days, "matches": matches,
        "summary": {"consumed": consumed, "uses": own["uses"], "logins": own["logins"],
                    "tx_count": len(txs)},
        "models": model_list,
        "series": series,
        "peer": {
            "active_accounts": peer["active"],
            "total_consumed": peer["total"],
            "avg_consumed": round(peer["total"] / active, 1),
            "avg_uses": round(peer["uses"] / active, 1),
            "rank": rank,
        },
        "recent": recent,
    }


# ==============================================================================
# ZH: v2.9 模型對應表 —— 廠商原始代碼 ↔ 顯示名稱/供應商/類別
# EN: v2.9 model map — vendor raw code ↔ display name / provider / category
# ZH: 只在顯示時套用，不改寫 myai_transactions；對錯了改一改即可，來源永遠是原始碼。
# ==============================================================================
PROVIDERS = ["Anthropic", "OpenAI", "Google", "xAI", "Perplexity", "平台工具", "其他"]
CATEGORIES = ["對話", "影像", "影音", "搜尋", "簡報", "文件", "程式", "其他"]

# ZH: (代碼前綴, 供應商, 類別) —— 只用來「建議」，admin 可自行改 | EN: guess rules only
_GUESS_RULES = [
    ("claude_",        "Anthropic",  "對話"),
    ("chatgpt_",       "OpenAI",     "對話"),
    ("gpt_image",      "OpenAI",     "影像"),
    ("gpt_",           "OpenAI",     "對話"),
    ("speech_to_text", "OpenAI",     "影音"),
    ("nano_banana",    "Google",     "影像"),
    ("gemini_",        "Google",     "對話"),
    ("grok_",          "xAI",        "對話"),
    ("perplexity_",    "Perplexity", "搜尋"),
    ("slide",          "平台工具",    "簡報"),
    ("speak_",         "平台工具",    "影音"),
    ("programming",    "平台工具",    "程式"),
]
# ZH: 完全比對的平台工具（非模型，廠商自建功能）| EN: exact-match vendor tools
_GUESS_EXACT = {
    "editor":               ("平台工具", "文件"),
    "official_document":    ("平台工具", "文件"),
    "meeting_minutes":      ("平台工具", "文件"),
    "writing_optimizer":    ("平台工具", "文件"),
    "translate_to_english": ("平台工具", "文件"),
    "gai_article_detector": ("平台工具", "文件"),
}
_NAME_FIX = {"gpt": "GPT", "chatgpt": "ChatGPT", "ai": "AI", "gai": "GAI", "xai": "xAI"}


def _pretty_name(code: str) -> str:
    """ZH: 由代碼猜好讀名稱：連續數字段併成版號（claude_opus_4_8 → Claude Opus 4.8）。
       EN: guess a friendly name; consecutive numeric tokens join as a version.

    @node job-scheduler/app/routers/external_ai.py::_pretty_name
    """
    parts = [p for p in (code or "").replace("-", "_").split("_") if p]
    out: list[str] = []
    for p in parts:
        if p.isdigit() and out and re.fullmatch(r"[\d.]+", out[-1]):
            out[-1] = f"{out[-1]}.{p}"      # ZH: 4 + 8 → 4.8
            continue
        if p.isdigit() or " " in p or any(ch.isupper() for ch in p):
            out.append(p)                   # ZH: 已含空白/大寫(如 "preprocess (OpenAI)") → 原樣保留
        else:
            out.append(_NAME_FIX.get(p.lower(), p.capitalize()))
    return " ".join(out)


def _guess_model_meta(code: str) -> dict:
    """ZH: 依代碼猜供應商/類別/顯示名稱（僅為建議值，admin 可覆寫）。

    @node job-scheduler/app/routers/external_ai.py::_guess_model_meta
    """
    low = (code or "").strip().lower()
    provider, category = "其他", "其他"
    if low in _GUESS_EXACT:
        provider, category = _GUESS_EXACT[low]
    else:
        for prefix, prov, cat in _GUESS_RULES:
            if low.startswith(prefix):
                provider, category = prov, cat
                break
    return {"code": code, "display_name": _pretty_name(code),
            "provider": provider, "category": category}


class ModelMapEntry(BaseModel):
    code: str
    display_name: str | None = None
    provider: str | None = None
    category: str | None = None
    note: str | None = None


def _model_usage_counts(db: Session) -> dict:
    """ZH: 每個代碼在交易紀錄中的筆數（給對應表顯示，方便判斷哪些真的在用）。

    @node job-scheduler/app/routers/external_ai.py::_model_usage_counts
    """
    from sqlalchemy import func
    rows = (db.query(models.MyaiTransaction.model, func.count(models.MyaiTransaction.id))
              .filter(models.MyaiTransaction.event_type == "ai_usage",
                      models.MyaiTransaction.model.isnot(None))
              .group_by(models.MyaiTransaction.model).all())
    return {m: n for m, n in rows if m}


@router.get("/admin/model-map")
def list_model_map(
    db: Session = Depends(get_db),
    _: models.User = Depends(require_admin),
) -> Any:
    """ZH: 列出對應表 + 交易中出現但「未對應」的代碼（除錯用）。
       EN: list the map plus codes seen in transactions but not mapped.

    @node job-scheduler/app/routers/external_ai.py::list_model_map
    """
    counts = _model_usage_counts(db)
    entries = db.query(models.MyaiModelMap).all()
    mapped = {(e.code or "") for e in entries}
    items = [{
        "id": e.id, "code": e.code, "display_name": e.display_name,
        "provider": e.provider, "category": e.category, "note": e.note,
        "tx_count": counts.get(e.code, 0),
        "seen": e.code in counts,          # ZH: 交易紀錄中是否真的出現過
    } for e in entries]
    items.sort(key=lambda x: (-x["tx_count"], x["code"] or ""))
    # ZH: 未對應 = 交易有、對應表沒有 → 附上建議值，前端可一鍵帶入
    unmapped = [{**_guess_model_meta(c), "tx_count": n}
                for c, n in sorted(counts.items(), key=lambda kv: -kv[1]) if c not in mapped]
    return {
        "items": items, "unmapped": unmapped,
        "providers": PROVIDERS, "categories": CATEGORIES,
        "total_codes": len(counts), "mapped_count": len(items),
        "unmapped_count": len(unmapped),
    }


@router.post("/admin/model-map")
def upsert_model_map(
    entry: ModelMapEntry,
    db: Session = Depends(get_db),
    _: models.User = Depends(require_admin),
) -> Any:
    """ZH: 新增/更新一列（以 code 為鍵，重複即更新）| EN: upsert one row by code.

    @node job-scheduler/app/routers/external_ai.py::upsert_model_map
    """
    code = (entry.code or "").strip()
    if not code:
        raise HTTPException(status_code=400, detail="代碼不可空白")
    row = db.query(models.MyaiModelMap).filter(models.MyaiModelMap.code == code).first()
    if not row:
        row = models.MyaiModelMap(code=code)
        db.add(row)
    row.display_name = (entry.display_name or "").strip() or None
    row.provider = (entry.provider or "").strip() or None
    row.category = (entry.category or "").strip() or None
    row.note = (entry.note or "").strip() or None
    db.commit(); db.refresh(row)
    return {"status": "ok", "id": row.id, "code": row.code}


@router.delete("/admin/model-map/{entry_id}")
def delete_model_map(
    entry_id: str,
    db: Session = Depends(get_db),
    _: models.User = Depends(require_admin),
) -> Any:
    """ZH: 刪除一列（原始交易資料不受影響）| EN: delete a row (raw tx unaffected).

    @node job-scheduler/app/routers/external_ai.py::delete_model_map
    """
    row = db.query(models.MyaiModelMap).filter(models.MyaiModelMap.id == entry_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="找不到該對應")
    db.delete(row); db.commit()
    return {"status": "ok"}


@router.post("/admin/model-map/seed")
def seed_model_map(
    db: Session = Depends(get_db),
    _: models.User = Depends(require_admin),
) -> Any:
    """ZH: 把交易中「未對應」的代碼用建議值一次帶入（已存在的列不動，不會覆蓋你的修改）。
       EN: bulk-insert guessed rows for unmapped codes; never overwrites existing rows.

    @node job-scheduler/app/routers/external_ai.py::seed_model_map
    """
    counts = _model_usage_counts(db)
    mapped = {c for (c,) in db.query(models.MyaiModelMap.code).all()}
    created = 0
    for code in counts:
        if code in mapped:
            continue
        g = _guess_model_meta(code)
        db.add(models.MyaiModelMap(code=code, display_name=g["display_name"],
                                   provider=g["provider"], category=g["category"]))
        created += 1
    db.commit()
    return {"status": "ok", "created": created, "skipped": len(counts) - created}
