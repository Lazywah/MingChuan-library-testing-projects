"""
==============================================================================
Router: 管理員路由群組 (Admin Routes)
==============================================================================
ZH: 用途：提供管理員專屬的使用者、任務、模型管理與數據分析端點
EN: Purpose: Admin-only endpoints for user/job/model management and analytics

ZH: 所有端點均需 JWT 認證且 role=admin，透過 require_admin Depends 強制執行
EN: All endpoints require JWT auth and role=admin, enforced via require_admin Depends

ZH: 端點清單：
    GET    /users                → 列出所有使用者（含 Token 狀態，JOIN 單查詢，支援分頁）
    PUT    /users/{id}           → 更新使用者（email/role/active/limit/password）
    PUT    /users/batch/tokens   → 批量設定 Token
    POST   /users/{id}/delete    → 刪除使用者（需驗管理員密碼）
    POST   /users/{id}/reset     → 初始化帳號（重置密碼 + 歸零用量）
    POST   /users/provision      → 配發新帳號
    POST   /verify               → 管理員密碼驗證
    GET    /jobs                 → 列出所有任務（支援分頁）
    POST   /jobs/{id}/cancel     → 強制取消任務
    PUT    /jobs/{id}/priority   → 調整任務優先級
    GET    /models               → 列出所有模型
    POST   /models               → 新增模型
    PUT    /models/{id}          → 更新模型
    DELETE /models/{id}          → 刪除模型
    GET    /cluster/stats        → 叢集 GPU 節點狀態（Worker heartbeat）
    GET    /analytics            → 數據分析（學系/工具分布）
==============================================================================
"""

from fastapi import (APIRouter, Depends, HTTPException, BackgroundTasks,
                     Query, Body, Request, UploadFile, File, Form)
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import case, func
from sqlalchemy.exc import IntegrityError
from datetime import datetime, timedelta, timezone, date as ddate, time as dtime
from typing import Any, Optional
import csv
import io
import json
import logging
import re

from .. import models, schemas, crud
from ..auth import get_current_user, require_admin as _require_admin
from ..config import SSO_POLICY, settings
from ..database import get_db
from ..services import email_service

logger = logging.getLogger(__name__)

router = APIRouter(tags=["管理員 Admin"])


# ==============================================================================
# ZH: 管理員身份驗證 Depends（取代舊的普通函式，確保注入鏈完整）
# EN: Admin auth Depends (replaces plain function to stay within FastAPI DI chain)
# ==============================================================================

# ZH: v3.8 判定搬到 auth.py 並改看 `is_admin` 旗標（身分與權限拆開）。
#     這裡保留同名的重新匯出 —— 本檔有 40+ 個端點 Depends 它,
#     全部改成 auth.require_admin 只是製造一個大 diff 而沒有任何好處。
require_admin = _require_admin


# ==============================================================================
# ZH: v3.1 step 6 — 系統設定（營運旋鈕，存 SystemConfig，runtime 生效、.env 為 fallback）
# EN: v3.1 step 6 — system settings (operational knobs; SystemConfig overrides .env at runtime)
# ==============================================================================
@router.get("/system-settings", summary="讀取營運型系統設定")
def get_system_settings(
    db: Session = Depends(get_db),
    _admin: models.User = Depends(require_admin),
):
    """ZH: 回傳每個旋鈕的 生效值/預設值/範圍/是否已覆寫。

    @node job-scheduler/app/routers/admin.py::get_system_settings
    """
    # ZH: 分組定義一起送 —— **有順序**，前端照著畫分區，不自己維護一份對照表。
    #     前端自己維護的話，後端新增旋鈕時前端那份會漏掉，
    #     而漏掉的表現是那個旋鈕安靜地不出現在任何一區。
    return {"settings": crud.get_all_settings(db), "groups": crud.SETTING_GROUPS}


@router.put("/system-settings", summary="更新營運型系統設定（值留空＝回退預設）")
def put_system_settings(
    payload: dict = Body(..., description="{key: value} 部分更新；值為空＝清除覆寫、回退 .env 預設"),
    db: Session = Depends(get_db),
    _admin: models.User = Depends(require_admin),
):
    """@node job-scheduler/app/routers/admin.py::put_system_settings"""
    try:
        updated = crud.set_settings(db, payload)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    return {"settings": updated}


# ==============================================================================
# ZH: v3.2 GPU 節點管理 — 可排程時段/開關/池別 + 狀態總覽
# EN: v3.2 GPU node management — schedule windows/switch/pool + status overview
# ==============================================================================
# ==============================================================================
# ZH: v3.8 個人組織資料的一次性解鎖
# EN: v3.8 one-shot unlock for a user's own org fields
# ==============================================================================
@router.get("/users/{user_id}/profile-unlock", summary="查目前的一次性解鎖狀態")
def get_profile_unlock(
    user_id: str,
    db: Session = Depends(get_db),
    _admin: models.User = Depends(require_admin),
):
    """
    ZH: 這個人現在有沒有還沒用掉的解鎖,以及最近一次的紀錄。

    ZH: 為什麼不併進使用者清單：清單一次回幾百人,每人多一次查詢不划算。
        跟配額、實驗室一樣,開啟帳號詳細時再非同步補。

    ZH: `last_used` 是給管理者判斷用的 —— 「他上次申請是什麼時候、改完了沒」,
        沒有這個就只看得到「現在鎖著」,分不出「從沒申請過」與「剛用掉」。

    @node job-scheduler/app/routers/admin.py::get_profile_unlock
    """
    # ZH: 查不存在的帳號要 404,與核可那支一致 —— 回 200 帶一堆 null 的話,
    #     打錯 id 看起來就像「這個人沒有解鎖」,而不是「沒有這個人」。
    if not db.query(models.User).filter(models.User.id == user_id).first():
        raise HTTPException(status_code=404, detail="找不到這個使用者")

    active = crud.active_unlock(db, user_id)
    last_used = (db.query(models.ProfileUnlock)
                 .filter(models.ProfileUnlock.user_id == user_id,
                         models.ProfileUnlock.used_at.isnot(None))
                 .order_by(models.ProfileUnlock.used_at.desc()).first())
    def _row(r):
        if r is None:
            return None
        return {"fields": [f for f in (r.fields or "").split(",") if f],
                "reason": r.reason, "granted_at": r.granted_at, "used_at": r.used_at}
    return {"unlockable": list(crud.UNLOCKABLE_FIELDS),
            "active": _row(active), "last_used": _row(last_used)}


@router.post("/users/{user_id}/profile-unlock", summary="核可一次性修改個人組織資料")
def grant_profile_unlock(
    user_id: str,
    payload: dict = Body(..., description='{"fields": ["campus"], "reason": "轉系"}'),
    db: Session = Depends(get_db),
    admin: models.User = Depends(require_admin),
):
    """
    ZH: 開放某位使用者修改一次自己的校區／學系／行政單位。

    ZH: **申請入口刻意不另外做** —— 使用者用既有的「問題回報」送單即可,
        那邊本來就有送出→管理端可見→回覆的完整流程。多做一套申請單
        只會讓使用者有兩個地方可以送、而管理者有兩個地方要看。

    ZH: 解鎖在使用者**成功存檔一次**的當下用掉,不是過多久自動失效。

    @node job-scheduler/app/routers/admin.py::grant_profile_unlock
    """
    user = crud.get_user_by_id(db, user_id) if hasattr(crud, "get_user_by_id") else         db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="找不到這個使用者")
    try:
        row = crud.grant_profile_unlock(db, user, payload.get("fields") or [],
                                        admin, payload.get("reason") or "")
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    # ZH: 稽核。本檔沒有 _log_admin_action 這個 helper（那支在 storage_lifecycle）,
    #     直接寫 AdminAction —— 與本檔其他三處一致。
    db.add(models.AdminAction(
        admin_id=admin.id,
        target_user=user.id,
        action="grant_profile_unlock",
        payload=json.dumps({"fields": row.fields, "reason": row.reason},
                            ensure_ascii=False),
    ))
    db.commit()
    return {"fields": row.fields.split(","), "granted_at": row.granted_at,
            "used_at": row.used_at}


@router.post("/users/{user_id}/profile-unlock/revoke", summary="收回還沒用掉的解鎖")
def revoke_profile_unlock(
    user_id: str,
    db: Session = Depends(get_db),
    admin: models.User = Depends(require_admin),
):
    """
    ZH: 把還沒用掉的解鎖收回來（核可錯了、或那個人已經不需要改了）。

    ZH: 🔴 **標成已用掉,不刪紀錄。** 刪掉的話「這個人曾經被開過一次」就查不到了 ——
        而「開過但沒用」正是稽核時想知道的事（例如核可之後才發現不該開）。

    @node job-scheduler/app/routers/admin.py::revoke_profile_unlock
    """
    row = crud.active_unlock(db, user_id)
    if row is None:
        raise HTTPException(status_code=404, detail="這個帳號目前沒有未使用的解鎖")
    row.used_at = datetime.now(timezone.utc)
    db.add(models.AdminAction(
        admin_id=admin.id,
        target_user=user_id,
        action="revoke_profile_unlock",
        payload=json.dumps({"fields": row.fields}, ensure_ascii=False),
    ))
    db.commit()
    return {"revoked": True, "fields": [f for f in (row.fields or "").split(",") if f]}


def _bounce_status(db) -> dict:
    """
    ZH: 退信回收現在的狀態 —— 給管理端顯示用，**不做任何連線**。

    ZH: 這裡刻意只回「設定推導出來的結果」與「上次掃描的紀錄」，
        不當場去連 IMAP —— 一個顯示用的端點不該因為外部服務掛掉就卡住
        （小基的 status 端點也是同一個原則）。

    ZH: `host` 是空字串就代表**停用**。它預設由 SMTP 主機推導，
        所以管理者改一次 SMTP 就可能連帶把退信回收關掉，而且完全不報錯 ——
        這一行的存在就是為了讓那件事在畫面上看得見。

    @node job-scheduler/app/routers/admin.py::_bounce_status
    """
    from ..services import bounce_reader
    try:
        cfg = bounce_reader.imap_config(db)
    except Exception:  # noqa: BLE001 - 顯示用，取不到就當停用
        cfg = {}
    last = crud.get_system_config(db, "bounce_last_scan", "") or ""
    return {
        "host": cfg.get("host") or "",
        "folder": cfg.get("folder") or "",
        "enabled": bool(cfg.get("host") and cfg.get("user") and cfg.get("password")),
        "interval_minutes": crud.get_setting(db, "bounce_scan_minutes"),
        # ZH: 格式 "<ISO 時間>|<scanned>|<bounces>|<applied>"，由 scan_bounces 寫入。
        #     空字串＝服務啟動後還沒掃過（不是「掃過但沒東西」——那兩件事要分得出來）。
        "last_scan": last,
    }


@router.get("/email-log", summary="寄信紀錄（誰、何時、結果）")
def get_email_log(
    limit: int = Query(200, ge=1, le=1000),
    status: Optional[str] = Query(None, description="sent / refused / failed / mock / bounced / deferred"),
    db: Session = Depends(get_db),
    _admin: models.User = Depends(require_admin),
):
    """
    ZH: v3.4 寄信紀錄。⚠️ `sent` = **已交付中繼伺服器，不代表送達**；
        「網域存在但信箱不存在」會被中繼接受、稍後非同步退信到寄件人信箱，程式端看不到。
        v3.5 起會由退信回收（IMAP）把非同步退信回填成 `bounced`(永久，5.x.x：信箱不存在)
        或 `deferred`(暫時，4.x.x：稍後可能仍會送達，**不代表不存在**)。
    EN: Outbound email log; `sent` means accepted by relay, not delivered. v3.5 back-fills
        real bounces read over IMAP.

    @node job-scheduler/app/routers/admin.py::get_email_log
    """
    q = db.query(models.EmailLog)
    if status:
        q = q.filter(models.EmailLog.status == status)
    rows = q.order_by(models.EmailLog.created_at.desc()).limit(limit).all()
    _smtp_cfg = crud.effective_smtp(db)
    counts = dict(
        db.query(models.EmailLog.status, func.count(models.EmailLog.id))
        .group_by(models.EmailLog.status).all()
    )
    return {
        "counts": counts,
        # ZH: v3.8 顯示生效值 —— 這一頁就是管理者用來確認「我剛改的設定生效了沒」的地方，
        #     顯示 .env 的舊值會讓他以為沒存到，然後再存一次。
        "smtp_configured": bool(_smtp_cfg["server"]),
        "from_email": _smtp_cfg["from_email"],
        # ZH: v3.8 —— 退信回收的狀態。
        #
        # ZH: 為什麼要放在這一頁：退信回收是「這封信到底送到了沒」的**唯一事實來源**，
        #     但它在介面上完全看不見。2026-08-27 稽核時它停了 40 分鐘
        #     （SMTP 主機被改成推導不出 IMAP 的值），而我是靠翻容器日誌才發現的 ——
        #     管理者不會去翻容器日誌。
        #
        # ZH: `host` 空字串＝停用。IMAP 主機**預設由 SMTP 推導**，
        #     所以改了 SMTP 就可能連帶把它關掉，這一行讓那件事看得見。
        "bounce": _bounce_status(db),
        "logs": [{
            "id": r.id, "to_email": r.to_email, "username": r.username,
            "kind": r.kind, "subject": r.subject, "status": r.status,
            "detail": r.detail,
            "created_at": r.created_at.isoformat() if r.created_at else None,
            "bounced_at": r.bounced_at.isoformat() if r.bounced_at else None,
        } for r in rows],
    }


@router.post("/email-log/scan-bounces", summary="v3.5 立即用 IMAP 掃退信並回填")
def scan_bounces_now(
    db: Session = Depends(get_db),
    _admin: models.User = Depends(require_admin),
):
    """
    ZH: 手動觸發一次退信回收（平常由排程每 bounce_scan_minutes 分鐘自動跑）。
        唯讀：只讀退信、只標記已讀，絕不刪信/移動/寄信。
        回 {scanned, bounces, applied}；applied=0 代表退信對不到任何寄件紀錄
        （例如那些信寄出時還沒有 email_log）——此時**不會**新增假紀錄。
    EN: Trigger one bounce-scan pass now; read-only, never fabricates log rows.

    @node job-scheduler/app/routers/admin.py::scan_bounces_now
    """
    from ..services import bounce_reader
    try:
        return bounce_reader.scan_bounces(db)
    except bounce_reader.BounceReaderError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/bootstrap-status", summary="初始管理員是否仍在使用預設密碼")
def bootstrap_admin_status(
    db: Session = Depends(get_db),
    _admin: models.User = Depends(require_admin),
):
    """
    ZH: v3.3 給 admin UI 判斷是否顯示「請盡快改密碼」橫幅。
        判定方式＝拿 .env 的 BOOTSTRAP_ADMIN_PASSWORD 去驗證 admin 帳號的雜湊；
        驗得過代表初始密碼仍有效 → 顯示提醒。改過密碼後自動驗不過 → 橫幅自動消失。
        （不回傳密碼本身，只回布林。）
    EN: Whether the bootstrap admin still uses the initial password (banner trigger).

    @node job-scheduler/app/routers/admin.py::bootstrap_admin_status
    """
    pw = (settings.BOOTSTRAP_ADMIN_PASSWORD or "").strip()
    if not pw:
        return {"using_initial_password": False, "username": None}
    user = crud.get_user_by_username(db, "admin")
    if not user or not user.is_admin:
        return {"using_initial_password": False, "username": None}
    try:
        still = crud.verify_password(pw, user.hashed_password)
    except Exception:  # noqa: BLE001
        still = False
    return {"using_initial_password": bool(still), "username": "admin" if still else None}


@router.get("/gpu-nodes", summary="GPU 節點狀態與設定總覽")
def list_gpu_nodes(
    db: Session = Depends(get_db),
    _admin: models.User = Depends(require_admin),
):
    """ZH: 每節點回 設定+心跳即時值+狀態(四態)+下次開關時間+執行中任務+撞名警示。

    @node job-scheduler/app/routers/admin.py::list_gpu_nodes
    """
    return {"nodes": crud.list_gpu_nodes_with_status(db)}


@router.put("/gpu-nodes/{node_id}", summary="更新單一 GPU 節點設定")
def put_gpu_node(
    node_id: str,
    payload: dict = Body(..., description="可含 display_name/note/enabled/pool_override/schedule/dispatch_buffer_min"),
    db: Session = Depends(get_db),
    _admin: models.User = Depends(require_admin),
):
    """
    ZH: 更新節點設定（schedule 格式見 gpu_schedule.py；清空=全天可排）。
        節點未心跳過也可先建設定（pre-provision）。成功後回全列表讓前端整頁刷新。

    @node job-scheduler/app/routers/admin.py::put_gpu_node
    """
    try:
        crud.update_gpu_node(db, node_id, payload)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    return {"nodes": crud.list_gpu_nodes_with_status(db)}


# ==============================================================================
# ZH: 使用者管理 | EN: User Management
# ==============================================================================

# v2.1 在線狀態修正：admin 不再讀 DB 內 online_status 欄位（會 stale），
# 改用 last_activity 動態計算「10 分鐘內活躍 = 在線」
from datetime import timedelta as _td
_ONLINE_THRESHOLD = _td(minutes=10)

def _compute_online(user: models.User) -> Optional[int]:
    """
    ZH: 用 last_activity 動態判斷在線
    EN: Compute online from last_activity

    v2.1 修正：admin 從未登入過 user UI（last_login_time 為 None）→ 回 None，
    讓 admin UI 顯示「—」而非誤導性的「離線」。
    admin 一旦登入過 user UI（即使後來離線），仍回 0/1 正常計算。

    @node job-scheduler/app/routers/admin.py::_compute_online
    """
    # ZH: ⚠️ v3.8 拆開身分與權限後這裡**刻意仍看 `role`** ——
    #     講的是「只用後台、從沒進過前台的系統帳號」,不是權限。
    if user.role == "admin" and user.last_login_time is None:
        return None
    last = getattr(user, "last_activity", None)
    if not last:
        return 0
    if last.tzinfo is None:
        last = last.replace(tzinfo=timezone.utc)
    return 1 if (datetime.now(timezone.utc) - last) < _ONLINE_THRESHOLD else 0


def _yaml_mock_usernames() -> set:
    """
    ZH: 從 SSO_POLICY 抓出目前 yaml 內所有 mock SSO 帳號的 student_id 集合
    EN: Set of student_ids currently allowed via mock SSO in yaml

    v2.1: 用於 filter 已從 yaml 移除但 DB 仍有 row 的 sso_mock 使用者
    (保留 orphan row 不破壞既有聊天歷史 / 任務 FK，僅在列表中隱藏)

    @node job-scheduler/app/routers/admin.py::_yaml_mock_usernames
    """
    try:
        users = (SSO_POLICY or {}).get("mock", {}).get("users", []) or []
        return {str(u.get("student_id")) for u in users if u.get("student_id")}
    except Exception as e:
        logger.warning(f"_yaml_mock_usernames failed: {e}")
        return set()


@router.get("/users", response_model=list[schemas.AdminUserListItem])
def get_all_users(
    skip: int = Query(0, ge=0, description="ZH: 跳過筆數 | EN: Records to skip"),
    limit: int = Query(100, ge=1, le=500, description="ZH: 每頁筆數 | EN: Records per page"),
    auth_source: Optional[str] = Query(
        None,
        description=(
            "ZH: 依登入來源過濾（local / sso_oidc / sso_cas / sso_mock）"
            " | EN: Filter by auth_source"
        ),
    ),
    db: Session = Depends(get_db),
    _: models.User = Depends(require_admin),
) -> Any:
    """
    ZH: 列出所有使用者，單次 JOIN 查詢避免 N+1，支援分頁
    EN: List all users with token usage via single JOIN query, supports pagination

    v2.1 擴充：
    - 可用 ?auth_source=local|sso_oidc|sso_mock 過濾分類
    - sso_mock 帳號額外做 yaml filter：若 username 已從 sso_policy.yaml 移除則隱藏

    @node job-scheduler/app/routers/admin.py::get_all_users
    """
    query = (
        db.query(models.User, models.TokenUsage)
        .outerjoin(models.TokenUsage, models.TokenUsage.user_id == models.User.id)
    )
    if auth_source:
        query = query.filter(models.User.auth_source == auth_source)
    rows = (
        query.order_by(models.User.created_at.desc())
        .offset(skip)
        .limit(limit)
        .all()
    )

    # v2.1 yaml filter: 若使用者已從 yaml 移除，從列表隱藏（DB row 仍保留）
    yaml_usernames = _yaml_mock_usernames()

    # ZH: 學系 → 英文名。一次查完做成表，不要每一列查一次（500 列就是 500 次）。
    # ZH: department 是**自由字串沒有外鍵**，所以只能靠名字對。對不到就是 None，
    #     前端退回中文 —— 那是正常情況（打錯字、或還沒進對照表）。
    dept_en = {d.name: d.name_en for d in db.query(models.OrgDepartment).all() if d.name_en}

    result = []
    for u, t in rows:
        if u.auth_source == "sso_mock" and u.username not in yaml_usernames:
            continue  # yaml 已移除 → 列表隱藏
        result.append(
            schemas.AdminUserListItem(
                id=u.id,
                username=u.username,
                email=u.email,
                # ZH: v4.3 常用信箱 —— 上下兩處警告講的就是這種欄位，這次還是踩了
                #     （schema 加了、這裡漏帶，管理端顯示永遠 None）。
                contact_email=getattr(u, "contact_email", None),
                role=u.role,
                role_source=getattr(u, "role_source", None),
                is_admin=int(getattr(u, "is_admin", 0) or 0),
                is_active=u.is_active,
                online_status=_compute_online(u),  # v2.1: 動態計算
                last_login_time=u.last_login_time,
                last_login_ip=u.last_login_ip,
                department=u.department,
                # ZH: ⚠ 上面那句警告就是在講這裡 —— schema 加了欄位還要在這裡帶上，
                #     不然它會靜靜地永遠是 None。
                department_en=dept_en.get(u.department),
                created_at=u.created_at,
                tokens_used=t.tokens_used if t else 0,
                tokens_limit=t.tokens_limit if t else 0,
                auth_source=getattr(u, "auth_source", "local") or "local",  # v2.1: 3-tab 分頁
                # ⚠ ZH: 這裡是**手工建構**的，光在 schema 加欄位沒有用 ——
                #   會靜靜地永遠回 None。這個坑在 v3.6 踩過兩次（metrics、has_model）。
                expires_at=u.expires_at,
                temp_purpose=u.temp_purpose,
            )
        )
    return result


# ==============================================================================
# v2.2: 使用者管理 Excel/CSV 匯出（欄位 + 範圍 admin 可勾選）
# v2.2: User-management export to Excel/CSV (admin chooses columns + scope)
# ==============================================================================

# 欄位白名單 — 防止 admin 隨便丟未授權欄位名稱（避免存取 hashed_password 等敏感欄）
# Whitelist of allowed export columns — prevents injection of sensitive attribute names
_EXPORT_COLUMNS = {
    # key: (顯示標題, getter function)
    "username":            ("帳號名稱", lambda u, t: u.username),
    "email":               ("Email", lambda u, t: u.email),
    "role":                ("角色", lambda u, t: u.role),
    # ZH: v3.8 讓管理者用匯出就能複查「哪些人的角色是自動判的」——
    #     自動判定的依據是我們自己組出來的信箱,不是學校給的權威資料。
    "role_source":         ("角色來源", lambda u, t: getattr(u, "role_source", None) or "未記錄"),
    "is_admin":            ("管理權限", lambda u, t: bool(getattr(u, "is_admin", 0))),
    "auth_source":         ("登入來源", lambda u, t: getattr(u, "auth_source", "local") or "local"),
    "is_active":           ("是否啟用", lambda u, t: bool(u.is_active)),
    "department":          ("學系", lambda u, t: u.department or ""),
    "last_login_time":     ("最後登入時間", lambda u, t: u.last_login_time.isoformat() if u.last_login_time else ""),
    "last_login_ip":       ("最後登入 IP", lambda u, t: u.last_login_ip or ""),
    "created_at":          ("建立日期", lambda u, t: u.created_at.isoformat() if u.created_at else ""),
    "login_count":         ("登入次數", lambda u, t: u.login_count or 0),
    "tokens_used":         ("Token 已用", lambda u, t: (t.tokens_used if t else 0)),
    "tokens_limit":        ("Token 配額", lambda u, t: (t.tokens_limit if t else 0)),
    "lifetime_tokens_used":("歷史累計 Token", lambda u, t: u.lifetime_tokens_used or 0),
    "online_status":       ("線上狀態", lambda u, t: _compute_online(u)),
}


@router.get("/users/export", summary="v2.2 — 匯出使用者管理資料 (Excel / CSV)")
def export_users(
    columns: str = Query(
        "username,email,role,auth_source,is_active,department,last_login_time,tokens_used,tokens_limit",
        description="ZH: 逗號分隔欄位名稱 (見 _EXPORT_COLUMNS 白名單) | EN: comma-separated column names",
    ),
    fmt: str = Query("xlsx", pattern="^(xlsx|csv)$", description="ZH: xlsx 或 csv | EN: xlsx | csv"),
    scope: str = Query("filter", pattern="^(filter|all)$", description="ZH: filter=套用 auth_source 篩選 / all=全部 | EN: filter | all"),
    auth_source: Optional[str] = Query(None, description="ZH: 當 scope=filter 時的篩選值 | EN: filter value when scope=filter"),
    db: Session = Depends(get_db),
    current_admin: models.User = Depends(require_admin),
) -> Any:
    """
    ZH: 把使用者管理列表匯出成 Excel (.xlsx) 或 CSV。
    EN: Export user-management list as Excel (.xlsx) or CSV.

    使用範例 / Examples:
      GET /api/v1/admin/users/export?fmt=xlsx&columns=username,email,tokens_used
      GET /api/v1/admin/users/export?fmt=csv&scope=filter&auth_source=local

    安全 / Security:
      - 限 admin (require_admin)
      - 欄位走白名單 (_EXPORT_COLUMNS)，避免拉到 hashed_password 等敏感欄位
      - 寫入 audit log

    @node job-scheduler/app/routers/admin.py::export_users
    """
    # 解析 + 驗證 columns
    requested = [c.strip() for c in columns.split(",") if c.strip()]
    if not requested:
        raise HTTPException(status_code=400, detail="columns 不可為空 / columns must not be empty")
    unknown = [c for c in requested if c not in _EXPORT_COLUMNS]
    if unknown:
        raise HTTPException(
            status_code=400,
            detail=f"未知欄位 / Unknown columns: {unknown}. 允許 / Allowed: {list(_EXPORT_COLUMNS)}",
        )

    # 查資料（沿用 get_all_users 的 join + filter 邏輯）
    query = (
        db.query(models.User, models.TokenUsage)
        .outerjoin(models.TokenUsage, models.TokenUsage.user_id == models.User.id)
    )
    if scope == "filter" and auth_source:
        query = query.filter(models.User.auth_source == auth_source)
    rows = query.order_by(models.User.created_at.desc()).all()

    # v2.1 yaml filter（同 get_all_users）
    yaml_usernames = _yaml_mock_usernames()
    visible = [
        (u, t) for u, t in rows
        if not (u.auth_source == "sso_mock" and u.username not in yaml_usernames)
    ]

    # 組標題列 + 資料列
    headers = [_EXPORT_COLUMNS[c][0] for c in requested]
    data_rows = [
        [_EXPORT_COLUMNS[c][1](u, t) for c in requested]
        for u, t in visible
    ]

    # audit log
    try:
        import json as _json
        db.add(models.AdminAction(
            admin_id=current_admin.id,
            target_user=None,
            action="export_users",
            payload=_json.dumps({
                "format": fmt,
                "scope": scope,
                "auth_source": auth_source,
                "columns": requested,
                "row_count": len(visible),
            }, ensure_ascii=False),
            timestamp=datetime.now(timezone.utc),
        ))
        db.commit()
    except Exception as e:
        # audit log 失敗不阻止匯出
        logger.warning(f"audit log write failed for export_users: {e}")
        db.rollback()

    # 產生檔案
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    filename = f"users-export-{timestamp}.{fmt}"

    if fmt == "csv":
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(headers)
        writer.writerows(data_rows)
        # 加 UTF-8 BOM 讓 Excel 開 CSV 時正確顯示中文
        content = ("﻿" + buf.getvalue()).encode("utf-8")
        return StreamingResponse(
            io.BytesIO(content),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    # fmt == "xlsx"
    try:
        from openpyxl import Workbook
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl 未安裝 / openpyxl not installed; pip install openpyxl")

    wb = Workbook()
    ws = wb.active
    ws.title = "Users"
    ws.append(headers)
    for row in data_rows:
        ws.append(row)

    # 簡單格式化：標題列粗體 + 凍結首列
    from openpyxl.styles import Font, PatternFill
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDEBF7")
    ws.freeze_panes = "A2"

    # auto width（粗略）
    for col_idx, col_name in enumerate(headers, start=1):
        max_len = max(
            [len(str(col_name))]
            + [len(str(row[col_idx - 1])) for row in data_rows[:200]]  # 看前 200 行
        )
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/users/export/columns", summary="v2.2 — 列出可匯出的欄位清單（給前端 UI 用）")
def export_users_columns(
    _: models.User = Depends(require_admin),
) -> Any:
    """ZH: 回傳 [{key, label}, ...] 給前端 modal 動態建勾選清單

    @node job-scheduler/app/routers/admin.py::export_users_columns
    """
    return [{"key": k, "label": v[0]} for k, v in _EXPORT_COLUMNS.items()]


@router.put("/users/batch/tokens")
def batch_update_tokens(
    payload: schemas.BatchTokenUpdate,
    db: Session = Depends(get_db),
    _: models.User = Depends(require_admin),
) -> Any:
    """
    ZH: 管理員批量更新指定使用者的 Token 狀態
    EN: Admin batch update token state for specified users

    ZH: action = reset_usage → 將 tokens_used 歸零
    ZH: action = set_limit   → 將 tokens_limit 設為 payload.value

    @node job-scheduler/app/routers/admin.py::batch_update_tokens
    """
    if not payload.user_ids:
        raise HTTPException(status_code=400, detail="ZH: 至少要選一個帳號 | EN: user_ids cannot be empty")
    if payload.action not in ("reset_usage", "set_limit"):
        raise HTTPException(status_code=400, detail="ZH: action 只能是 reset_usage 或 set_limit | EN: action must be 'reset_usage' or 'set_limit'")

    now = datetime.now(timezone.utc)

    if payload.action == "reset_usage":
        updated = (
            db.query(models.TokenUsage)
            .filter(models.TokenUsage.user_id.in_(payload.user_ids))
            .update(
                {models.TokenUsage.tokens_used: 0, models.TokenUsage.last_updated: now},
                synchronize_session=False,
            )
        )
    else:  # set_limit
        updated = (
            db.query(models.TokenUsage)
            .filter(models.TokenUsage.user_id.in_(payload.user_ids))
            .update(
                {models.TokenUsage.tokens_limit: payload.value, models.TokenUsage.last_updated: now},
                synchronize_session=False,
            )
        )

    db.commit()
    logger.info("batch_update_tokens: action=%s value=%s updated=%d", payload.action, payload.value, updated)
    return {"updated_count": updated, "action": payload.action, "value": payload.value}


@router.put("/users/{user_id}")
def admin_update_user(
    user_id: str,
    update_data: schemas.AdminUserUpdate,
    db: Session = Depends(get_db),
    _: models.User = Depends(require_admin),
) -> Any:
    """ZH: 管理員修改使用者資訊 | EN: Admin update user details

    @node job-scheduler/app/routers/admin.py::admin_update_user
    """
    db_user = db.query(models.User).filter(models.User.id == user_id).first()
    if not db_user:
        raise HTTPException(status_code=404, detail="ZH: 找不到這個帳號 | EN: User not found")

    if update_data.email is not None:
        db_user.email = update_data.email
    if update_data.is_admin is not None:
        # ZH: v3.8 管理權限與身分分開設。這是**唯一**能寫這個欄位的地方 ——
        #     使用者端的 UserUpdate 沒有這個欄位,PUT /auth/me 表達不出它。
        db_user.is_admin = 1 if update_data.is_admin else 0
    if update_data.role is not None:
        db_user.role = update_data.role
        # ZH: v3.8 管理者改過就不再是「自動判定」—— 複查清單要把他排除,
        #     否則每次複查都會再看到同一個已經確認過的人。
        db_user.role_source = "admin"
    if update_data.is_active is not None:
        db_user.is_active = update_data.is_active
    if update_data.department is not None:
        db_user.department = update_data.department
    if update_data.password is not None and update_data.password.strip():
        db_user.hashed_password = crud.get_password_hash(update_data.password)

    if update_data.tokens_limit is not None:
        usage = crud.get_token_usage(db, user_id)
        if usage:
            usage.tokens_limit = update_data.tokens_limit

    db.commit()
    db.refresh(db_user)

    usage = crud.get_token_usage(db, user_id)
    return {
        "id": db_user.id,
        "username": db_user.username,
        "email": db_user.email,
        "role": db_user.role,
        "is_active": db_user.is_active,
        "tokens_used": usage.tokens_used if usage else 0,
        "tokens_limit": usage.tokens_limit if usage else 0,
    }


@router.post("/users/{user_id}/delete")
def admin_delete_user(
    user_id: str,
    payload: schemas.AdminDeleteUser,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin),
) -> Any:
    """ZH: 管理員刪除使用者 (需驗證密碼) | EN: Admin delete user (requires password verification)

    @node job-scheduler/app/routers/admin.py::admin_delete_user
    """
    if not crud.verify_password(payload.admin_password, current_user.hashed_password):
        raise HTTPException(status_code=403, detail="ZH: 管理員密碼不對 | EN: Invalid admin password")
    if user_id == current_user.id:
        raise HTTPException(status_code=400, detail="ZH: 不能刪除自己的帳號 | EN: Cannot delete yourself")

    db_user = db.query(models.User).filter(models.User.id == user_id).first()
    if not db_user:
        raise HTTPException(status_code=404, detail="ZH: 找不到這個帳號 | EN: User not found")

    username = db_user.username

    # ==========================================================================
    # ZH: v3.3 完整刪除流程（原本只刪 token_usage → 會 FK 失敗或留下孤兒）
    #   1. Lab 資料先「封存」（容器移除、volume 原地保留 N 天，可還原）
    #   2. 解開稽核類外鍵參照（admin_actions / quota_grants 為 ON DELETE NO ACTION，
    #      不處理會讓刪除直接 IntegrityError → 500；target_user 可為 NULL，
    #      故解參照而非刪紀錄，稽核軌跡得以保留）
    #   2b. v3.4 issue_reports **不在這裡處理**，因為它的兩個 users FK 都宣告了
    #       ondelete="SET NULL"，DB 自己會解 —— 與第 4 點同類，不是漏掉。
    #       （admin_actions / quota_grants 之所以要手動處理，是因為它們沒宣告
    #        ondelete，預設 NO ACTION，刪除會直接 IntegrityError。差別在這裡。）
    #       選 SET NULL 而非 CASCADE 是刻意的：**帳號刪了，問題可能還在**，
    #       那仍是管理者的待辦。username_at_report 留下送出當下的快照，
    #       所以解參照之後仍看得出是誰報的；replied_by 解掉但回應內容保留。
    #       ⚠ 這條路徑由 tests/test_reports.py 兩支刪帳號測試守著，而那兩支
    #        **要有 conftest 的 PRAGMA foreign_keys=ON 才驗得到**（見該處註解）。
    #   3. 清掉無 FK 約束、不會自動 cascade 的表（chat_history / training_jobs / token_usage）
    #   4. 其餘（external_ai_accounts / lab_sessions / user_secrets / user_session_usage /
    #      user_storage_state / quota_grants.user_id）由 DB 的 ON DELETE CASCADE 處理
    # EN: v3.3 full deletion: archive Lab data, unlink audit FKs (else IntegrityError),
    #     purge non-cascading tables, then delete the user.
    # ==========================================================================
    archived = None
    try:
        from ..services import lab_manager
        archived = lab_manager.archive_user_lab(
            db, db_user, retention_days=crud.get_setting(db, "lab_archive_days"),
            reason="admin_delete",
        )
    except Exception as e:  # noqa: BLE001 - Lab 封存失敗不應阻擋帳號刪除
        logger.error(f"Lab 封存失敗（仍繼續刪除帳號）: {e}")

    # ZH: v3.6 —— 資料集與訓練產出的**實體檔案**。
    #     DB 那邊 datasets 是 ON DELETE CASCADE、training_jobs 在下面直接刪，
    #     但**磁碟上的檔案不會跟著消失** —— 刪了帳號，那些 zip 與模型檔會永遠留著。
    #     這與 Lab volume 封存是同一類：DB 與實體儲存要一起處理。
    #     ⚠ 與 Lab 不同的是**不做封存**：資料集是使用者自己上傳的原始檔，
    #       他要保留應該自己留一份；平台沒有理由替他保管一份副本。
    try:
        import shutil as _sh
        from .datasets import DATASET_DIR as _dsd
        import os as _os
        _sh.rmtree(_os.path.join(_dsd, user_id), ignore_errors=True)
    except Exception as e:  # noqa: BLE001 - 清檔案失敗不該擋住帳號刪除
        logger.error(f"ZH: 清除使用者資料集檔案失敗（仍繼續刪除帳號）: {e}")
    try:
        from . import worker as _wr
        for _j in db.query(models.TrainingJob).filter(
                models.TrainingJob.user_id == user_id,
                models.TrainingJob.artifact_bytes.isnot(None)).all():
            _wr.remove_artifact_file(_j.id)
    except Exception as e:  # noqa: BLE001
        logger.error(f"ZH: 清除使用者模型檔失敗（仍繼續刪除帳號）: {e}")

    try:
        db.query(models.AdminAction).filter(
            models.AdminAction.target_user == user_id
        ).update({models.AdminAction.target_user: None}, synchronize_session=False)
        db.query(models.QuotaGrant).filter(
            models.QuotaGrant.granted_by == user_id
        ).delete(synchronize_session=False)

        db.query(models.ChatHistory).filter(models.ChatHistory.user_id == user_id).delete(synchronize_session=False)
        db.query(models.TrainingJob).filter(models.TrainingJob.user_id == user_id).delete(synchronize_session=False)
        db.query(models.TokenUsage).filter(models.TokenUsage.user_id == user_id).delete(synchronize_session=False)

        # ==================================================================
        # ZH: v3.8 **刪除本身要留稽核。** 在此之前,刪帳號是管理端破壞性最強的
        #     動作,卻是唯一完全不留痕跡的 —— 誰刪的、什麼時候、刪了誰,查不到。
        #     而且上面那一步還會把先前提到這個人的稽核紀錄 target_user 洗成 NULL,
        #     於是連間接的線索都沒有。（2026-08-27 實地發現:一個帳號不見了,
        #     翻遍稽核只有 export_users / create_temp_account,沒有任何刪除紀錄。）
        #
        # ZH: 🔴 **target_user 必須留 NULL**：它是 FK,而這一列指向的使用者
        #     下一行就要被刪掉了 —— 填了會直接 IntegrityError（NO ACTION）。
        #     身分資訊全部放進 payload,那是 Text 欄位,不會被上面的解參照洗掉。
        #
        # ZH: 快照要**在刪之前取**,而且要夠完整 —— 刪完之後這些資訊
        #     沒有任何地方查得到。含封存的 volume 名,因為那是唯一還能救回資料的線索。
        # ==================================================================
        db.add(models.AdminAction(
            admin_id=current_user.id,
            target_user=None,
            action="delete_user",
            payload=json.dumps({
                "deleted_user_id": user_id,
                "username": db_user.username,
                "email": db_user.email,
                "role": db_user.role,
                "is_admin": int(getattr(db_user, "is_admin", 0) or 0),
                "auth_source": getattr(db_user, "auth_source", None),
                "department": db_user.department,
                "unit": getattr(db_user, "unit", None),
                "created_at": db_user.created_at.isoformat() if db_user.created_at else None,
                "last_login_time": (db_user.last_login_time.isoformat()
                                    if db_user.last_login_time else None),
                # ZH: 封存的 volume 名 —— 逾期銷毀前這是唯一救得回資料的線索。
                "lab_archive": (archived or {}).get("volume") if isinstance(archived, dict) else None,
                "lab_archive_expires": ((archived or {}).get("expires_at")
                                        if isinstance(archived, dict) else None),
            }, ensure_ascii=False, default=str),
        ))

        db.delete(db_user)
        db.commit()
    except IntegrityError as e:
        db.rollback()
        logger.error(f"刪除使用者 {username} 失敗（外鍵約束）: {e}")
        raise HTTPException(
            status_code=409,
            detail="無法刪除：此帳號仍被其他資料引用。請聯絡開發者檢查關聯資料。",
        )

    logger.info(f"User {username} deleted by admin {current_user.username} (lab_archived={bool(archived)})")
    return {
        "message": f"User {username} deleted",
        "deleted_id": user_id,
        "lab_archived": archived,   # ZH: None=無 Lab 資料；否則含 volume/大小/到期
    }


# ==============================================================================
# ZH: v3.3 Lab 資料封存管理（刪除帳號後的緩衝區：可檢視 / 還原 / 立即銷毀）
# EN: v3.3 archived Lab volumes — list / restore / purge-now
# ==============================================================================
@router.get("/lab-archives", summary="列出封存中的 Lab 資料")
def list_lab_archives(
    db: Session = Depends(get_db),
    _admin: models.User = Depends(require_admin),
):
    """@node job-scheduler/app/routers/admin.py::list_lab_archives"""
    from ..services import lab_manager
    return {"archives": lab_manager.list_archives(db),
            "retention_days": crud.get_setting(db, "lab_archive_days")}


@router.post("/lab-archives/{volume_name}/restore", summary="把封存的 Lab 還原給指定使用者")
def restore_lab_archive(
    volume_name: str,
    payload: dict = Body(..., description='{"target_user_id": "..."}'),
    db: Session = Depends(get_db),
    _admin: models.User = Depends(require_admin),
):
    """ZH: SSO 使用者刪除後會以新 uuid 回來，故還原＝複製進目標使用者現有的 volume。

    @node job-scheduler/app/routers/admin.py::restore_lab_archive
    """
    from ..services import lab_manager
    target = (payload or {}).get("target_user_id")
    if not target:
        raise HTTPException(status_code=400, detail="缺少 target_user_id")
    try:
        return lab_manager.restore_archive(db, volume_name, target)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:  # noqa: BLE001
        logger.exception("Lab 還原失敗")
        raise HTTPException(status_code=500, detail=f"還原失敗：{e}")


@router.delete("/lab-archives/{volume_name}", summary="立即銷毀某筆封存（不等到期）")
def delete_lab_archive(
    volume_name: str,
    db: Session = Depends(get_db),
    _admin: models.User = Depends(require_admin),
):
    """@node job-scheduler/app/routers/admin.py::delete_lab_archive"""
    from ..services import lab_manager
    if not lab_manager.delete_archive_now(db, volume_name):
        raise HTTPException(status_code=404, detail="找不到該封存紀錄")
    return {"message": "已銷毀", "volume_name": volume_name}


@router.post("/verify")
def admin_verify_action(
    payload: schemas.AdminVerify,
    current_user: models.User = Depends(require_admin),
) -> Any:
    """ZH: 管理員密碼驗證（解鎖敏感操作）| EN: Admin password verification (unlock sensitive actions)

    @node job-scheduler/app/routers/admin.py::admin_verify_action
    """
    if not crud.verify_password(payload.admin_password, current_user.hashed_password):
        raise HTTPException(status_code=403, detail="ZH: 管理員密碼不對 | EN: Invalid admin password")
    return {"message": "Verification successful"}


@router.post("/users/{user_id}/extend", summary="延長臨時帳號的到期日")
def extend_temp_user(
    user_id: str,
    data: schemas.AdminExtendTempAccount,
    db: Session = Depends(get_db),
    admin: models.User = Depends(require_admin),
) -> Any:
    """
    ZH: 把臨時帳號的到期日往後延。

    ZH: 這支現在收的是**絕對到期日**（expires_on），不是「再加幾天」。
        擁有者 2026-08-24 裁定：建立與延期都在日曆上挑一天。

    ZH: 舊版有一個「已過期就從現在起算」的阱阱（防止「過期一個月的帳號
        延長 7 天之後仍然是過期的」）。**絕對日期讓這個問題消失了** ——
        選了一個未來的日子，結果就是那一天，跟舊值無關。

    ZH: 🔴 但這一個仍然成立，漏了就是**靜默做錯事**：
        **要把 is_active 設回 1。** 每日排程會把過期帳號標成停用；
        只改到期日而不解除停用，帳號依舊登不進來 ——
        按了沒反應、而且沒有任何錯誤訊息。

    ZH: 只對臨時帳號有意義（一般帳號沒有 expires_at）。

    @node job-scheduler/app/routers/admin.py::extend_temp_user
    """
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="ZH: 找不到這個帳號 | EN: User not found")
    if user.expires_at is None:
        raise HTTPException(
            status_code=400,
            detail="ZH: 這不是臨時帳號，沒有到期日可以延長 | EN: Not a temporary account",
        )

    now = datetime.now(timezone.utc)
    user.expires_at = data.expires_at_utc
    # ZH: 見上面 🔴 那段 —— 排程可能已經把它停用了。
    user.is_active = 1

    db.add(models.AdminAction(
        admin_id=admin.id,
        target_user=user.id,
        action="extend_temp_account",
        payload=json.dumps({
            # ZH: 稽核記錄要看得出管理者**選了哪一天**，
            #     而不是只留一個換算後的 UTC 時間戳。
            "expires_on": data.expires_on.isoformat(),
            "new_expires_at": user.expires_at.isoformat(),
        }, ensure_ascii=False),
        timestamp=now,
    ))
    db.commit()

    logger.info("延長臨時帳號 %s 至 %s by %s",
                user.username, user.expires_at.isoformat(), admin.username)
    return {"username": user.username, "expires_at": user.expires_at.isoformat()}


# ZH: 匯入時接受的身分寫法 → 平台角色。中文寫法一併收 ——
#     匯入檔多半是人用 Excel 打的，逼他寫英文代碼只會多一輪來回。
#     🔴 沒有 admin：管理員永遠手動開（打錯一格就多一個管理員不可接受）。
_IMPORT_ROLES = {
    "student": "student", "學生": "student",
    "teacher": "teacher", "老師": "teacher", "教師": "teacher",
    "staff": "staff", "職員": "staff",
    "guest": "guest", "訪客": "guest",
}

_IMPORT_HEADERS = ("帳號名稱", "帳號", "名稱", "username", "name")
_IMPORT_MAX_ROWS = 200
_EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")


def _cell_str(v) -> str:
    """ZH: Excel 儲存格轉字串。整數值的 float 去掉 `.0` ——
       純數字密碼/帳號在 xlsx 裡是數值型，直接 str() 會變 `12345678.0`。

    @node job-scheduler/app/routers/admin.py::_cell_str
    """
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _parse_import_file(filename: str, blob: bytes) -> list:
    """ZH: 把上傳的 CSV / XLSX 解成列（[username, email, password, role] 字典）。

    ZH: CSV 編碼：先試 utf-8-sig，再退 cp950 —— 台灣的 Excel 存 CSV
        預設是 Big5，不接的話中文身分欄會整批亂碼。
    ZH: 首列若是欄位名（帳號名稱/username…）自動略過，與範例檔對齊。
    ZH: 解析失敗丟 ValueError（訊息給人看），由端點包成 400。

    @node job-scheduler/app/routers/admin.py::_parse_import_file
    """
    name = (filename or "").lower()
    grid: list[list[str]] = []
    if name.endswith(".xlsx"):
        import openpyxl
        from io import BytesIO
        try:
            wb = openpyxl.load_workbook(BytesIO(blob), read_only=True, data_only=True)
        except Exception:
            raise ValueError("這個檔案不是有效的 Excel（.xlsx）")
        ws = wb.worksheets[0]
        for row in ws.iter_rows(values_only=True):
            grid.append([_cell_str(c) for c in row])
        wb.close()
    elif name.endswith(".csv"):
        try:
            text = blob.decode("utf-8-sig")
        except UnicodeDecodeError:
            try:
                text = blob.decode("cp950")
            except UnicodeDecodeError:
                raise ValueError("CSV 編碼看不懂（請用 UTF-8 或 Big5 存檔）")
        for line in text.splitlines():
            if line.strip():
                grid.append([c.strip() for c in line.split(",")])
    else:
        raise ValueError("只接受 .csv 或 .xlsx 檔")

    rows = []
    for i, cells in enumerate(grid):
        c = list(cells) + [""] * (4 - len(cells))
        if not any(x for x in c[:4]):
            continue
        if i == 0 and c[0].strip().lower() in _IMPORT_HEADERS:
            continue
        rows.append({"username": c[0], "email": c[1],
                     "password": c[2], "role": c[3] or "student"})
    return rows


@router.get("/users/temporary/import-template", summary="下載臨時帳號匯入範例檔")
def temp_import_template(
    fmt: str = Query("csv", pattern="^(csv|xlsx)$"),
    admin: models.User = Depends(require_admin),
) -> Any:
    """ZH: 範例檔（含欄位列＋兩列示範）。CSV 帶 BOM —— 台灣的 Excel
       沒有 BOM 會把 UTF-8 中文開成亂碼。

    @node job-scheduler/app/routers/admin.py::temp_import_template
    """
    from fastapi.responses import Response
    header = ["帳號名稱", "信箱(可空)", "密碼(可空=系統產生)", "身分(學生/老師/職員/訪客)"]
    demo = [["visitor01", "guest01@example.edu", "Str0ngPa55", "訪客"],
            ["visitor02", "", "", "訪客"]]
    if fmt == "xlsx":
        import openpyxl
        from io import BytesIO
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "臨時帳號"
        ws.append(header)
        for r in demo:
            ws.append(r)
        # ZH: 欄寬撐開，不然打開來全被截字。
        for col, w in zip("ABCD", (16, 28, 24, 28)):
            ws.column_dimensions[col].width = w
        buf = BytesIO()
        wb.save(buf)
        return Response(
            buf.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument"
                       ".spreadsheetml.sheet",
            headers={"Content-Disposition":
                     "attachment; filename=temp-accounts-template.xlsx"})
    csv_text = "\n".join(",".join(r) for r in [header] + demo) + "\n"
    return Response(
        csv_text.encode("utf-8-sig"),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition":
                 "attachment; filename=temp-accounts-template.csv"})


@router.post("/users/temporary/import", summary="批次匯入臨時帳號（CSV/XLSX，預覽＋建立）")
async def import_temp_users(
    request: Request,
    file: UploadFile = File(...),
    purpose: str = Form(...),
    expires_on: str = Form(...),
    dry_run: bool = Form(False),
    db: Session = Depends(get_db),
    admin: models.User = Depends(require_admin),
) -> Any:
    """
    ZH: 批次建立臨時帳號（擁有者需求 2026-09-02，改檔案匯入 2026-09-03）。
        上傳 CSV 或 XLSX：每列 名稱/信箱/密碼/身分；用途與到期日整批共用。
        `dry_run=true` 只驗證回報、不寫入。範例檔見 import-template 端點。

    ZH: 🔴 **全有或全無**：任何一列驗不過就整批不建（400 帶逐列錯誤）。
        建到一半停下來的話，管理者得自己對「哪幾個建了」——那正是
        批次匯入要消滅的工作。預覽（dry_run）讓錯誤在送出前就現形。

    ZH: 密碼規則：有給就用（≥8 字）；留空＝系統產生，**回應帶一次明文**
        （與單筆建立同一個「只顯示這一次」契約）。有給密碼的列**不回傳**
        密碼——管理者的檔案裡本來就有，多回一次只是多一份外洩面。

    ZH: 絕不寄信（臨時帳號的鐵則，理由見單筆端點）。

    @node job-scheduler/app/routers/admin.py::import_temp_users
    """
    import secrets as _secrets
    import uuid as _uuid
    from datetime import date as _date

    # ── 共用欄位驗證（multipart 拿不到 pydantic body，這裡手動對齊單筆規則）──
    purpose = (purpose or "").strip()
    if not purpose:
        raise HTTPException(status_code=400, detail="請說明這批臨時帳號的用途")
    try:
        exp_date = _date.fromisoformat((expires_on or "").strip())
        schemas._check_expires_on(exp_date)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e) or "到期日格式不對（YYYY-MM-DD）")
    expires_at = schemas.expires_on_to_utc(exp_date)

    blob = await file.read()
    try:
        raw_rows = _parse_import_file(file.filename, blob)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    if not raw_rows:
        raise HTTPException(status_code=400, detail="檔案裡沒有任何資料列")
    if len(raw_rows) > _IMPORT_MAX_ROWS:
        raise HTTPException(status_code=400,
                            detail=f"一次最多 {_IMPORT_MAX_ROWS} 列（收到 {len(raw_rows)}）")

    # ── 逐列驗證（先全部驗完，一筆都還沒寫）──────────────────────────
    report = []
    errors = 0
    seen_users = set()
    seen_emails = set()
    for i, row in enumerate(raw_rows, start=1):
        uname = row["username"].strip()
        email = row["email"].strip()
        pw = row["password"].strip()
        role_raw = row["role"].strip() or "student"
        problems = []
        if not uname:
            problems.append("帳號名稱空白")
        elif uname.lower() in seen_users:
            problems.append("檔案內帳號重複")
        elif crud.get_user_by_username(db, uname):
            problems.append("帳號已存在")
        seen_users.add(uname.lower())
        if email:
            if not _EMAIL_RE.match(email):
                problems.append("Email 格式不對")
            elif email.lower() in seen_emails:
                problems.append("檔案內 Email 重複")
            elif crud.get_user_by_email(db, email):
                problems.append("Email 已存在")
            seen_emails.add(email.lower())
        if pw and len(pw) < 8:
            problems.append("密碼太短（至少 8 字）")
        role = _IMPORT_ROLES.get(role_raw.lower() if role_raw.isascii() else role_raw)
        if role is None:
            problems.append("身分看不懂（可用：學生/老師/職員/訪客）")
        if problems:
            errors += 1
        report.append({"line": i, "username": uname, "role": role,
                       "has_email": bool(email), "will_generate_pw": not pw,
                       "errors": problems})

    if dry_run or errors:
        result = {"ok": errors == 0, "total": len(report),
                  "error_rows": errors, "rows": report}
        if errors and not dry_run:
            # ZH: 直接送建立卻驗不過 → 400，內容與預覽同構，前端好顯示。
            raise HTTPException(status_code=400, detail=result)
        return result

    # ── 建立（全數通過才走到這裡；一次交易）──────────────────────────
    created = []
    for row in raw_rows:
        uname = row["username"].strip()
        email_given = row["email"].strip()
        email = email_given or f"temp-{uname}@ai-base.invalid"
        pw_given = row["password"].strip()
        pw = pw_given or _secrets.token_urlsafe(9)
        role_raw = row["role"].strip() or "student"
        role = _IMPORT_ROLES[role_raw.lower() if role_raw.isascii() else role_raw]
        user = models.User(
            id=str(_uuid.uuid4()),
            username=uname,
            email=email,
            hashed_password=crud.get_password_hash(pw),
            role=role,
            is_active=1,
            expires_at=expires_at,
            temp_purpose=purpose,
        )
        db.add(user)
        db.flush()          # ZH: AdminAction 的外鍵需要 user.id 先落地（同單筆端點）
        db.add(models.AdminAction(
            admin_id=admin.id,
            target_user=user.id,
            action="create_temp_account",
            payload=json.dumps({
                "username": uname, "purpose": purpose,
                "expires_on": exp_date.isoformat(),
                "role": role, "batch_import": True,
            }, ensure_ascii=False),
            timestamp=datetime.now(timezone.utc),
            ip_address=(request.client.host if request.client else None),
        ))
        created.append({"username": uname, "role": role,
                        "has_email": bool(email_given),
                        # ZH: 只有系統產生的才回明文（見 docstring）。
                        "password": (None if pw_given else pw)})
    db.commit()

    logger.info("批次匯入臨時帳號 %d 個（到期 %s，用途：%s，檔案 %s）by %s",
                len(created), exp_date.isoformat(), purpose,
                file.filename, admin.username)
    return {"ok": True, "total": len(created),
            "expires_at": expires_at.isoformat(), "created": created}


@router.post("/users/temporary", summary="建立臨時帳號")
def create_temp_user(
    data: schemas.AdminTempUserCreate,
    request: Request,
    db: Session = Depends(get_db),
    admin: models.User = Depends(require_admin),
) -> Any:
    """
    ZH: 建立有到期日的臨時帳號（校外人士、長官視察、例外用途）。

    ZH: 與 `/users/provision` 的三個差別，每一個都有理由：
        1. **email 可以不填** —— 校外人士多半沒有學校信箱
        2. **絕不寄信** —— 填假信箱會真的寄出去然後退信
        3. **密碼直接回傳** —— 沒有信可寄，只能當面交給對方

    ZH: 🔴 `users.email` 是 NOT NULL + UNIQUE，改成可空要重建整張表。
        所以沒填 email 時合成一個 RFC 2606 保留網域的位址（`.invalid`）——
        那個網域**永遠不可能存在**，而 `send_email` 早就有保留網域閘門擋著。
        管理端看到 `.invalid` 會顯示成「無信箱」而不是把假位址秀出來。

    ZH: **不用 is_test_account** —— 帶那個旗標的帳號每次服務重啟就被刪掉。

    @node job-scheduler/app/routers/admin.py::create_temp_user
    """
    if crud.get_user_by_username(db, data.username):
        raise HTTPException(status_code=400,
                            detail="ZH: 這個帳號名稱已經有人用了 | EN: Username already exists")
    if data.email and crud.get_user_by_email(db, data.email):
        raise HTTPException(status_code=400,
                            detail="ZH: 這個 Email 已經有人用了 | EN: Email already exists")

    import secrets as _secrets
    import uuid as _uuid
    temp_password = _secrets.token_urlsafe(9)

    # ZH: 見上面的說明 —— 合成位址只是為了滿足 NOT NULL + UNIQUE，
    #     `.invalid` 是 RFC 2606 保留、永遠不會存在的網域。
    email = data.email or f"temp-{_uuid.uuid4().hex[:12]}@invalid"

    # ZH: 轉換在 schemas.expires_on_to_utc 裡只做一次（台灣時間當天 23:59:59）。
    expires_at = data.expires_at_utc

    user = models.User(
        id=str(_uuid.uuid4()),
        username=data.username,
        email=email,
        hashed_password=crud.get_password_hash(temp_password),
        role=data.role or "student",
        department=data.department,
        is_active=1,
        expires_at=expires_at,
        temp_purpose=data.purpose,
    )
    db.add(user)
    # ZH: 🔴 一定要先 flush。`admin_actions.target_user` 有指向 `users.id` 的外鍵，
    #     而同一個交易裡 User 還沒寫出去 —— 直接 add 稽核列會 FOREIGN KEY constraint failed。
    db.flush()

    # ZH: 稽核 —— 「誰、什麼時候、為了什麼開了這個帳號」要留得住。
    db.add(models.AdminAction(
        admin_id=admin.id,
        target_user=user.id,
        action="create_temp_account",
        payload=json.dumps({
            "username": data.username,
            "purpose": data.purpose,
            "expires_on": data.expires_on.isoformat(),
            "role": user.role,
            "expires_at": expires_at.isoformat(),
        }, ensure_ascii=False),
        timestamp=datetime.now(timezone.utc),
        ip_address=(request.client.host if request.client else None),
    ))
    db.commit()

    logger.info("建立臨時帳號 %s（到期 %s，用途：%s）by %s",
                data.username, data.expires_on.isoformat(), data.purpose, admin.username)

    return {
        "id": user.id,
        "username": user.username,
        # ZH: 只有這一次拿得到明文 —— 畫面要提醒管理者現在就抄走。
        "password": temp_password,
        "expires_at": expires_at.isoformat(),
        "purpose": data.purpose,
        "role": user.role,
        # ZH: 讓前端知道要不要顯示信箱（合成的那個不該給人看）
        "has_email": bool(data.email),
    }


@router.post("/users/provision")
def provision_user(
    data: schemas.AdminProvisionUser,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    _: models.User = Depends(require_admin),
) -> Any:
    """ZH: 管理員配發新帳號（預先建立，待 SSO 接管）| EN: Admin provision a new user account

    @node job-scheduler/app/routers/admin.py::provision_user
    """
    if crud.get_user_by_username(db, data.username):
        raise HTTPException(status_code=400, detail="ZH: 這個帳號名稱已經有人用了 | EN: Username already exists")
    if crud.get_user_by_email(db, data.email):
        raise HTTPException(status_code=400, detail="ZH: 這個 Email 已經有人用了 | EN: Email already exists")

    import secrets
    temp_password = data.password if data.password else secrets.token_urlsafe(12)

    user_create = schemas.UserCreate(
        username=data.username,
        email=data.email,
        password=temp_password,
        role=data.role or "student",
    )
    db_user = crud.create_user(db, user_create)
    db_user.is_test_account = 0
    db.commit()

    email_queued = bool(crud.mail_address_for(db_user))
    if email_queued:
        background_tasks.add_task(
            email_service.send_temp_password,
            crud.mail_address_for(db_user), db_user.username, temp_password, True,
        )

    # ZH: 僅在無法發送 Email 時才在回應中回傳明文密碼，避免密碼出現在瀏覽器記錄中
    # EN: Only return plaintext password when email cannot be sent (avoids it appearing in browser logs)
    logger.info(
        "provision_user: created %s (email_queued=%s)", db_user.username, email_queued
    )
    return {
        "id": db_user.id,
        "username": db_user.username,
        "email": db_user.email,
        "role": db_user.role,
        "temp_password": temp_password if not email_queued else "[已寄送至 Email | sent via email]",
        "email_sent": email_queued,
    }


@router.post("/users/{user_id}/reset")
def reset_user_account(
    user_id: str,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin),
) -> Any:
    """ZH: 初始化帳號 — 重置密碼 + 歸零 Token 用量 | EN: Reset password and clear token usage

    @node job-scheduler/app/routers/admin.py::reset_user_account
    """
    db_user = db.query(models.User).filter(models.User.id == user_id).first()
    if not db_user:
        raise HTTPException(status_code=404, detail="ZH: 找不到這個帳號 | EN: User not found")

    import secrets
    temp_password = secrets.token_urlsafe(12)
    db_user.hashed_password = crud.get_password_hash(temp_password)

    usage = crud.get_token_usage(db, user_id)
    if usage:
        usage.tokens_used = 0

    db.commit()

    email_queued = bool(crud.mail_address_for(db_user))
    if email_queued:
        background_tasks.add_task(
            email_service.send_temp_password,
            crud.mail_address_for(db_user), db_user.username, temp_password, False,
        )

    logger.info(
        "reset_user_account: %s reset by admin %s (email_queued=%s)",
        db_user.username, current_user.username, email_queued,
    )
    return {
        "username": db_user.username,
        "temp_password": temp_password if not email_queued else "[已寄送至 Email | sent via email]",
        "email_sent": email_queued,
        "message": f"Account {db_user.username} has been initialized",
    }


# ==============================================================================
# ZH: 任務管理 | EN: Job Management
# ==============================================================================

@router.get("/jobs", response_model=list[schemas.AdminJobListItem])
def get_all_jobs(
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=500),
    db: Session = Depends(get_db),
    _: models.User = Depends(require_admin),
) -> Any:
    """ZH: 列出所有任務，支援分頁 | EN: List all jobs with pagination

    @node job-scheduler/app/routers/admin.py::get_all_jobs
    """
    jobs = (
        db.query(models.TrainingJob)
        .order_by(models.TrainingJob.created_at.desc())
        .offset(skip)
        .limit(limit)
        .all()
    )
    return [
        schemas.AdminJobListItem(
            job_id=j.id,
            job_name=j.job_name,
            model_name=j.model_name,
            user_id=j.user_id,
            status=j.status,
            priority=j.priority,
            progress=j.progress,
            gpu_server=j.gpu_server,
            created_at=j.created_at,
            started_at=j.started_at,
            completed_at=j.completed_at,
            error_message=j.error_message,
        )
        for j in jobs
    ]


@router.post("/jobs/{job_id}/cancel")
def admin_cancel_job(
    job_id: str,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin),
) -> Any:
    """ZH: 管理員強制取消任務 | EN: Admin force-cancel a job

    @node job-scheduler/app/routers/admin.py::admin_cancel_job
    """
    job = db.query(models.TrainingJob).filter(models.TrainingJob.id == job_id).first()
    if not job:
        raise HTTPException(status_code=404, detail="ZH: 找不到這張任務 | EN: Job not found")
    if job.status not in ("pending", "queued"):
        raise HTTPException(
            status_code=400,
            detail=f"ZH: 這張任務的狀態是「{job.status}」，只有還在排隊的才能取消 | EN: Cannot cancel job with status '{job.status}'. Only pending/queued jobs can be cancelled.",
        )

    job.status = "cancelled"
    db.commit()
    db.refresh(job)

    logger.info(f"Admin {current_user.username} cancelled job {job_id[:8]}")
    return {"job_id": job.id, "status": job.status, "message": "Job cancelled"}


@router.put("/jobs/{job_id}/priority")
def admin_update_job_priority(
    job_id: str,
    data: schemas.AdminJobPriority,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin),
) -> Any:
    """ZH: 管理員修改任務優先級 | EN: Admin update job priority

    @node job-scheduler/app/routers/admin.py::admin_update_job_priority
    """
    job = db.query(models.TrainingJob).filter(models.TrainingJob.id == job_id).first()
    if not job:
        raise HTTPException(status_code=404, detail="ZH: 找不到這張任務 | EN: Job not found")
    if job.status not in ("pending", "queued"):
        raise HTTPException(
            status_code=400,
            detail=f"ZH: 這張任務的狀態是「{job.status}」，已經不能調整優先序 | EN: Cannot reprioritize job with status '{job.status}'.",
        )

    old_priority = job.priority
    job.priority = data.priority
    db.commit()
    db.refresh(job)

    logger.info(f"Admin {current_user.username} changed job {job_id[:8]} priority: {old_priority} -> {data.priority}")
    return {"job_id": job.id, "priority": job.priority, "old_priority": old_priority}


# ==============================================================================
# ZH: 模型管理 | EN: Model Management
# ==============================================================================

@router.get("/models")
def get_all_models(
    db: Session = Depends(get_db),
    _: models.User = Depends(require_admin),
) -> Any:
    """ZH: 列出所有模型 | EN: List all models

    @node job-scheduler/app/routers/admin.py::get_all_models
    """
    mdls = db.query(models.Model).order_by(models.Model.created_at.desc()).all()
    return [
        {
            "id": m.id, "name": m.name, "model_type": m.model_type,
            "description": m.description, "framework": m.framework,
            "storage_path": m.storage_path, "size_bytes": m.size_bytes,
            "uploaded_by": m.uploaded_by, "is_public": m.is_public,
            "tool_types": m.tool_types or "chat",
            "api_provider": m.api_provider, "api_endpoint": m.api_endpoint,
            "api_model_id": m.api_model_id, "created_at": m.created_at,
        }
        for m in mdls
    ]


@router.post("/models")
def admin_create_model(
    data: schemas.AdminModelCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin),
) -> Any:
    """ZH: 管理員新增模型 | EN: Admin create model

    @node job-scheduler/app/routers/admin.py::admin_create_model
    """
    if db.query(models.Model).filter(models.Model.name == data.name).first():
        raise HTTPException(status_code=400, detail=f"ZH: 已經有一個叫「{data.name}」的模型了 | EN: Model '{data.name}' already exists")

    new_model = models.Model(
        name=data.name, model_type=data.model_type or "local",
        description=data.description, framework=data.framework,
        storage_path=data.storage_path or "", is_public=data.is_public or 0,
        tool_types=data.tool_types or "chat",
        uploaded_by=current_user.id, api_provider=data.api_provider,
        api_endpoint=data.api_endpoint, api_model_id=data.api_model_id,
    )
    db.add(new_model)
    db.commit()
    db.refresh(new_model)

    logger.info(f"Admin {current_user.username} created model '{data.name}'")
    return {"id": new_model.id, "name": new_model.name, "model_type": new_model.model_type}


@router.put("/models/{model_id}")
def admin_update_model(
    model_id: str,
    data: schemas.AdminModelUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin),
) -> Any:
    """ZH: 管理員更新模型資訊 | EN: Admin update model info

    @node job-scheduler/app/routers/admin.py::admin_update_model
    """
    mdl = db.query(models.Model).filter(models.Model.id == model_id).first()
    if not mdl:
        raise HTTPException(status_code=404, detail="ZH: 找不到這個模型 | EN: Model not found")

    if data.name is not None:
        dup = db.query(models.Model).filter(
            models.Model.name == data.name, models.Model.id != model_id
        ).first()
        if dup:
            raise HTTPException(status_code=400, detail=f"ZH: 「{data.name}」這個模型名稱已經有人用了 | EN: Model name '{data.name}' already taken")
        mdl.name = data.name

    for field in ("description", "model_type", "framework", "storage_path",
                  "is_public", "tool_types", "api_provider", "api_endpoint", "api_model_id"):
        val = getattr(data, field, None)
        if val is not None:
            setattr(mdl, field, val)

    db.commit()
    db.refresh(mdl)

    logger.info(f"Admin {current_user.username} updated model '{mdl.name}'")
    return {"id": mdl.id, "name": mdl.name, "model_type": mdl.model_type}


@router.delete("/models/{model_id}")
def admin_delete_model(
    model_id: str,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin),
) -> Any:
    """ZH: 管理員刪除模型 | EN: Admin delete model

    @node job-scheduler/app/routers/admin.py::admin_delete_model
    """
    mdl = db.query(models.Model).filter(models.Model.id == model_id).first()
    if not mdl:
        raise HTTPException(status_code=404, detail="ZH: 找不到這個模型 | EN: Model not found")

    model_name = mdl.name
    db.delete(mdl)
    db.commit()

    logger.info(f"Admin {current_user.username} deleted model '{model_name}'")
    return {"message": f"Model '{model_name}' deleted", "deleted_id": model_id}


# ==============================================================================
# ZH: 叢集狀態 | EN: Cluster Status
# ==============================================================================

@router.get("/cluster/stats")
def get_cluster_stats(
    db: Session = Depends(get_db),
    _: models.User = Depends(require_admin),
) -> Any:
    """ZH: 取得各 GPU 即時狀態（每張卡一筆，供前端 cluster 卡片）
       EN: Per-GPU live stats (one entry per card) for the admin cluster panel.

    @node job-scheduler/app/routers/admin.py::get_cluster_stats
    """
    nodes = db.query(models.WorkerHeartbeat).order_by(models.WorkerHeartbeat.last_seen.desc()).all()
    out: list[dict] = []
    for n in nodes:
        status = "online" if n.is_online else "offline"
        try:
            detail = json.loads(n.gpus_detail or "[]")
        except Exception:
            detail = []
        if detail:
            for g in detail:
                out.append({
                    "node_id": n.node_id,
                    "gpu_id": g.get("gpu_id", "-"),
                    "name": g.get("name", "GPU"),
                    "utilization": g.get("utilization", 0),
                    "temperature": g.get("temperature", 0),
                    "memory_used": g.get("memory_used", 0),
                    "memory_total": g.get("memory_total", 0),
                    "last_seen": n.last_seen,
                    "status": status,
                })
        else:
            # ZH: 舊版 worker（無 per-GPU 詳細）→ 退回單一聚合卡，避免顯示 undefined
            # EN: Old worker (no per-GPU detail) → single aggregate card so the UI isn't "undefined"
            out.append({
                "node_id": n.node_id,
                "gpu_id": "-",
                "name": n.node_id,
                "utilization": n.gpu_utilization or 0,
                "temperature": 0,
                "memory_used": 0,
                "memory_total": 0,
                "last_seen": n.last_seen,
                "status": status,
            })
    return out


# ==============================================================================
# ZH: 數據分析 | EN: Analytics
# ==============================================================================

@router.get("/users/{user_id}/analytics")
def get_user_analytics(
    user_id: str,
    db: Session = Depends(get_db),
    _: models.User = Depends(require_admin),
) -> Any:
    """
    ZH: 取得指定使用者的細粒度數據分析（Token、Sessions、工具分布）
    EN: Detailed per-user analytics — token quota, sessions, tool breakdown

    @node job-scheduler/app/routers/admin.py::get_user_analytics
    """
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="ZH: 找不到這個帳號 | EN: User not found")

    usage = crud.get_token_usage(db, user_id)

    # ZH: 依工具類型彙總訊息數與 Token 消耗 | EN: Aggregate by tool_type
    tool_rows = (
        db.query(
            models.ChatHistory.tool_type,
            func.count(models.ChatHistory.id).label("message_count"),
            func.sum(models.ChatHistory.tokens_used).label("tokens_sum"),
        )
        .filter(models.ChatHistory.user_id == user_id)
        .group_by(models.ChatHistory.tool_type)
        .all()
    )

    # ZH: Top-10 Sessions（依 Token 消耗降冪）| EN: Top-10 sessions by token cost
    session_rows = (
        db.query(
            models.ChatHistory.session_id,
            func.min(models.ChatHistory.created_at).label("started_at"),
            func.count(models.ChatHistory.id).label("message_count"),
            func.sum(models.ChatHistory.tokens_used).label("tokens_sum"),
        )
        .filter(models.ChatHistory.user_id == user_id)
        .group_by(models.ChatHistory.session_id)
        .order_by(func.sum(models.ChatHistory.tokens_used).desc())
        .limit(10)
        .all()
    )

    # ZH: 對話 Session 總數 | EN: Total distinct sessions
    total_sessions = (
        db.query(func.count(func.distinct(models.ChatHistory.session_id)))
        .filter(models.ChatHistory.user_id == user_id)
        .scalar()
    ) or 0

    tokens_used  = usage.tokens_used  if usage else 0
    tokens_limit = usage.tokens_limit if usage else 0
    usage_pct    = round(tokens_used / tokens_limit * 100, 1) if tokens_limit > 0 else 0.0

    return {
        "user": {
            "id":                   user.id,
            "username":             user.username,
            "email":                user.email,
            "role":                 user.role,
            "department":           user.department,
            "is_active":            user.is_active,
            "login_count":          user.login_count,
            "lifetime_tokens_used": user.lifetime_tokens_used,
            "last_login_time":      user.last_login_time,
            "last_login_ip":        user.last_login_ip,
            "created_at":           user.created_at,
        },
        "token_quota": {
            "tokens_used":  tokens_used,
            "tokens_limit": tokens_limit,
            "usage_pct":    usage_pct,
            "reset_date":   usage.reset_date if usage else None,
        },
        "total_sessions": total_sessions,
        "tool_breakdown": [
            {
                "tool_type":     r.tool_type or "chat",
                "message_count": r.message_count,
                "tokens_sum":    int(r.tokens_sum or 0),
            }
            for r in tool_rows
        ],
        "top_sessions": [
            {
                "session_id":    r.session_id,
                "started_at":    r.started_at,
                "message_count": r.message_count,
                "tokens_sum":    int(r.tokens_sum or 0),
            }
            for r in session_rows
        ],
    }


def _analytics_day(v):
    """ZH: "YYYY-MM-DD" → date；空的或格式不對回 None（不要拋，回退成「不篩」）。

    ZH: ⚠ 格式錯就當成沒給，不要回 400 —— 這個參數來自網址列，
        手打錯一個字就整頁壞掉太脆。看得到全部資料，比看到一個錯誤畫面好。
    """
    if not v:
        return None
    try:
        return ddate.fromisoformat(str(v)[:10])
    except ValueError:
        return None


@router.get("/analytics")
def get_analytics(
    group_by: str = Query("department", description="department / college / unit"),
    department: Optional[str] = Query("all"),
    days: int = Query(30, ge=0, le=3650, description="ZH: 0 = 全部 | EN: 0 = all time"),
    start: Optional[str] = Query(None, description="YYYY-MM-DD"),
    end: Optional[str] = Query(None, description="YYYY-MM-DD"),
    db: Session = Depends(get_db),
    _: models.User = Depends(require_admin),
) -> Any:
    """
    ZH: 平台使用統計。v3.8 起可以選**依學院／學系／行政單位**分組（#13）。

    ZH: 學院**不是 users 的欄位** —— 由 `department` 外連 `org_departments` 推出來。
        所以改對照表就會全站生效,不必回填幾千筆使用者。

    ZH: 🔴 外連（outerjoin）不是內連：對不到對照表的學系**必須留下來**,
        用內連會讓那些人從統計裡**安靜地消失** —— 人數對不上,而畫面上看不出少了誰。
        對不到的一律回 None,由前端顯示成「未分類」。

    ZH: 行政單位只有職員會有值,所以依單位分組時,學生那一大群會全部落在「未分類」。
        那是正確的,不是 bug。

    @node job-scheduler/app/routers/admin.py::get_analytics
    """
    if group_by not in ("department", "college", "unit", "campus"):
        raise HTTPException(status_code=400,
                            detail=f"不支援的分組方式：{group_by}"
                                   "（可用：department / college / unit / campus）")

    # ZH: v4.5 依校區（擁有者需求 2026-09-03）。校區是**關聯表**
    #     （user_campuses，教職員可多校區）——直接 join 會把跨校區的人
    #     算兩次，總數對不上、圓餅也切不完整。所以先彙總成每人一列：
    #     單校區歸該校區、**多校區自成一桶**、沒設＝未分類（None）。
    camp_sub = None
    if group_by == "campus":
        camp_sub = (db.query(
            models.UserCampus.user_id.label("uid"),
            func.count(models.UserCampus.campus).label("n"),
            func.min(models.UserCampus.campus).label("one"),
        ).group_by(models.UserCampus.user_id).subquery())
        label_col = case((camp_sub.c.n > 1, "多校區"), else_=camp_sub.c.one)
    elif group_by == "college":
        label_col = models.OrgDepartment.college
    elif group_by == "unit":
        label_col = models.User.unit
    else:
        label_col = models.User.department

    base_q = db.query(
        label_col.label("grp"),
        func.count(models.User.id).label("user_count"),
        func.sum(models.User.login_count).label("total_logins"),
        func.sum(models.User.lifetime_tokens_used).label("total_tokens"),
    )
    if group_by == "college":
        base_q = base_q.outerjoin(
            models.OrgDepartment,
            models.User.department == models.OrgDepartment.name)
    elif group_by == "campus":
        base_q = base_q.outerjoin(
            camp_sub, models.User.id == camp_sub.c.uid)

    # ZH: `department` 這個篩選沿用舊參數名（前端與匯出都在用）——
    #     它篩的一律是**學系**，與 group_by 無關：先篩人，再看要怎麼分組。
    if department != "all":
        base_q = base_q.filter(models.User.department == department)
    rows = base_q.group_by(label_col).all()

    # ZH: v3.9 分組名的英文（擁有者裁定 2026-08-30）。
    # ZH: ⚠ **只做學系與學院**。行政單位的英文名只有 53/97，做了會讓同一欄
    #     一半英文一半中文 —— 比全中文更難讀。等補齊再把 unit 加進來。
    group_en = {}
    if group_by == "department":
        group_en = {d.name: d.name_en
                    for d in db.query(models.OrgDepartment).all() if d.name_en}
    elif group_by == "college":
        group_en = {d.college: d.college_en
                    for d in db.query(models.OrgDepartment).all() if d.college_en}
    elif group_by == "campus":
        from .. import org_seed
        group_en = dict(org_seed.CAMPUS_EN)
        group_en["多校區"] = "Multi-campus"

    group_stats = [
        {
            "group": r.grp,                    # ZH: None = 未分類，文案由前端決定
            # ZH: 對不到就是 None，前端退回中文（unit 一律是 None，見上面）。
            "group_en": group_en.get(r.grp),
            "user_count": r.user_count,
            "total_logins": r.total_logins or 0,
            "total_tokens": r.total_tokens or 0,
        }
        for r in rows
    ]

    # ══════════════════════════════════════════════════════════════════
    # ZH: v3.9 各分類的實際使用（擁有者裁定 2026-08-30）
    #
    # ZH: 上面那三個數字回答的是「有沒有人」與「有沒有登入」。
    #     下面這些回答的是**「有沒有在用」** —— 那是完全不同的問題，
    #     而後者才決定要去哪個系推廣。
    #
    # ZH: 🔴 每一項都由 user_id 接回 users.department，再依 group_by 分組。
    #     所以改組織對照表就會全站生效，不必回填任何一筆使用紀錄。
    # ══════════════════════════════════════════════════════════════════
    d_start, d_end = _analytics_day(start), _analytics_day(end)
    if d_start and d_end and d_start > d_end:
        # ZH: 選反了就幫他換過來，不要回一張空表（與 MYAI 那支同一個做法）
        d_start, d_end = d_end, d_start

    def _win(q, col, *, naive_vendor_time: bool = False):
        """ZH: 套上期間。

        ZH: 🔴 `naive_vendor_time` 是給 myai_transactions 用的 ——
            那張表的 occurred_at 是**廠商當地時間**不是 UTC。
            用 UTC 的 now() 去比會差 8 小時，日界附近的筆數就會跑掉。
        """
        now = datetime.now() if naive_vendor_time else datetime.now(timezone.utc)
        if d_start or d_end:
            if d_start:
                q = q.filter(col >= datetime.combine(d_start, dtime.min))
            if d_end:
                q = q.filter(col <= datetime.combine(d_end, dtime.max))
        elif days > 0:
            q = q.filter(col >= now - timedelta(days=days))
        return q

    def _by_group(q, join_user_on):
        """ZH: 把任何一張帶 user_id 的表接回分組欄位。"""
        q = q.join(models.User, join_user_on == models.User.id)
        if group_by == "college":
            q = q.outerjoin(models.OrgDepartment,
                            models.User.department == models.OrgDepartment.name)
        elif group_by == "campus":
            q = q.outerjoin(camp_sub, models.User.id == camp_sub.c.uid)
        if department != "all":
            q = q.filter(models.User.department == department)
        return q.group_by(label_col)

    def _counts(model, user_col, time_col, *, extra=None, naive=False):
        """ZH: {分組 -> 筆數}。統一走這一支，四張表就不會各寫一份 join。"""
        # ZH: 🔴 `select_from` 不能省。第一個 query 欄位是 users.department，
        #     SQLAlchemy 會據此把 `users` 當成 FROM，接著 _by_group 又 join 一次
        #     users —— 於是 `ambiguous column name: users.department`。
        #     明確講出「從這張事件表出發」就不會猜。
        q = (db.query(label_col.label("grp"), func.count(model.id).label("n"))
             .select_from(model))
        q = _by_group(q, user_col)
        q = _win(q, time_col, naive_vendor_time=naive)
        if extra is not None:
            q = q.filter(extra)
        return {r.grp: r.n for r in q.all()}

    def _actives(model, user_col, time_col, *, naive=False):
        """ZH: {分組 -> 有幾個**不同的人**}。滲透率的分子。"""
        q = (db.query(label_col.label("grp"),
                      func.count(func.distinct(user_col)).label("n"))
             .select_from(model))
        q = _by_group(q, user_col)
        q = _win(q, time_col, naive_vendor_time=naive)
        return {r.grp: r.n for r in q.all()}

    visits = _counts(models.MyaiVisit, models.MyaiVisit.user_id,
                     models.MyaiVisit.occurred_at)
    jobs = _counts(models.TrainingJob, models.TrainingJob.user_id,
                   models.TrainingJob.created_at)
    lab_gpu = _counts(models.LabUsageLog, models.LabUsageLog.user_id,
                      models.LabUsageLog.started_at,
                      extra=(models.LabUsageLog.used_gpu == 1))
    lab_cpu = _counts(models.LabUsageLog, models.LabUsageLog.user_id,
                      models.LabUsageLog.started_at,
                      extra=(models.LabUsageLog.used_gpu == 0))

    # ZH: MYAI 的交易接不到 user_id —— 要繞 external_ai_accounts 的綁定表
    #     （vendor_sn 對得起來的才算）。沒綁定的廠商帳號算不進任何分類，
    #     那是**事實**不是漏算：平台不知道那個帳號是誰的。
    tx_q = (db.query(label_col.label("grp"),
                     func.count(models.MyaiTransaction.id).label("n"),
                     # ZH: 🔴 **在 SQL 裡就只加負的**。先全部加總再取負值是錯的：
                     #     `-500 + 9000` 會變成 +8500，取 min(0, …) 之後是 0 ——
                     #     於是「有用又被補過點的系」顯示成完全沒用。
                     #     第一版就是這樣寫的，測試抓到了。
                     func.sum(case(
                         (models.MyaiTransaction.points_delta < 0,
                          -models.MyaiTransaction.points_delta),
                         else_=0)).label("pts"))
            .select_from(models.MyaiTransaction)
            .join(models.ExternalAiAccount,
                  models.MyaiTransaction.vendor_sn == models.ExternalAiAccount.myai_vendor_sn))
    tx_q = _by_group(tx_q, models.ExternalAiAccount.user_id)
    tx_q = _win(tx_q, models.MyaiTransaction.occurred_at, naive_vendor_time=True)
    tx_rows = {r.grp: (r.n, r.pts or 0) for r in tx_q.all()}

    myai_actives_q = (db.query(label_col.label("grp"),
                               func.count(func.distinct(models.ExternalAiAccount.user_id)).label("n"))
                      .select_from(models.MyaiTransaction)
                      .join(models.ExternalAiAccount,
                            models.MyaiTransaction.vendor_sn == models.ExternalAiAccount.myai_vendor_sn))
    myai_actives_q = _by_group(myai_actives_q, models.ExternalAiAccount.user_id)
    myai_actives_q = _win(myai_actives_q, models.MyaiTransaction.occurred_at,
                          naive_vendor_time=True)
    myai_actives = {r.grp: r.n for r in myai_actives_q.all()}

    visit_actives = _actives(models.MyaiVisit, models.MyaiVisit.user_id,
                             models.MyaiVisit.occurred_at)
    job_actives = _actives(models.TrainingJob, models.TrainingJob.user_id,
                           models.TrainingJob.created_at)
    lab_actives = _actives(models.LabUsageLog, models.LabUsageLog.user_id,
                           models.LabUsageLog.started_at)

    for g in group_stats:
        k = g["group"]
        _n_tx, pts = tx_rows.get(k, (0, 0))
        # ZH: 交易**筆數**刻意不送。它與點數高度相關（用得多筆數就多），
        #     而畫面上已經有九欄 —— 多一欄只增加寬度不增加資訊。
        #     要的話一行就加得回來，但送出去卻沒有人讀等於死欄位。
        # ZH: pts 已經是「用掉的點數」（正數）—— 只加負的 delta，見上面的 case。
        #     加點（管理員補點）不是使用量，混進來會讓「補過點的系」
        #     看起來用得特別多，而那正好是用得少所以要補的那些系。
        g["myai_points"] = int(pts or 0)
        g["myai_visits"] = visits.get(k, 0)
        g["jobs"] = jobs.get(k, 0)
        g["lab_gpu"] = lab_gpu.get(k, 0)
        g["lab_cpu"] = lab_cpu.get(k, 0)
        # ZH: 🔴 滲透率的分子是「**至少做過一件事**的人數」，不是各項相加 ——
        #     同一個人既跳轉又開實驗室，相加會把他算兩次，
        #     於是「有用的人」可能超過「總人數」。
        #     ⚠ 這裡取各項的最大值當**下限**（無法在 SQL 端跨表 distinct union
        #     而不引入一個很貴的子查詢）。所以這個數字是保守的：
        #     真實的活躍人數 ≥ 這個值。欄位名用 active_users_min 講明它是下限。
        g["active_users_min"] = max(
            myai_actives.get(k, 0), visit_actives.get(k, 0),
            job_actives.get(k, 0), lab_actives.get(k, 0),
        )

    # ZH: 🔴 兩種比例的分母不同，前端不要自己算 —— 算錯了看不出來。
    #       share_*  = 這一組佔全平台的幾 %（大系一定贏）
    #       adoption = 這一組有多少 % 的人在用（滲透率，看得出哪裡還沒推開）
    # ZH: ⚠ adoption 的分母是「**平台上**這一組的人數」，不是全系人數。
    #     SSO 是第一次登入才建帳號，所以完全沒來過的人不在分母裡 ——
    #     這個數字會**高估**滲透率，文案上要講明。
    def _share(v, total):
        return round(v * 100.0 / total, 1) if total else 0.0

    tot_points = sum(g["myai_points"] for g in group_stats)
    tot_visits = sum(g["myai_visits"] for g in group_stats)
    tot_jobs = sum(g["jobs"] for g in group_stats)
    for g in group_stats:
        g["share_points"] = _share(g["myai_points"], tot_points)
        g["share_visits"] = _share(g["myai_visits"], tot_visits)
        g["share_jobs"] = _share(g["jobs"], tot_jobs)
        # ZH: Lab 的佔比不送 —— GPU 與 CPU 分兩欄顯示，掛一個合計的百分比
        #     在其中任一欄都會被讀成「那一欄自己的佔比」。
        g["adoption"] = _share(g["active_users_min"], g["user_count"])

    # ZH: 這兩張表是 v3.9 才開始記的，**沒有歷史**。選的期間比第一筆還早時，
    #     前端要能講出「自 X 日起才有資料」—— 不講的話那兩欄的 0 會被當成
    #     「這個系都沒在用」，而其實是「那時候還沒開始記」。
    # ZH: 日期由資料本身推，不寫死 —— 寫死的話換一台機器部署就錯了。
    tracking_since = {
        "myai_visits": db.query(func.min(models.MyaiVisit.occurred_at)).scalar(),
        "lab_usage": db.query(func.min(models.LabUsageLog.started_at)).scalar(),
    }

    tool_q = db.query(
        models.ChatHistory.tool_type,
        func.count(models.ChatHistory.id).label("usage_count"),
    )
    if department != "all":
        tool_q = tool_q.join(models.User, models.ChatHistory.user_id == models.User.id).filter(
            models.User.department == department
        )
    tool_stats = tool_q.group_by(models.ChatHistory.tool_type).all()

    return {
        "group_by": group_by,
        "department_filter": department,
        "period": {"days": days, "start": start, "end": end},
        # ZH: 兩張新表的第一筆時間（沒有資料就是 None）。見上面的說明。
        "tracking_since": {
            k: (v.isoformat() if v is not None else None)
            for k, v in tracking_since.items()
        },
        "group_stats": group_stats,
        "tools_breakdown": [
            {"tool_type": s.tool_type or "chat", "usage_count": s.usage_count}
            for s in tool_stats
        ],
    }


# ==============================================================================
# ZH: v2.0 Lab 模組 — 13 個 admin endpoints
# EN: v2.0 Lab module — 13 admin endpoints
# ==============================================================================

from pydantic import BaseModel, Field
from ..services import quota_service, storage_lifecycle, lab_manager, secrets_service


# ---- pydantic 請求 / 回應模型 ----

class QuotaGrantRequest(BaseModel):
    user_id: str
    extra_quota_gb: int = Field(..., gt=0)
    reason: str = Field(..., min_length=5)
    expires_at: Optional[datetime] = None


class StorageStateActionRequest(BaseModel):
    user_id: str
    reason: Optional[str] = None
    admin_password: Optional[str] = None  # 永久刪除需驗證


# ---- 配額提權 ----

@router.post("/quota/grant")
def grant_quota(
    payload: QuotaGrantRequest,
    db: Session = Depends(get_db),
    admin: models.User = Depends(require_admin),
) -> Any:
    """ZH: 為使用者額外提權配額（保留歷史） | EN: Grant extra disk quota to user

    @node job-scheduler/app/routers/admin.py::grant_quota
    """
    target = db.query(models.User).filter(models.User.id == payload.user_id).first()
    if not target:
        raise HTTPException(404, "Target user not found")
    grant = quota_service.grant_quota(
        db,
        user_id=payload.user_id,
        extra_quota_gb=payload.extra_quota_gb,
        granted_by=admin.id,
        reason=payload.reason,
        expires_at=payload.expires_at,
    )
    return {"id": grant.id, "extra_quota_gb": grant.extra_quota_gb, "granted_at": grant.granted_at}


@router.delete("/quota/grant/{grant_id}")
def revoke_quota(
    grant_id: str,
    db: Session = Depends(get_db),
    admin: models.User = Depends(require_admin),
) -> Any:
    """ZH: 撤銷一筆配額提權 | EN: Revoke a quota grant

    @node job-scheduler/app/routers/admin.py::revoke_quota
    """
    success = quota_service.revoke_quota(db, grant_id=grant_id, revoked_by=admin.id)
    if not success:
        raise HTTPException(404, "Grant not found or already revoked")
    return {"status": "revoked", "grant_id": grant_id}


@router.get("/quota/{user_id}")
def get_user_quota(
    user_id: str,
    db: Session = Depends(get_db),
    _: models.User = Depends(require_admin),
) -> Any:
    """ZH: 查看使用者目前生效配額與所有提權紀錄 | EN: View user effective quota + all grants

    @node job-scheduler/app/routers/admin.py::get_user_quota
    """
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(404, "User not found")
    grants = quota_service.list_grants(db, user_id=user_id)
    return {
        "user_id": user_id,
        "base_quota_gb": user.disk_quota_gb,
        "effective_quota_gb": quota_service.get_effective_quota_gb(db, user_id),
        # ZH: v3.9 實際用量。**每日 03:00 掃描時更新**，不是即時值 ——
        #     即時量的話這支端點要跑 docker df，管理者點一個人就等好幾秒。
        # ZH: ⚠ 回 `None` 代表「從沒量到過」（沒開過 Lab，或量測失敗），
        #     不是 0 —— 前端要分得出來，否則「沒量到」看起來像「沒在用」。
        "used_gb": (lambda st: st.current_size_gb if st else None)(
            db.query(models.UserStorageState)
            .filter(models.UserStorageState.user_id == user_id).first()),
        "grants": [
            {
                "id": g.id,
                "extra_quota_gb": g.extra_quota_gb,
                "reason": g.reason,
                "granted_by": g.granted_by,
                "granted_at": g.granted_at,
                "expires_at": g.expires_at,
                "revoked_at": g.revoked_at,
            }
            for g in grants
        ],
    }


# ---- 儲存生命週期 ----

# ==============================================================================
# ZH: 儲存生命週期的四個動作 —— 🔴 回傳值一定要用
#
# ZH: 2026-08-27 稽核發現：這四支**全部忽略 `storage_lifecycle` 的回傳值**，
#     不管實際發生什麼都回 `{"status": "…"}`。於是連函式明確拒絕的情況
#     （不是 frozen 狀態、學期中不准歸檔）管理員看到的也是「成功」——
#     他會以為處理好了，然後去做下一步。
#
# ZH: ⚠️ 更大的問題不在這一層：`freeze` 只改狀態**不會真的讓儲存變唯讀**，
#     `archive` 只寫路徑**不會真的打包**。見 storage_lifecycle 各函式的 docstring。
# ==============================================================================
@router.post("/storage/freeze")
def storage_freeze(
    payload: StorageStateActionRequest,
    db: Session = Depends(get_db),
    admin: models.User = Depends(require_admin),
    request: Any = None,
) -> Any:
    """ZH: 凍結使用者儲存（停用 lab session 但保留檔案） | EN: Freeze storage

    @node job-scheduler/app/routers/admin.py::storage_freeze
    """
    ok = storage_lifecycle.freeze(db, user_id=payload.user_id, admin_id=admin.id, reason=payload.reason)
    # ZH: ⚠️ `frozen` 只是狀態標記 —— 使用者的儲存**不會真的變成唯讀**（尚未實作）。
    #     回傳裡明講，管理者才不會以為已經擋住了。
    return {"status": "frozen" if ok else "unchanged", "user_id": payload.user_id,
            "enforced": False,
            "note": "ZH: 狀態已標記，但唯讀限制尚未實作 | EN: state only; read-only is not enforced yet"}


@router.post("/storage/archive")
def storage_archive(
    payload: StorageStateActionRequest,
    db: Session = Depends(get_db),
    admin: models.User = Depends(require_admin),
) -> Any:
    """ZH: 歸檔到冷儲存（HDD） | EN: Archive to cold storage

    @node job-scheduler/app/routers/admin.py::storage_archive
    """
    ok = storage_lifecycle.archive(db, user_id=payload.user_id, admin_id=admin.id, reason=payload.reason)
    if not ok:
        # ZH: 拒絕的原因有三種（尚未實作打包／不是 frozen 狀態／學期中），
        #     函式已寫進 log。這裡不猜是哪一種,只誠實說「沒有歸檔」。
        # ZH: 用數字而不是 `status.HTTP_409_CONFLICT` —— 這個檔案沒有匯入 `status`，
        #     而那種錯誤 py_compile 與 import 都抓不到（它在函式內才求值），
        #     只有真的走到這條分支時才 NameError。全檔都用數字，跟著慣例走。
        raise HTTPException(
            status_code=409,
            detail="ZH: 未歸檔（尚未實作實際打包，或狀態／時段不允許）—— 請看伺服器日誌 | "
                   "EN: Not archived (packing not implemented, or state/season disallows it)")
    return {"status": "archived", "user_id": payload.user_id}


@router.post("/storage/restore")
def storage_restore(
    payload: StorageStateActionRequest,
    db: Session = Depends(get_db),
    admin: models.User = Depends(require_admin),
) -> Any:
    """ZH: 從凍結/歸檔還原 | EN: Restore from frozen/archived

    @node job-scheduler/app/routers/admin.py::storage_restore
    """
    ok = storage_lifecycle.restore(db, user_id=payload.user_id, admin_id=admin.id, reason=payload.reason)
    return {"status": "active" if ok else "unchanged", "user_id": payload.user_id}


@router.post("/storage/permanent-delete")
def storage_permanent_delete(
    payload: StorageStateActionRequest,
    db: Session = Depends(get_db),
    admin: models.User = Depends(require_admin),
) -> Any:
    """ZH: 永久刪除（需 admin 密碼驗證） | EN: Permanent delete (requires admin password)

    @node job-scheduler/app/routers/admin.py::storage_permanent_delete
    """
    if not payload.admin_password:
        raise HTTPException(400, "admin_password required for permanent delete")
    if not crud.verify_password(payload.admin_password, admin.hashed_password):
        raise HTTPException(403, "Admin password verification failed")
    ok = storage_lifecycle.permanent_delete(db, user_id=payload.user_id, admin_id=admin.id, reason=payload.reason)
    return {"status": "deleted" if ok else "unchanged", "user_id": payload.user_id}


@router.get("/storage/states")
def list_storage_states(
    state: Optional[str] = Query(None, description="active/frozen/archived/pending_delete"),
    db: Session = Depends(get_db),
    _: models.User = Depends(require_admin),
) -> Any:
    """ZH: 列出所有使用者儲存狀態 | EN: List all storage states

    @node job-scheduler/app/routers/admin.py::list_storage_states
    """
    states = storage_lifecycle.list_states(db, filter_state=state)
    return {"states": states}


# ---- Lab Sessions 監控 ----

@router.get("/lab/sessions")
def list_lab_sessions(
    db: Session = Depends(get_db),
    _: models.User = Depends(require_admin),
) -> Any:
    """ZH: 列出當前所有 lab sessions | EN: List all current lab sessions

    @node job-scheduler/app/routers/admin.py::list_lab_sessions
    """
    sessions = lab_manager.list_all_sessions(db)
    return {"sessions": sessions}


@router.post("/lab/sessions/{user_id}/force-stop")
def force_stop_session(
    user_id: str,
    session: Optional[str] = Query(None, description="ZH: 要關哪一份存檔；留空=目前在跑的那一份"),
    db: Session = Depends(get_db),
    admin: models.User = Depends(require_admin),
) -> Any:
    """ZH: 強制停止特定使用者 lab session | EN: Force-stop a user's lab session

    ZH: 🔴 `lab_manager.force_stop` **原本不存在**，這個端點從上線起每次都是 500。
        沒有測試涵蓋，所以測試一直綠著。2026-08-21 補上（見該函式的註解）。

    ZH: v3.6 —— `session` 留空代表「他目前正在跑的那一份」，不是 default。
        一次只開一份，所以「正在跑的那一份」是明確的。

    @node job-scheduler/app/routers/admin.py::force_stop_session
    """
    success = lab_manager.force_stop(db, user_id=user_id, admin_id=admin.id, session=session)
    if not success:
        raise HTTPException(404, "Session not found or already stopped")
    return {"status": "stopped", "user_id": user_id}


# ---- Secrets 監控（admin 也不可看 value） ----

@router.get("/secrets/{user_id}/names")
def list_user_secret_names(
    user_id: str,
    db: Session = Depends(get_db),
    _: models.User = Depends(require_admin),
) -> Any:
    """ZH: 列出某使用者的 secret 名稱（不回 value） | EN: List user secret names (no values)

    @node job-scheduler/app/routers/admin.py::list_user_secret_names
    """
    secrets_meta = secrets_service.list_secrets_masked(db, user_id=user_id)
    return {"user_id": user_id, "secrets": secrets_meta}


@router.delete("/secrets/{user_id}/{name}")
def admin_delete_user_secret(
    user_id: str,
    name: str,
    db: Session = Depends(get_db),
    admin: models.User = Depends(require_admin),
) -> Any:
    """ZH: 管理員刪除使用者的特定 secret（離校等情境） | EN: Admin delete a user's secret

    @node job-scheduler/app/routers/admin.py::admin_delete_user_secret
    """
    success = secrets_service.delete_secret(db, user_id=user_id, name=name, admin_id=admin.id)
    if not success:
        raise HTTPException(404, "Secret not found")
    return {"status": "deleted", "user_id": user_id, "name": name}


# ---- Audit log ----

@router.get("/audit")
def get_audit_log(
    admin_id: Optional[str] = Query(None),
    target_user: Optional[str] = Query(None),
    action: Optional[str] = Query(None),
    limit: int = Query(100, le=500),
    offset: int = Query(0, ge=0),
    db: Session = Depends(get_db),
    _: models.User = Depends(require_admin),
) -> Any:
    """ZH: 查詢 admin 操作審計 log（支援篩選與分頁） | EN: Query admin audit log

    @node job-scheduler/app/routers/admin.py::get_audit_log
    """
    q = db.query(models.AdminAction)
    if admin_id:
        q = q.filter(models.AdminAction.admin_id == admin_id)
    if target_user:
        q = q.filter(models.AdminAction.target_user == target_user)
    if action:
        q = q.filter(models.AdminAction.action == action)
    total = q.count()
    rows = q.order_by(models.AdminAction.timestamp.desc()).offset(offset).limit(limit).all()

    # ZH: 🔴 一起回**名字**，不要只回 UUID。
    #     稽核紀錄的用途就是「誰、對誰、做了什麼」——
    #     只顯示 `3ad36141` 的話，看的人得再去查一次那是誰，
    #     而那正是他打開這一頁想省下的動作。
    #
    # ZH: 一次撈完再對照，不要在迴圈裡逐筆查（100 筆就是 200 次查詢）。
    ids = {r.admin_id for r in rows} | {r.target_user for r in rows if r.target_user}
    names = {}
    if ids:
        names = {u.id: u.username for u in
                 db.query(models.User).filter(models.User.id.in_(ids)).all()}

    return {
        "total": total,
        "limit": limit,
        "offset": offset,
        "items": [
            {
                "id": r.id,
                "admin_id": r.admin_id,
                # ZH: 查不到就是 None —— 帳號被刪掉了。**不要回 UUID 當名字**，
                #     那會讓畫面顯示一串亂碼而看起來像正常的使用者名稱。
                "admin_username": names.get(r.admin_id),
                "target_user": r.target_user,
                "target_username": names.get(r.target_user) if r.target_user else None,
                "action": r.action,
                "payload": r.payload,
                "timestamp": r.timestamp,
                "ip_address": r.ip_address,
            }
            for r in rows
        ],
    }


# ==============================================================================
# ZH: v3.9 MYAI 手動補齊點數
# ==============================================================================
@router.post("/myai/topup", summary="手動把所有綁定帳號補到指定點數（預設先預覽）")
async def myai_manual_topup(
    target: int = Body(..., embed=True, description="要補到的點數（每人補到這個水位）"),
    dry_run: bool = Body(True, embed=True, description="true=只預覽不送出"),
    db: Session = Depends(get_db),
    admin: models.User = Depends(require_admin),
):
    """
    ZH: 手動補齊 —— 給例外狀況用（活動、補償、排程當天服務沒起來…）。

    ZH: `dry_run=true`（預設）只回報會補誰、補多少。要真的送出必須明確傳 false，
        與組織對照表匯入同一套慣例：**不可逆的操作先讓人看一眼**。

    ZH: 🔴 重複按不會重複發放 —— 這是「補到 N」而不是「加 N」的性質：
        第一次跑完所有人都在 N，第二次算差額就是空的。

    @node job-scheduler/app/routers/admin.py::myai_manual_topup
    """
    from ..services import myai_sync
    try:
        return await myai_sync.manual_topup(db, target, admin.id, dry_run=dry_run)
    except myai_sync.MyaiSyncError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/users/{user_id}/myai/grant", summary="給單一使用者加點（個別需求）")
async def myai_grant_points(
    user_id: str,
    points: int = Body(..., embed=True, description="要加的點數"),
    reason: str = Body("", embed=True, description="原因（會寫進稽核與廠商備註）"),
    db: Session = Depends(get_db),
    admin: models.User = Depends(require_admin),
):
    """
    ZH: 給一個人加點。

    ZH: 🔴 **這是「加 N」不是「補到 N」，按兩次就發兩次。**
        與手動補齊（/myai/topup）刻意不同 —— 那支重按是安全的，這支不是。
        介面上必須先確認再送。

    @node job-scheduler/app/routers/admin.py::myai_grant_points
    """
    from ..services import myai_sync
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="ZH: 找不到這個使用者 | EN: user not found")
    try:
        return await myai_sync.grant_points(db, user, points, admin.id, reason)
    except myai_sync.MyaiSyncError as e:
        raise HTTPException(status_code=400, detail=str(e))
