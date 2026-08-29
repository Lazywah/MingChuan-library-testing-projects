"""
==============================================================================
Service: MYAI 廠商平台 headless 同步 (v2.8) — 唯讀
==============================================================================
ZH: 用途：以管理者帳密 headless 登入廠商 (myai168) 管理後台 → 匯出使用者清單
    (.xlsx，含「點數」= Token 餘額) → 解析 → upsert 進 myai_accounts 供平台顯示。

    流程（已實測廠商端為標準表單登入、無驗證碼、無 CSRF）：
      1. GET  /mcu/ai/user/login            （取得初始 session cookie，若有）
      2. POST /mcu/ai/user/login_info        （form: email + password）→ 設 session
      3. GET  /mcu/gt_sdk/admin_168/user/export_user_list  → .xlsx
      4. 解析 9 欄 → 以 email 對應本平台使用者

    安全：帳密只從 .env 讀 (MYAI_ADMIN_EMAIL / MYAI_ADMIN_PASSWORD)，不存明文於碼。
          全程唯讀 — 只 login + export GET，絕不呼叫 transfer/register/edit/delete。
EN: Headless-login to the vendor admin, export the user list (.xlsx incl. token
    points), parse and upsert into myai_accounts for display. Read-only.
==============================================================================
"""

from __future__ import annotations

import io
import json
import logging
import os
import re
from datetime import datetime, timedelta, timezone

import httpx
from sqlalchemy.orm import Session

from .. import crud, models
from ..config import settings, SSO_POLICY
from . import email_service

logger = logging.getLogger(__name__)

# ZH: admin 交易日誌路徑（全體、逐筆、含備註/模型；filter: date_start/date_end/keyword）
ADMIN_TX_PATH = "/mcu/ai/admin/transaction"

# ZH: v2.8 交易列解析改用 lxml —— 廠商已把交易頁從 <b>欄位：</b>值 改版成 CSS grid
#     （kbx-row / kbx-cell / kbx-dt / kbx-time / kbx-muted…）。實作見 parse_transactions。
#     欄位（桌面 kbx-dt 依序）：時間 / 點數 / 餘額 / 備註 / 帳號(email + 名稱・sn:序號) / IP。
#     一律不取、不存 IP（隱私）。


class MyaiSyncError(Exception):
    """ZH: 同步流程的可預期錯誤（帳密未設、登入失敗、格式不符）| expected sync errors"""


# ZH: 廠商匯出欄位 → 我們欄位 | vendor export header → our column
COLUMN_MAP = {
    "編號": "vendor_sn",
    "類型": "user_type",
    "名稱": "name",
    "電子郵件": "email",
    "點數": "points",
    "有效期間": "expiry",
    "狀態": "status",
    "電子報": "newsletter",
    "備註": "note",
}


# ── v2.8 Session 快取：登入一次、cookie 重用，只有被導回登入頁(失效)才重登。──
# ZH: 消除「每次同步都登入」→ 更快，且不再把「Login was success.」洗版進廠商交易紀錄。
#     cookie 另存到 /data（與 DB 同區、volume 保存）→ 連 scheduler 重啟都免重登。全程唯讀。
# EN: Cache the vendor session cookie and reuse it across syncs; only re-login when the
#     response looks like the login page. Persisted next to the DB so it survives restarts.

# ZH: cookie 持久化檔（放 DB 同目錄，通常是 volume 掛載的 /data）
_COOKIE_FILE = os.path.join(os.path.dirname(settings.DATABASE_PATH) or ".", "myai_session.json")


def _save_cookies(cookies) -> None:
    """ZH: 把 vendor session cookie 存成 JSON（best-effort，失敗不影響同步）。

    @node job-scheduler/app/services/myai_sync.py::_save_cookies
    """
    try:
        items = [{"name": c.name, "value": c.value, "domain": c.domain, "path": c.path}
                 for c in cookies.jar]
        if not items:
            return
        with open(_COOKIE_FILE, "w", encoding="utf-8") as f:
            json.dump(items, f)
    except Exception as e:  # noqa: BLE001
        logger.debug("MYAI cookie 儲存失敗（略過）：%s", e)


def _load_cookies():
    """ZH: 啟動時載入上次的 cookie；沒有或壞掉就回 None（會自動重登）。

    @node job-scheduler/app/services/myai_sync.py::_load_cookies
    """
    try:
        if not os.path.exists(_COOKIE_FILE):
            return None
        with open(_COOKIE_FILE, "r", encoding="utf-8") as f:
            items = json.load(f)
        jar = httpx.Cookies()
        for it in items:
            jar.set(it["name"], it["value"], domain=it.get("domain") or "", path=it.get("path") or "/")
        return jar if len(jar.jar) else None
    except Exception as e:  # noqa: BLE001
        logger.debug("MYAI cookie 載入失敗（略過，將重登）：%s", e)
        return None


_MYAI_COOKIES = _load_cookies()  # type: httpx.Cookies | None  # 啟動即載入持久化 cookie


def _login_ctx():
    """ZH: 回 (base, login_page, headers)。廠商防跨站 → 登入 POST 必須帶對的 Referer/Origin。

    @node job-scheduler/app/services/myai_sync.py::_login_ctx
    """
    base = settings.MYAI_BASE_URL.rstrip("/")
    login_page = settings.MYAI_LOGIN_PATH.rsplit("/", 1)[0] + "/login"  # /mcu/ai/user/login
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                      "(KHTML, like Gecko) Chrome/149.0 Safari/537.36",
        "Referer": base + login_page,
        "Origin": base,
        "X-Requested-With": "XMLHttpRequest",
        "Accept": "application/json, text/plain, */*",
    }
    return base, login_page, headers


async def _session_request(do_fetch, is_valid):
    """ZH: 帶快取 cookie 送出請求；若被導回登入頁(is_valid=False) 才登入一次再抓。
           do_fetch(client)->Response、is_valid(Response)->bool。回最終 Response。
       EN: issue a request with cached cookies; re-login once only if invalid.

    @node job-scheduler/app/services/myai_sync.py::_session_request
    """
    global _MYAI_COOKIES
    if not settings.MYAI_ADMIN_EMAIL or not settings.MYAI_ADMIN_PASSWORD:
        raise MyaiSyncError("MYAI_ADMIN_EMAIL / MYAI_ADMIN_PASSWORD 未設定（請填入 .env）")
    base, login_page, headers = _login_ctx()
    async with httpx.AsyncClient(
        base_url=base, follow_redirects=True, timeout=httpx.Timeout(30.0),
        headers=headers, cookies=_MYAI_COOKIES,
    ) as client:
        # (1) 先用快取 cookie 直接抓（多數同步不需登入）
        if _MYAI_COOKIES is not None:
            try:
                resp = await do_fetch(client)
                if is_valid(resp):
                    _MYAI_COOKIES = client.cookies
                    _save_cookies(client.cookies)
                    return resp
            except httpx.HTTPError:
                pass  # 落到重新登入
        # (2) 沒 cookie 或已失效 → 登入一次再抓
        try:
            await client.get(login_page)  # 讓伺服器發初始 cookie（無則略過）
            await client.post(
                settings.MYAI_LOGIN_PATH,
                data={"email": settings.MYAI_ADMIN_EMAIL, "password": settings.MYAI_ADMIN_PASSWORD},
            )
        except httpx.HTTPError as e:
            raise MyaiSyncError(f"登入請求失敗：{e}")
        try:
            resp = await do_fetch(client)
        except httpx.HTTPError as e:
            raise MyaiSyncError(f"資料請求失敗：{e}")
        _MYAI_COOKIES = client.cookies
        _save_cookies(client.cookies)
        return resp


async def fetch_export_bytes() -> bytes:
    """ZH: 取得 export_user_list 的 .xlsx bytes（session 快取，失效才登入）。失敗拋 MyaiSyncError。
       EN: download export_user_list (.xlsx) reusing the cached session. Raises on failure.

    @node job-scheduler/app/services/myai_sync.py::fetch_export_bytes
    """
    async def _do(client):
        """@node job-scheduler/app/services/myai_sync.py::fetch_export_bytes.<nested@165>._do"""
        return await client.get(settings.MYAI_EXPORT_PATH)

    def _valid(r):  # 是 xlsx(ZIP 魔術數字 PK) = 登入有效
        """@node job-scheduler/app/services/myai_sync.py::fetch_export_bytes.<nested@168>._valid"""
        return r.status_code == 200 and r.content[:2] == b"PK"

    r = await _session_request(_do, _valid)
    if r.status_code != 200:
        raise MyaiSyncError(f"匯出回應 {r.status_code}（可能登入失敗或權限不足）")
    if r.content[:2] != b"PK":  # 非 xlsx → 多半被導回登入頁(HTML)
        raise MyaiSyncError("匯出內容非 xlsx（多半是帳密錯誤被導回登入頁，請確認 .env）")
    return r.content


def parse_xlsx(body: bytes) -> list[dict]:
    """ZH: 解析匯出 .xlsx → list[dict]（已對應欄位、points 轉 int）。
       EN: parse the exported .xlsx into mapped dict rows.

    @node job-scheduler/app/services/myai_sync.py::parse_xlsx
    """
    from openpyxl import load_workbook  # ZH: 延遲匯入 | lazy import

    wb = load_workbook(io.BytesIO(body), read_only=True, data_only=True)
    try:
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        headers = [str(h).strip() if h is not None else "" for h in next(it)]
        out: list[dict] = []
        for raw in it:
            if raw is None or all(c is None for c in raw):
                continue
            rec: dict = {}
            for h, v in zip(headers, raw):
                key = COLUMN_MAP.get(h)
                if key:
                    rec[key] = v
            if not rec.get("vendor_sn"):
                continue
            # points → int
            try:
                rec["points"] = int(float(rec.get("points") or 0))
            except (ValueError, TypeError):
                rec["points"] = 0
            # 其餘轉字串去空白
            for k in ("vendor_sn", "email", "name", "user_type", "expiry", "status", "newsletter", "note"):
                if rec.get(k) is not None:
                    rec[k] = str(rec[k]).strip()
            out.append(rec)
        return out
    finally:
        wb.close()


async def sync(db: Session) -> dict:
    """ZH: 完整同步：登入 → 匯出 → 解析 → upsert 進 myai_accounts。
       EN: full sync: login → export → parse → upsert into myai_accounts.

    @node job-scheduler/app/services/myai_sync.py::sync
    """
    body = await fetch_export_bytes()
    records = parse_xlsx(body)
    created = updated = 0
    now = datetime.now(timezone.utc)
    for rec in records:
        sn = rec["vendor_sn"]
        row = db.query(models.MyaiAccount).filter(models.MyaiAccount.vendor_sn == sn).first()
        if row:
            for k, v in rec.items():
                setattr(row, k, v)
            row.synced_at = now
            updated += 1
        else:
            db.add(models.MyaiAccount(synced_at=now, **rec))
            created += 1
    db.commit()
    logger.info("MYAI sync: total=%d created=%d updated=%d", len(records), created, updated)

    # ZH: 同步後自動以 email 配對綁定 | EN: auto-bind by email after sync
    match = auto_match(db)

    # ZH: 交易日誌同步（逐筆、含模型；失敗不影響帳號同步）
    # EN: also sync the per-event transaction log (best-effort)
    tx = {"fetched": 0, "created": 0}
    try:
        tx = await sync_transactions(db, days=settings.MYAI_TX_SYNC_DAYS)
    except Exception as e:  # noqa: BLE001
        logger.warning("MYAI tx-sync skipped: %s", e)

    return {
        "status": "ok",
        "total": len(records),
        "created": created,
        "updated": updated,
        "matched_created": match["matched_created"],
        "backfilled": match["backfilled"],
        "tx_fetched": tx.get("fetched", 0),
        "tx_created": tx.get("created", 0),
        "synced_at": now.isoformat(),
    }


def auto_match(db: Session) -> dict:
    """ZH: 以 email 自動配對「myai 帳號 ↔ 平台使用者」，建立/回填 external_ai_accounts 綁定。
       規則：myai.email == user.email(不分大小寫) 且該使用者尚未綁定 → 自動建綁定
       (vendor_username=email, myai_vendor_sn=vendor_sn)；已綁且 email 相符但缺 sn → 回填 sn。
       只寫本平台 DB；絕不碰廠商。回傳 {matched_created, backfilled}。
       EN: Auto-bind myai accounts to platform users by email. Writes our DB only.

    @node job-scheduler/app/services/myai_sync.py::auto_match
    """
    created = backfilled = 0
    myai_rows = (
        db.query(models.MyaiAccount)
        .filter(models.MyaiAccount.email.isnot(None))
        .all()
    )
    for m in myai_rows:
        email = (m.email or "").strip()
        if not email:
            continue
        user = db.query(models.User).filter(models.User.email.ilike(email)).first()
        if not user:
            continue  # ZH: 廠商端帳號在平台無對應使用者(如純管理員) | no platform user
        acc = (
            db.query(models.ExternalAiAccount)
            .filter(models.ExternalAiAccount.user_id == user.id)
            .first()
        )
        if not acc:
            db.add(models.ExternalAiAccount(
                user_id=user.id, vendor_username=email,
                myai_vendor_sn=m.vendor_sn, status="active", note="auto-matched",
            ))
            created += 1
        elif not acc.myai_vendor_sn and (acc.vendor_username or "").strip().lower() == email.lower():
            acc.myai_vendor_sn = m.vendor_sn  # ZH: 既有綁定回填穩定鍵 | backfill stable key
            backfilled += 1
    db.commit()
    logger.info("MYAI auto-match: created=%d backfilled=%d", created, backfilled)
    return {"matched_created": created, "backfilled": backfilled}


# ==============================================================================
# ZH: v2.8 交易日誌同步 —— 逐筆(全體、含模型)；不存 IP
# EN: v2.8 transaction-log sync — per event (all users, incl. model); no IP
# ==============================================================================
def _classify(note: str, pts: int) -> tuple[str, str | None]:
    """ZH: 由備註判斷事件類型與模型 | EN: classify event/model from note.

    @node job-scheduler/app/services/myai_sync.py::_classify
    """
    low = (note or "").lower()
    if "login" in low:
        return "login", None
    # ZH: Transfer(轉點) 與 Top Up(加值) 都是「配點」——前端 transfer 就是顯示成「配點」，
    #     語意相符，不必新增列舉值。原本 Top Up 落到 other，5 筆加值被歸成「其他」。
    #     一併改成不分大小寫：原本 startswith("Transfer") 是大小寫敏感，
    #     而上面的 login 判斷是小寫比對，同一函式兩套規則遲早出事。
    if low.startswith("transfer") or low.startswith("top up") or low.startswith("topup"):
        return "transfer", None
    if pts < 0:
        return "ai_usage", note.strip() or None
    return "other", None


def _to_int(s: str) -> int:
    """ZH: '2,100,000' / '-1,234' / '0' → int（去掉逗號等非數字字元）。

    @node job-scheduler/app/services/myai_sync.py::_to_int
    """
    try:
        return int(re.sub(r"[^\d\-]", "", (s or "").strip()) or "0")
    except ValueError:
        return 0


def parse_transactions(html: str) -> list[dict]:
    """ZH: 解析交易日誌 HTML → list[dict]（**不取 IP**）。支援兩種版型：
           v2 表格版（2026-08 廠商改版後的 <table>）優先，解不出來才退回 v1 的 kbx-grid。
           保留舊版是因為廠商回退或別的分頁沿用舊版時不該又壞一次。
       EN: dual-layout parser. Table layout first, kbx-grid as fallback. No IP stored.

    @node job-scheduler/app/services/myai_sync.py::parse_transactions
    """
    rows = _parse_tx_table(html) or _parse_tx_kbx(html)
    return _disambiguate_keys(rows)


def _disambiguate_keys(rows: list[dict]) -> list[dict]:
    """ZH: 同一組 (時間|sn|點數|備註) 出現多列時，第 2 筆起加 `|#N` 後綴。

       為什麼需要：實測 202 列只有 194 個相異鍵——8 組是**同一秒的兩次登入**
       （點數 0、餘額相同，資料上完全同一），會被去重掉，「登入次數」因此少算 8 次。

       ⚠ **第 1 筆刻意不加後綴**，這樣既有資料的 key 完全不變。
         舊 parser 每組只存得進第 1 筆，所以 DB 裡的就是「第 1 筆」的 key——
         加了後綴才會與既有 194 筆對不上，下次同步整批重複插入。
       序號是「組內出現序」，與查詢的日期範圍無關，所以不同窗口算出來一樣。

    @node job-scheduler/app/services/myai_sync.py::_disambiguate_keys
    """
    seen: dict[str, int] = {}
    for r in rows:
        base = r["dedup_key"]
        n = seen.get(base, 0) + 1
        seen[base] = n
        if n > 1:
            r["dedup_key"] = f"{base}|#{n}"
    return rows


def tx_row_count(html: str) -> int:
    """ZH: 頁面上「看起來有幾列交易」——與解析無關，只數 DOM。
           用來偵測「頁面有資料但一列都解不出來」＝版型又變了（見 sync_transactions）。

    ZH: 🔴 v3.8 —— **解不開的頁面要拋錯，不能回 0。**

    ZH: 這支是「解析失敗」的偵測器。但它原本在 HTML 解不開時
        `except: return 0` —— 而它要偵測的那些 parser 失敗時也回空清單。
        於是 `seen=0, rows=[]`，`if seen and not rows` 不成立、不拋錯，
        流程只印 `fetched=0`，讀起來就是「沒有新資料」。

    ZH: **這正是 2026-08 那場 29 天 / 201 筆事故的失敗模式**，
        只是觸發原因不同（那次是版型變了，這條是頁面根本解不開）。
        偵測器跟它要偵測的東西用同一種方式壞掉，等於沒有偵測器。

    ZH: 空字串是最實際的觸發點：session 過期被導向、gateway 打嗝、
        廠商回 204 —— `lxml` 對空文件會拋 `ParserError`。

    @node job-scheduler/app/services/myai_sync.py::tx_row_count
    """
    from lxml import html as lxml_html
    try:
        doc = lxml_html.fromstring(html)
    except Exception as e:  # noqa: BLE001
        raise MyaiSyncError(
            f"交易日誌頁面解不開（長度 {len(html or '')}）：{e}"
        ) from e
    n = len(doc.xpath("//tbody/tr"))
    n += len(doc.xpath("//div[contains(concat(' ', normalize-space(@class), ' '), ' kbx-row ')]"))
    return n


def _parse_tx_table(html: str) -> list[dict]:
    """ZH: v2 表格版型。欄序：時間 / 點數 / 餘額 / 備註 / 帳號 / IP。

       ⚠ 兩個會咬人的細節：
         1. 時間欄是 `<td>2026-08-16<br>10:23:45</td>`，直接取文字會黏成
            `2026-08-1610:23:45`——strptime 會炸，dedup_key 也會與舊資料對不上。
         2. **IP 是第 6 欄，一律不讀不存**（既有原則）。

    @node job-scheduler/app/services/myai_sync.py::_parse_tx_table
    """
    from lxml import html as lxml_html
    try:
        doc = lxml_html.fromstring(html)
    except Exception:  # noqa: BLE001
        return []

    out: list[dict] = []
    for tr in doc.xpath("//tbody/tr"):
        tds = tr.xpath("./td")
        if len(tds) < 5:                      # 少於 5 欄不是交易列（IP 欄可有可無）
            continue
        # ZH: 用 itertext 併空白，<br> 造成的斷行才會變成分隔而不是黏在一起
        t = " ".join(x.strip() for x in tds[0].itertext() if x.strip())
        pts = _to_int(tds[1].text_content())
        bal = _to_int(tds[2].text_content())
        note = tds[3].text_content().strip()

        acct_parts = [x.strip() for x in tds[4].itertext() if x.strip()]
        email = next((x for x in acct_parts if "@" in x), "")
        meta = next((x for x in acct_parts if "sn:" in x), "")
        msn = re.search(r"sn:(\d+)", meta)
        sn = msn.group(1) if msn else ""
        name = re.sub(r"・?\s*sn:\d+\s*$", "", meta).strip()
        # tds[5] 是 IP —— 刻意不讀

        try:
            occ = datetime.strptime(t, "%Y-%m-%d %H:%M:%S")
        except ValueError:
            occ = None
        if occ is None and not sn:
            continue                          # 兩個關鍵欄都沒有＝不是交易列（表頭之類）
        ev, model = _classify(note, pts)
        out.append({
            "occurred_at": occ, "vendor_sn": sn, "email": email, "name": name,
            "points_delta": pts, "balance": bal, "note": note,
            "event_type": ev, "model": model,
            # ZH: 去重鍵與 v1 完全相同的組成與格式，舊資料才不會被重複插入
            "dedup_key": f"{t}|{sn}|{pts}|{note}",
        })
    return out


def _parse_tx_kbx(html: str) -> list[dict]:
    """ZH: v1 kbx-grid 版型（2026-08 前）。保留作為 fallback。

    @node job-scheduler/app/services/myai_sync.py::_parse_tx_kbx
    """
    from lxml import html as lxml_html  # ZH: 延遲匯入 | lazy import

    def _has(el, token: str) -> bool:  # ZH: class 是否含某 token
        """@node job-scheduler/app/services/myai_sync.py::parse_transactions.<nested@327>._has"""
        return token in (el.get("class") or "").split()

    try:
        doc = lxml_html.fromstring(html)
    except Exception:  # noqa: BLE001
        return []

    out: list[dict] = []
    for row in doc.xpath("//div[contains(concat(' ', normalize-space(@class), ' '), ' kbx-row ')]"):
        grids = row.xpath(".//div[contains(concat(' ', normalize-space(@class), ' '), ' kbx-grid ')]")
        if not grids:
            continue
        grid = grids[0]
        # ZH: 時間 —— class 恰為 'kbx-time'（IP 欄雖也含 kbx-time 但同時含 kbx-dt，排除）
        tl = grid.xpath(".//div[normalize-space(@class)='kbx-time']/text()")
        t = tl[0].strip() if tl else ""
        # ZH: 桌面欄位（kbx-cell + kbx-dt）依序：點數 / 餘額 / 備註 / 帳號 / IP
        cells = [c for c in grid.xpath(".//div") if _has(c, "kbx-cell") and _has(c, "kbx-dt")]
        if len(cells) < 4:
            continue
        pts = _to_int(cells[0].text_content())
        bal = _to_int(cells[1].text_content())
        note = cells[2].text_content().strip()
        acct = cells[3]
        # ZH: 帳號欄 = email 子 div（含 @、非 kbx-muted）＋「名稱・sn:序號」的 kbx-muted 子 div
        email = ""
        for d in acct.xpath("./div"):
            txt = d.text_content().strip()
            if "@" in txt and not _has(d, "kbx-muted"):
                email = txt
                break
        name, sn = "", ""
        muted = [d for d in acct.xpath(".//div") if _has(d, "kbx-muted")]
        if muted:
            mt = muted[0].text_content().strip()   # ZH: 例："NyaLazy・sn:1003387"
            msn = re.search(r"sn:(\d+)", mt)
            sn = msn.group(1) if msn else ""
            name = re.sub(r"・?\s*sn:\d+\s*$", "", mt).strip()
        try:
            occ = datetime.strptime(t, "%Y-%m-%d %H:%M:%S")
        except ValueError:
            occ = None
        ev, model = _classify(note, pts)
        out.append({
            "occurred_at": occ, "vendor_sn": sn, "email": email, "name": name,
            "points_delta": pts, "balance": bal, "note": note,
            "event_type": ev, "model": model,
            "dedup_key": f"{t}|{sn}|{pts}|{note}",   # ZH: 去重鍵（不含 IP）
        })
    return out


def _tx_logged_in(r) -> bool:
    """ZH: 交易頁(已登入)含「交易紀錄／備註」；登入頁不含 → 用來判斷 session 是否有效。
       EN: the logged-in tx page contains these labels; the login page does not.

    @node job-scheduler/app/services/myai_sync.py::_tx_logged_in
    """
    if r.status_code != 200:
        return False
    t = r.text
    return ("交易紀錄" in t) or ("備註" in t)


async def fetch_transactions_html(date_start: str, date_end: str) -> str:
    """ZH: GET admin 交易日誌(日期範圍) → 回 HTML（session 快取，失效才登入）。唯讀。
       EN: GET the admin transaction log reusing the cached session. Read-only.

    @node job-scheduler/app/services/myai_sync.py::fetch_transactions_html
    """
    async def _do(client):
        """@node job-scheduler/app/services/myai_sync.py::fetch_transactions_html.<nested@392>._do"""
        return await client.get(ADMIN_TX_PATH, params={"date_start": date_start, "date_end": date_end})

    r = await _session_request(_do, _tx_logged_in)
    if r.status_code != 200:
        raise MyaiSyncError(f"交易日誌回應 {r.status_code}（可能登入失敗或權限不足）")
    if not _tx_logged_in(r):
        raise MyaiSyncError("交易日誌非預期內容（多半被導回登入頁，請確認 .env）")
    return r.text


async def sync_transactions(db: Session, days: int = 90) -> dict:
    """ZH: 抓近 N 天交易日誌 → 解析 → 去重 upsert（不存 IP）。回統計。
       EN: fetch last N days of the tx log, parse, dedup-insert (no IP).

    @node job-scheduler/app/services/myai_sync.py::sync_transactions
    """
    days = max(1, min(int(days or 90), 730))
    end = datetime.now()
    start = end - timedelta(days=days)
    html = await fetch_transactions_html(start.strftime("%Y-%m-%d"), end.strftime("%Y-%m-%d"))
    rows = parse_transactions(html)
    # ZH: ⚠ 這一段是這次事故的直接教訓。
    #     2026-08 廠商把交易頁從 kbx-grid 改成 <table>，parser 一列都解不出來，
    #     但流程只印了 fetched=0——讀起來就像「沒有新資料」，而暑假期間那完全合理，
    #     於是同步靜靜死了 29 天、漏掉 201 筆，沒有任何人發現。
    #     所以：頁面上明明有列、卻一列都解不出來，必須**當成錯誤**，不能當成沒資料。
    seen = tx_row_count(html)
    if seen and not rows:
        raise MyaiSyncError(
            f"交易日誌解析失敗：頁面有 {seen} 列但解出 0 筆，廠商版型可能又改了"
        )
    existing = {k for (k,) in db.query(models.MyaiTransaction.dedup_key).all()}
    now = datetime.now(timezone.utc)
    created = 0
    # ZH: 既有列的分類自我修正 —— 去重會跳過已存在的 key，所以「分類規則改了」
    #     不會回頭修正舊資料（實測：Top Up 的 5 筆一直卡在 other）。
    #     這裡只比對本次窗口內、且分類與現行規則不符的列，成本與窗口同級。
    reclassified = 0
    if rows:
        by_key = {r["dedup_key"]: r for r in rows}
        for old in (db.query(models.MyaiTransaction)
                      .filter(models.MyaiTransaction.dedup_key.in_(list(by_key.keys())))
                      .all()):
            r = by_key.get(old.dedup_key)
            if r and (old.event_type != r["event_type"] or old.model != r["model"]):
                old.event_type, old.model = r["event_type"], r["model"]
                reclassified += 1
    for r in rows:
        if r["dedup_key"] in existing:
            continue
        db.add(models.MyaiTransaction(synced_at=now, **r))
        existing.add(r["dedup_key"])
        created += 1
    db.commit()

    # ZH: v2.8 用本次抓到的「最新一列餘額」更新 myai_accounts.points，讓當前餘額隨交易變新
    #     （供低點數提醒即時判斷）；只更新本窗口有活動的人，其餘維持不變。
    bal_updated = _refresh_points_from_tx(db, rows)

    logger.info("MYAI tx-sync: fetched=%d created=%d reclassified=%d bal_updated=%d (days=%d)",
                len(rows), created, reclassified, bal_updated, days)
    return {"status": "ok", "fetched": len(rows), "created": created,
            "reclassified": reclassified,
            "skipped": len(rows) - created, "balance_updated": bal_updated, "days": days}


def _refresh_points_from_tx(db: Session, rows: list[dict]) -> int:
    """ZH: 以每位使用者(vendor_sn)在本批交易中最新一列的餘額，更新 myai_accounts.points。
       EN: update myai_accounts.points to each user's latest tx balance in this batch.

    @node job-scheduler/app/services/myai_sync.py::_refresh_points_from_tx
    """
    latest: dict[str, tuple] = {}   # vendor_sn -> (occurred_at, balance)
    for r in rows:
        sn, occ = r.get("vendor_sn"), r.get("occurred_at")
        if not sn or occ is None:
            continue
        if sn not in latest or occ > latest[sn][0]:
            latest[sn] = (occ, r["balance"])
    updated = 0
    for sn, (_occ, bal) in latest.items():
        acc = db.query(models.MyaiAccount).filter(models.MyaiAccount.vendor_sn == sn).first()
        if acc and acc.points != bal:
            acc.points = bal
            updated += 1
    if updated:
        db.commit()
    return updated


# ==============================================================================
# ZH: v3.3 自動開通（首次登入即為學生建立 MYAI 帳號並綁定）
# EN: v3.3 auto-provision — create the student's MYAI account on first login & bind
# ------------------------------------------------------------------------------
# ZH: 使用廠商管理端「批次註冊」正式功能（非繞過註冊頁的 CAPTCHA）：
#       POST /mcu/gt_sdk/admin_168/user/register_batch_check   ← 上傳 xlsx（驗證/預覽）
#       → 確認送出（第二段）
#     Excel 格式（官方範本 register_batch.xlsx，**無標題列**，使用者已確認欄序）：
#       A=email、B=暱稱、C=密碼、D=備註
#     ⚠️ 這是對廠商端的「寫入」。原「唯讀」界線由使用者於 2026-08-05 有意識放寬，
#        僅限此官方批次註冊功能；transfer/top_up/delete 一律仍禁止。
# EN: Uses the vendor's official admin bulk-registration feature (not a CAPTCHA bypass).
#     Template columns (no header): A=email, B=nickname, C=password, D=remark.
# ==============================================================================
# ==============================================================================
# ⚠️⚠️ 端點辨識（2026-08-06 實地確認頁面標題）—— 廠商管理端有 **三個** Excel 上傳功能，
#      長得幾乎一樣，用錯會造成不可逆的點數損失。務必只用 register_*：
#
#   /user/register_batch         → 「批次註冊」✅ 本模組唯一該用的
#   /user/get_credit_batch       → 「批次**回收**點數」⛔ 名單＝**被扣點的人**
#                                   （點數從名單上的帳號收回管理者，是扣點不是發放）
#   /user/transfer_credit_batch  → 「批次轉移點數」⛔ 名單＝**收到點數的人**
#                                   （點數**從管理者自己的帳號**轉出給名單上的帳號＝發放）
#
#   ⚠️ 兩者的名單語意**完全相反**，誤用方向錯誤即造成不可逆損失：
#      · 把註冊名單誤送 get_credit  → 一次扣光全部學生的點數
#      · 把註冊名單誤送 transfer    → 從管理者帳號一次發出大量點數（池子被掏空）
#   三者各有獨立的 *_check 確認端點與 *.xlsx 範本，欄位格式亦不同，不可混用。
#   下方 _assert_register_endpoint() 會在每次送出前硬性驗證路徑，避免日後誤改。
# EN: THREE similar Excel-upload features; only register_* is used here. The two
#     credit endpoints have OPPOSITE list semantics — get_credit deducts FROM the
#     listed accounts; transfer_credit grants TO them from the admin's own balance.
#     Either misuse is irreversible.
# ==============================================================================
REGISTER_BATCH_PATH = "/mcu/gt_sdk/admin_168/user/register_batch"
REGISTER_BATCH_CHECK_PATH = "/mcu/gt_sdk/admin_168/user/register_batch_check"

# ZH: 明確列為禁用，若不慎被填進上面兩個常數，防呆會擋下
_FORBIDDEN_BATCH_TOKENS = ("credit", "transfer", "delete", "top_up", "topup")


def _assert_register_endpoint(path: str) -> None:
    """
    ZH: 送出前硬性驗證：路徑必須是註冊端點，且不得含任何點數/轉移/刪除字樣。
        目的是防止日後有人改錯常數、或複製貼上到別的批次功能而造成點數損失。
    EN: Hard guard — the upload target must be the registration endpoint and must
        not contain credit/transfer/delete tokens (irreversible point loss otherwise).

    @node job-scheduler/app/services/myai_sync.py::_assert_register_endpoint
    """
    p = (path or "").lower()
    if "register_batch" not in p:
        raise MyaiSyncError(f"拒絕送出：目標端點不是批次註冊（{path}）")
    for bad in _FORBIDDEN_BATCH_TOKENS:
        if bad in p:
            raise MyaiSyncError(f"拒絕送出：端點含禁用字樣 '{bad}'（{path}）—— 可能誤用點數相關功能")


def gen_initial_password(length: int = 12) -> str:
    """
    ZH: 產生 MYAI 初始密碼。廠商規則 8~20 字元；此處固定 12 碼並保證含大小寫+數字。
        刻意不用學號（公開資訊 → 任何人可登入他人帳號）。排除易混淆字元 0/O/1/l/I。
    EN: Random initial password (vendor allows 8-20). Never the student id (public).

    @node job-scheduler/app/services/myai_sync.py::gen_initial_password
    """
    import secrets as _secrets
    lower = "abcdefghijkmnopqrstuvwxyz"
    upper = "ABCDEFGHJKLMNPQRSTUVWXYZ"
    digits = "23456789"
    pool = lower + upper + digits
    while True:
        pwd = "".join(_secrets.choice(pool) for _ in range(length))
        if (any(c in lower for c in pwd) and any(c in upper for c in pwd)
                and any(c in digits for c in pwd)):
            return pwd


def build_register_xlsx(rows: list[dict]) -> bytes:
    """
    ZH: 依官方範本格式產生上傳用 xlsx（無標題列；A=email B=暱稱 C=密碼 D=備註）。
    EN: Build the upload workbook matching the vendor template (no header row).

    rows: [{"email":..., "nickname":..., "password":..., "remark":...}, ...]

    @node job-scheduler/app/services/myai_sync.py::build_register_xlsx
    """
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    for r in rows:
        ws.append([
            (r.get("email") or "").strip(),
            (r.get("nickname") or "").strip(),
            (r.get("password") or "").strip(),
            (r.get("remark") or "").strip(),
        ])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


REGISTER_CONFIRM_PATH = "/mcu/gt_sdk/admin_168/user/register_batch_info"


def _echoed_rows(html: str, fields: tuple) -> list[dict]:
    """
    ZH: 從第一段（預覽）的回應裡，抓廠商**自己解析出來**的那幾列。

    ZH: 廠商的預覽頁把每一列回吐成一組平行的 `name="xxx[]"` input，
        例如註冊是 emails[]/name_displays[]/passwords[]/remarks[]。
        把它們讀回來，就能在按下確認之前比對「他讀到的」跟「我送的」一不一樣。

    ZH: 🔴 這不是裝飾。**沒有這一步的話，第二段就是盲送** ——
        廠商若因為 email 重複、格式不符而少解析了一列，我們照樣會按確認，
        然後把「成功」寫進資料庫。預覽頁存在的意義就是給人對帳的，
        程式也該對一次。

    @node job-scheduler/app/services/myai_sync.py::_echoed_rows
    """
    cols = {}
    for f in fields:
        cols[f] = re.findall(
            r'<input[^>]*\bname="' + re.escape(f) + r'\[\]"[^>]*\bvalue="([^"]*)"',
            html, re.I)
    n = len(cols[fields[0]]) if fields else 0
    if any(len(v) != n for v in cols.values()):
        raise MyaiSyncError(
            "廠商預覽頁的欄位數量對不齊："
            + "、".join(f"{f}={len(v)}" for f, v in cols.items()))
    import html as _html
    return [{f: _html.unescape(cols[f][i]) for f in fields} for i in range(n)]


async def register_batch(rows: list[dict]) -> dict:
    """
    ZH: 送出批次註冊（兩段式，**第二段會真的建立帳號**）。
          第一段 POST xlsx → register_batch_check（廠商解析並回吐預覽）
          第二段 POST 表單 → register_batch_info（確認建立）
        兩段之間會比對廠商解析出來的 email 是否與送出的完全一致，不一致就中止。
    EN: Two-stage batch registration; stage 2 actually creates the accounts.
        The echoed preview rows are reconciled against what we sent before confirming.

    ZH: ⚠️ **第二段失敗時不重試** —— 重送可能建出重複帳號。
        回傳 {"ok", "status", "created"(bool), "rows"(廠商解析出來的列), "html"}。

    @node job-scheduler/app/services/myai_sync.py::register_batch
    """
    # ZH: 防呆 —— 廠商還有「批次回收點數 / 批次轉移點數」兩個長得一樣的 Excel 上傳功能，
    #     用錯會造成不可逆的點數損失。送出前硬性驗證目標端點。
    _assert_register_endpoint(REGISTER_BATCH_CHECK_PATH)
    _assert_register_endpoint(REGISTER_BATCH_PATH)
    _assert_register_endpoint(REGISTER_CONFIRM_PATH)

    xlsx = build_register_xlsx(rows)

    async def _do(client):
        """@node job-scheduler/app/services/myai_sync.py::register_batch.<nested@568>._do"""
        return await client.post(
            REGISTER_BATCH_CHECK_PATH,
            files={"upload_xls": ("register_batch.xlsx", xlsx,
                                  "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            headers={"Referer": settings.MYAI_BASE_URL.rstrip("/") + REGISTER_BATCH_PATH},
        )

    def _valid(r):
        """@node job-scheduler/app/services/myai_sync.py::register_batch.<nested@576>._valid"""
        return r.status_code == 200 and "Unauthorized" not in r.text[:200]

    r = await _session_request(_do, _valid)
    if not _valid(r):
        return {"ok": False, "status": r.status_code, "created": False,
                "rows": [], "html": r.text}

    fields = ("emails", "name_displays", "passwords", "remarks")
    echoed = _echoed_rows(r.text, fields)

    # ZH: 對帳 —— 廠商解析出來的 email 必須跟我們送的一模一樣（含筆數）。
    #     少一列代表他拒收了那一列（重複、格式不符…），這時候按確認就會
    #     建出「跟我們以為的不一樣」的結果，所以寧可整批中止。
    sent = [(x.get("email") or "").strip().lower() for x in rows]
    got = [(x["emails"] or "").strip().lower() for x in echoed]
    if sent != got:
        raise MyaiSyncError(
            f"廠商預覽的名單與送出的不符（送出 {len(sent)} 筆、解析 {len(got)} 筆）"
            f"：{got or '（空）'} —— 已中止，未建立任何帳號")

    # ZH: 第二段 —— 把預覽頁回吐的值原樣送回去確認。
    #     刻意用**廠商回吐的值**而不是我們手上的原值：如果他做了正規化
    #     （去空白、轉小寫…），照他的送回去才不會又觸發一次解析差異。
    form = []
    for row in echoed:
        for f in fields:
            form.append((f + "[]", row[f]))

    async def _confirm(client):
        """@node job-scheduler/app/services/myai_sync.py::register_batch.<nested>._confirm"""
        return await client.post(
            REGISTER_CONFIRM_PATH, data=form,
            headers={"Referer": settings.MYAI_BASE_URL.rstrip("/") + REGISTER_BATCH_CHECK_PATH},
        )

    logger.info("MYAI 批次註冊：確認建立 %d 個帳號", len(echoed))
    r2 = await _session_request(_confirm, _valid)
    ok = _valid(r2)
    if not ok:
        # ZH: 🔴 這裡**不重試**。第二段送出後帳號可能已經建了，
        #     重送一次會變成建兩個。交給呼叫端當失敗處理、人工對帳。
        logger.error("MYAI 批次註冊第二段失敗（回應 %s）—— 不重試，請人工到廠商後台確認",
                     r2.status_code)
    return {"ok": ok, "status": r2.status_code, "created": ok,
            "rows": echoed, "html": r2.text}


# ==============================================================================
# ZH: v3.9 批次轉移點數（發新帳號的初始點數）
# ==============================================================================
# ZH: 🔴 **這是唯一一個會讓平台的點數池變少的功能。** 動它之前先讀完這一段。
#
# ZH: 廠商官方範本 transfer_credit_batch.xlsx（2026-08-29 實際下載確認，**無標題列**）：
#       A = 收點數的帳號 email      例 mingtali@gmail.com
#       B = 點數（整數）            例 168
#       C = 備註                    例 remark1
#     ⚠️ **與註冊那支的欄位不同**（註冊是 A=email／B=暱稱／C=密碼／D=備註）。
#        C 欄在註冊是密碼、在轉點是備註 —— 混用會把密碼當備註送出去。
#
# ZH: 兩段式，但**第二段的路徑後綴與註冊不一樣**（實測 2026-08-29）：
#       註冊：register_batch_check       → register_batch_info
#       轉點：transfer_credit_batch_check → transfer_credit_batch_result
#     欄位也不同：emails[] / transferPoints[] / remarks[]（remarks 是 hidden）。
#
# ZH: 🔴 **點數從「平台登入廠商後台用的那個帳號」轉出**，不是逐筆指定的參數，
#     也就是 .env 的 MYAI_ADMIN_EMAIL。改那個值會連帶改變轉出來源，畫面上不會提示。
#
# ZH: 🔴 **不可逆。** 送出去的點數收不回來（收回是另一支 get_credit_batch，
#     而那支的名單語意是「被扣點的人」—— 方向相反，誤用會扣光學生的點數）。
#     所以：不重試、失敗就失敗、每一次發放都寫稽核。

TRANSFER_BATCH_PATH = "/mcu/gt_sdk/admin_168/user/transfer_credit_batch"
TRANSFER_BATCH_CHECK_PATH = "/mcu/gt_sdk/admin_168/user/transfer_credit_batch_check"
TRANSFER_CONFIRM_PATH = "/mcu/gt_sdk/admin_168/user/transfer_credit_batch_result"


def _assert_transfer_endpoint(path: str, confirm_grant: bool) -> None:
    """
    ZH: 轉點端點的硬性驗證。

    ZH: 🔴 為什麼要 `confirm_grant` 這個看似多餘的參數：
        這支函式擋得住「路徑打錯」，但擋不住「有人在別的地方順手呼叫它」。
        要求呼叫端明確傳 True，等於逼他在自己的程式碼裡寫下「我知道這會扣點」。
        沒有這個的話，一個 for 迴圈裡的手滑就能把池子送光。

    ZH: ⚠️ 白名單只開 `transfer_credit_batch` 這一族路徑。
        `get_credit`（回收，名單＝被扣點的人）與 `delete` 一律仍然全禁 ——
        那兩支的誤用方向是「扣光學生的點數」，比池子被掏空更難善後。

    @node job-scheduler/app/services/myai_sync.py::_assert_transfer_endpoint
    """
    if not confirm_grant:
        raise MyaiSyncError(
            "拒絕送出：轉移點數必須由呼叫端明確傳 confirm_grant=True"
            "（這個操作會從管理者帳號扣點且不可逆）")
    p = (path or "").lower()
    if "transfer_credit_batch" not in p:
        raise MyaiSyncError(f"拒絕送出：目標端點不是批次轉移點數（{path}）")
    for bad in ("get_credit", "delete", "top_up", "topup", "register"):
        if bad in p:
            raise MyaiSyncError(
                f"拒絕送出：端點含禁用字樣 '{bad}'（{path}）—— 可能誤用其他批次功能")


def _pool_balance(html: str):
    """
    ZH: 從轉點預覽頁讀出**轉出帳號自己還剩多少點**（廠商叫「您的點數」）。

    ZH: 讀不出來回 None —— **不要回 0**。0 會讓下面的檢查誤判成「池子空了」
        而中止每一次補點，那個失敗方向看起來像「功能壞了」但其實只是版型改了。
        讀不出來時我們選擇照送（廠商端本來就會擋不足額），並在日誌講明沒查到。

    ZH: 版型（2026-08-29 實測）：
          <div class="css_td">您的點數</div><div class="css_td">2,033,236</div>

    @node job-scheduler/app/services/myai_sync.py::_pool_balance
    """
    m = re.search(r"您的點數\s*</div>\s*<div[^>]*>\s*([\d,]+)\s*</div>", html)
    if not m:
        return None
    try:
        return int(m.group(1).replace(",", ""))
    except ValueError:
        return None


def build_transfer_xlsx(rows: list[dict]) -> bytes:
    """
    ZH: 依廠商範本組出轉點用的 xlsx（三欄、無標題列）。

    ZH: rows = [{"email":..., "points": int, "remark":...}, ...]

    ZH: ⚠️ points 一律轉成 int 且必須 > 0 —— 送 0 是白跑一趟，
        送負數在廠商端的行為未知（可能變成扣點），不要試。

    @node job-scheduler/app/services/myai_sync.py::build_transfer_xlsx
    """
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    for r in rows:
        email = (r.get("email") or "").strip()
        pts = int(r.get("points") or 0)
        if not email:
            raise MyaiSyncError("轉點名單有空的 email")
        if pts <= 0:
            raise MyaiSyncError(f"轉點數必須大於 0（{email} 給的是 {pts}）")
        ws.append([email, pts, (r.get("remark") or "").strip()])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def transfer_credit_batch(rows: list[dict], confirm_grant: bool = False) -> dict:
    """
    ZH: 對廠商送出批次轉移點數（兩段式）。**會從管理者帳號扣點，不可逆。**

    ZH: rows = [{"email", "points", "remark"}]；`confirm_grant` 必須明確傳 True。

    ZH: 與註冊同樣的對帳：第一段的預覽頁會回吐廠商解析出來的 email 與點數，
        兩者都要與送出的完全一致才按確認。點數對不上就中止 ——
        「多發了幾點」是收不回來的。

    ZH: ⚠️ **不重試。** 第二段失敗時點數**可能已經送出也可能沒有**，
        呼叫端必須當成「可能已發放」處理（去廠商後台對帳），
        **絕對不要自動再送一次**。重送一次 = 有機率發兩倍。

    @node job-scheduler/app/services/myai_sync.py::transfer_credit_batch
    """
    _assert_transfer_endpoint(TRANSFER_BATCH_CHECK_PATH, confirm_grant)
    _assert_transfer_endpoint(TRANSFER_CONFIRM_PATH, confirm_grant)
    if not rows:
        return {"ok": True, "granted": False, "count": 0, "points": 0}

    body = build_transfer_xlsx(rows)
    total = sum(int(r.get("points") or 0) for r in rows)
    logger.warning("MYAI 批次轉點：%d 個帳號、合計 %d 點（來源＝平台的廠商管理帳號）",
                   len(rows), total)

    async def _do(client):
        """@node job-scheduler/app/services/myai_sync.py::transfer_credit_batch.<nested>._do"""
        return await client.post(
            TRANSFER_BATCH_CHECK_PATH,
            files={"upload_xls": ("transfer_credit_batch.xlsx", body,
                                  "application/vnd.openxmlformats-officedocument."
                                  "spreadsheetml.sheet")},
            headers={"Referer": settings.MYAI_BASE_URL.rstrip("/") + TRANSFER_BATCH_PATH},
        )

    def _valid(r):
        """@node job-scheduler/app/services/myai_sync.py::transfer_credit_batch.<nested>._valid"""
        return r.status_code == 200 and "Unauthorized" not in r.text[:200]

    r = await _session_request(_do, _valid)
    if not _valid(r):
        raise MyaiSyncError(f"轉點預覽回應 {r.status_code}")

    # ZH: 🔴 送出去之前先看轉出帳號還剩多少 —— 池子不夠而硬送，
    #     廠商可能只轉一部分（誰拿到誰沒拿到我們無從得知），
    #     那比整批不送難善後得多。預覽頁本來就把餘額印在上面，用它。
    pool = _pool_balance(r.text)
    if pool is None:
        logger.warning("MYAI 轉點：預覽頁讀不到「您的點數」，跳過餘額檢查（版型可能改了）")
    elif total > pool:
        raise MyaiSyncError(
            f"拒絕送出：要轉 {total} 點，但轉出帳號只剩 {pool} 點 —— 已中止，未轉出任何點數")

    fields = ("emails", "transferPoints", "remarks")
    echoed = _echoed_rows(r.text, fields)

    sent = [((x.get("email") or "").strip().lower(), int(x.get("points") or 0))
            for x in rows]
    got = [((x["emails"] or "").strip().lower(), int(x["transferPoints"] or 0))
           for x in echoed]
    if sent != got:
        raise MyaiSyncError(
            f"廠商預覽的轉點名單與送出的不符（送出 {sent}、解析 {got}）—— 已中止，未轉出任何點數")

    form = []
    for row in echoed:
        for f in fields:
            form.append((f + "[]", row[f]))

    async def _confirm(client):
        """@node job-scheduler/app/services/myai_sync.py::transfer_credit_batch.<nested>._confirm"""
        return await client.post(
            TRANSFER_CONFIRM_PATH, data=form,
            headers={"Referer": settings.MYAI_BASE_URL.rstrip("/") + TRANSFER_BATCH_CHECK_PATH},
        )

    r2 = await _session_request(_confirm, _valid)
    ok = _valid(r2)
    if not ok:
        logger.error("MYAI 轉點第二段失敗（回應 %s）—— **不重試**，"
                     "點數可能已轉出，請人工到廠商後台對帳", r2.status_code)
    else:
        logger.warning("MYAI 轉點完成：%d 個帳號、合計 %d 點", len(echoed), total)
    return {"ok": ok, "granted": ok, "count": len(echoed), "points": total}


# ==============================================================================
# ZH: v3.9 每月補點 —— 把所有綁定帳號補到同一個水位
# ==============================================================================
# ZH: 規則（擁有者 2026-08-29 裁定）：
#       · 補到**固定值**（`myai_monthly_topup_to`），不是每人固定加
#       · **所有**綁定且啟用的帳號，不看這個月有沒有登入
#       · 每月第 `myai_monthly_topup_day` 天（預設 1 號，台北時間）
#       · 到期日不管 —— 那是廠商在處理的事
#
# ZH: 🔴 **冪等靠資料庫的事實，不靠排程準時。** 排程每小時醒一次，同一天會醒很多次；
#     容器重啟、時間調整、補跑都會讓「今天是不是 1 號」重複成立。
#     所以真正的閘門是「這個月做過了沒」，記在 SystemConfig。

TOPUP_MONTH_KEY = "myai_topup_last_month"


def _taipei_month(now=None) -> str:
    """ZH: 台北時間的 YYYY-MM。

    ZH: 🔴 用台北不用 UTC：「每月 1 號」是講給人聽的日期。
        UTC 的話台北時間 1 號早上 8 點之前都還算上個月，
        補點會在 1 號當天「還沒發生」，而看畫面的人已經在等了。

    @node job-scheduler/app/services/myai_sync.py::_taipei_month
    """
    from ..gpu_schedule import TZ_TAIPEI
    return (now or datetime.now(timezone.utc)).astimezone(TZ_TAIPEI).strftime("%Y-%m")


def _taipei_day(now=None) -> int:
    """ZH: 台北時間的日（1–31）。

    @node job-scheduler/app/services/myai_sync.py::_taipei_day
    """
    from ..gpu_schedule import TZ_TAIPEI
    return (now or datetime.now(timezone.utc)).astimezone(TZ_TAIPEI).day


def topup_targets(db: Session, target: int) -> list[dict]:
    """
    ZH: 算出這次要補誰、各補多少。回 [{"email", "points", "remark"}]。

    ZH: 只收**已綁定平台帳號**的人 —— 廠商後台還有我們的管理帳號與其他來源的
        帳號，補到那些身上是把點數送給不相干的人。綁定是「這是我們的學生」
        唯一可靠的判準。

    ZH: 已經 >= 目標的人不入列（補 0 是白跑，而且 build_transfer_xlsx 會拒絕）。

    ZH: ⚠️ 點數的新鮮度由呼叫端負責（先跑一次 sync）——
        拿舊資料算差額會補過頭，而多補的點數收不回來。

    @node job-scheduler/app/services/myai_sync.py::topup_targets
    """
    rows = []
    # ZH: 🔴 轉出帳號自己要排除。它同時也是一個綁定的平台帳號
    #     （MYAI_ADMIN_EMAIL 就是某個人的學號信箱），池子一旦低於目標值，
    #     補點就會變成「自己轉給自己」—— 廠商對這種操作的行為未知，不要試。
    source = (settings.MYAI_ADMIN_EMAIL or "").strip().lower()
    accs = (db.query(models.ExternalAiAccount)
              .filter(models.ExternalAiAccount.status == "active").all())
    for acc in accs:
        row = None
        if acc.myai_vendor_sn:
            row = (db.query(models.MyaiAccount)
                     .filter(models.MyaiAccount.vendor_sn == acc.myai_vendor_sn).first())
        if row is None and acc.vendor_username:
            row = (db.query(models.MyaiAccount)
                     .filter(models.MyaiAccount.email.ilike(acc.vendor_username)).first())
        if row is None:
            # ZH: 綁了但廠商端查不到 —— 不猜、不補，留給同步或人工處理。
            logger.info("MYAI 每月補點跳過 %s：廠商端查不到這個帳號", acc.vendor_username)
            continue
        email = (row.email or acc.vendor_username or "").strip()
        if source and email.lower() == source:
            logger.info("MYAI 每月補點跳過轉出帳號本身（%s）", email)
            continue
        have = int(row.points or 0)
        if have >= target:
            continue
        rows.append({"email": email,
                     "points": target - have,
                     "remark": "monthly-topup"})
    return rows


async def monthly_topup(db: Session, force: bool = False) -> dict:
    """
    ZH: 每月補點主流程。**一個月只會真的送出一次。**

    ZH: `force=True` 只給人工補跑用（跳過「今天是不是補點日」，但**不跳過**
        「這個月做過了沒」）—— 那道閘門不能繞，繞了就是重複發放。

    ZH: 🔴 送出之後**無論成敗都把這個月標成做過了**。第二段失敗時點數
        可能已經轉出，重跑一次就是發兩倍。失敗要人去對帳，不是自動重試。

    @node job-scheduler/app/services/myai_sync.py::monthly_topup
    """
    target = int(crud.get_setting(db, "myai_monthly_topup_to") or 0)
    if target <= 0:
        return {"status": "disabled"}

    month = _taipei_month()
    if crud.get_system_config(db, TOPUP_MONTH_KEY, "") == month:
        return {"status": "already_done", "month": month}

    if not force:
        day = int(crud.get_setting(db, "myai_monthly_topup_day") or 1)
        # ZH: 用 `!=` —— **只在當天跑**（擁有者裁定：服務 24/7 在線，
        #     不需要「錯過就補跑」的語意；而補跑會讓「月中才打開這個功能」
        #     當場補一次，那不是預期的行為）。
        # ZH: ⚠️ 代價要知道：補點日**整天**服務都沒起來的話，這個月就不補了，
        #     而且不會有錯誤訊息。真的遇到就用手動補齊（manual_topup）。
        if _taipei_day() != day:
            return {"status": "not_today", "day": day}

    # ZH: 先同步一次再算差額 —— 拿舊點數算會補過頭，而多補的收不回來。
    #     同步失敗就整個放棄：寧可這個月晚幾小時補，也不要照著錯的數字補。
    #     （這裡**不**標記月份，所以下一輪醒來會再試。）
    try:
        await sync(db)
    except Exception as e:  # noqa: BLE001
        logger.error("MYAI 每月補點中止：同步失敗，不拿舊點數算差額（%s）", e)
        return {"status": "sync_failed", "error": str(e)[:200]}

    rows = topup_targets(db, target)
    if not rows:
        crud.set_system_config(db, TOPUP_MONTH_KEY, month)
        logger.info("MYAI 每月補點：沒有人低於 %d 點，本月完成", target)
        return {"status": "nobody_below", "month": month, "target": target}

    total = sum(r["points"] for r in rows)
    logger.warning("MYAI 每月補點 %s：%d 人、合計 %d 點（補到 %d）",
                   month, len(rows), total, target)

    # ZH: 先標月份再送出。順序是刻意的 —— 標記失敗了就別送（還沒動到點數），
    #     但送出後才標記的話，中間掛掉就會重送。兩種錯法裡這一種安全得多：
    #     最壞是「這個月沒補到」，而不是「補了兩次」。
    crud.set_system_config(db, TOPUP_MONTH_KEY, month)

    try:
        res = await transfer_credit_batch(rows, confirm_grant=True)
    except Exception as e:  # noqa: BLE001
        logger.error("MYAI 每月補點失敗 %s：%s —— **不重試**，請人工對帳", month, e)
        return {"status": "failed", "month": month, "count": len(rows),
                "points": total, "error": str(e)[:200]}

    if not res.get("granted"):
        logger.error("MYAI 每月補點狀態未知 %s —— 點數可能已轉出，請到廠商後台對帳", month)
        return {"status": "unknown", "month": month, "count": len(rows), "points": total}

    logger.info("MYAI 每月補點完成 %s：%d 人、合計 %d 點", month, len(rows), total)
    return {"status": "done", "month": month, "count": len(rows),
            "points": total, "target": target}


async def manual_topup(db: Session, target: int, admin_id: str,
                       dry_run: bool = True) -> dict:
    """
    ZH: 手動把所有綁定帳號補到指定水位。給例外狀況用（活動、補償、排程漏跑…）。

    ZH: 與每月補點的差別只有兩個，其餘完全共用同一套邏輯：
          · **不看補點日、不看「這個月做過了沒」** —— 它就是為了例外而存在的
          · **有真正的執行者** → 稽核寫得進 admin_actions（自動補點沒有管理員，
            所以那條路只能記在帳號自己身上）

    ZH: 🔴 **重複按不會重複發放** —— 因為是「補到 N」不是「加 N」：
        第一次跑完所有人都在 N，第二次算差額就是空的。
        這是「補到固定值」相對於「固定加」最重要的性質，不要改成後者。

    ZH: `dry_run=True`（預設）只回報會補誰、補多少，不送出任何東西。

    @node job-scheduler/app/services/myai_sync.py::manual_topup
    """
    target = int(target or 0)
    if target <= 0:
        raise MyaiSyncError("補到的點數必須大於 0")

    # ZH: 先同步 —— 拿舊點數算差額會補過頭，而多補的收不回來。
    #     預覽也要同步：給管理者看的數字必須是他按下確認時會用的那一份。
    await sync(db)

    rows = topup_targets(db, target)
    total = sum(r["points"] for r in rows)
    preview = {"target": target, "count": len(rows), "points": total,
               "rows": [{"email": r["email"], "points": r["points"]} for r in rows]}

    if dry_run:
        return {"status": "preview", **preview}
    if not rows:
        return {"status": "nobody_below", **preview}

    logger.warning("MYAI 手動補齊：admin=%s 補到 %d 點，%d 人、合計 %d 點",
                   admin_id, target, len(rows), total)

    import json as _json

    def _audit(outcome: str):
        """ZH: 不可逆的操作一定要留下軌跡（誰、補到多少、幾個人、結果）。"""
        db.add(models.AdminAction(
            admin_id=admin_id, target_user=None, action="myai_manual_topup",
            payload=_json.dumps({"target": target, "count": len(rows),
                                 "points": total, "outcome": outcome},
                                ensure_ascii=False),
            timestamp=datetime.now(timezone.utc)))
        db.commit()

    try:
        res = await transfer_credit_batch(rows, confirm_grant=True)
    except Exception as e:  # noqa: BLE001
        # ZH: 中止在預覽階段（對帳不符、餘額不足…）→ 點數確定沒轉出。
        #     送出後才斷的話也走這裡，所以稽核記 "failed" 而不是 "not_sent" ——
        #     這兩種我們分不出來，就不要假裝分得出來。
        _audit(f"failed: {str(e)[:120]}")
        logger.error("MYAI 手動補齊失敗：%s —— **不重試**，請人工對帳", e)
        raise

    _audit("ok" if res.get("granted") else "unknown")
    if not res.get("granted"):
        logger.error("MYAI 手動補齊狀態未知 —— 點數可能已轉出，請到廠商後台對帳")
        return {"status": "unknown", **preview}
    logger.info("MYAI 手動補齊完成：%d 人、合計 %d 點", len(rows), total)
    return {"status": "done", **preview}


def _nickname_for(user) -> str:
    """ZH: 暱稱優先用平台顯示名，退回 username（學號）| EN: nickname for the vendor account

    @node job-scheduler/app/services/myai_sync.py::_nickname_for
    """
    for attr in ("display_name", "full_name", "name"):
        v = (getattr(user, attr, None) or "").strip()
        if v:
            return v[:60]
    return (user.username or "")[:60]


def store_initial_password(db: Session, acc, plaintext: str) -> None:
    """ZH: 加密暫存初始密碼（AES-256-GCM，同 user_secrets 的 KEK）並記發放時間。
       EN: Encrypt-at-rest the generated initial password with the shared KEK.

    @node job-scheduler/app/services/myai_sync.py::store_initial_password
    """
    from . import secrets_service
    acc.init_pwd_enc = secrets_service.encrypt_value(plaintext)
    acc.init_pwd_at = datetime.now(timezone.utc)
    acc.init_pwd_ack = 0
    db.commit()


def read_initial_password(db: Session, acc, retention_days: int) -> str | None:
    """
    ZH: 讀出未逾期且未被確認修改的初始密碼；逾期/已確認則就地清除並回 None。
        （學生始終可用 MYAI 自己的「忘記密碼」+ 學校信箱自助重設，故到期清除不會鎖死人。）
    EN: Return the initial password if still within retention and unacknowledged;
        otherwise purge it in place. Students can always self-serve via MYAI's own
        forgot-password using their school email.

    @node job-scheduler/app/services/myai_sync.py::read_initial_password
    """
    if not acc or not acc.init_pwd_enc:
        return None
    if acc.init_pwd_ack:
        clear_initial_password(db, acc)
        return None
    issued = acc.init_pwd_at
    if issued is not None and issued.tzinfo is None:
        issued = issued.replace(tzinfo=timezone.utc)
    if issued is None or (datetime.now(timezone.utc) - issued) > timedelta(days=retention_days):
        clear_initial_password(db, acc)
        return None
    from . import secrets_service
    try:
        return secrets_service.decrypt_value(acc.init_pwd_enc)
    except Exception as e:  # noqa: BLE001 - 金鑰換過/資料損壞 → 當作沒有
        logger.warning("初始密碼解密失敗（視為不存在）: %s", e)
        return None


def clear_initial_password(db: Session, acc) -> None:
    """ZH: 清除暫存的初始密碼（學生按「已修改」或逾期）| EN: purge the stored initial password

    @node job-scheduler/app/services/myai_sync.py::clear_initial_password
    """
    acc.init_pwd_enc = None
    acc.init_pwd_at = None
    db.commit()


def purge_expired_initial_passwords(db: Session, retention_days: int) -> int:
    """ZH: 批次清除逾期初始密碼（背景任務呼叫）| EN: purge expired initial passwords

    @node job-scheduler/app/services/myai_sync.py::purge_expired_initial_passwords
    """
    cutoff = datetime.now(timezone.utc) - timedelta(days=retention_days)
    rows = (
        db.query(models.ExternalAiAccount)
        .filter(models.ExternalAiAccount.init_pwd_enc.isnot(None))
        .all()
    )
    n = 0
    for acc in rows:
        issued = acc.init_pwd_at
        if issued is not None and issued.tzinfo is None:
            issued = issued.replace(tzinfo=timezone.utc)
        if acc.init_pwd_ack or issued is None or issued < cutoff:
            acc.init_pwd_enc = None
            acc.init_pwd_at = None
            n += 1
    if n:
        db.commit()
        logger.info("已清除 %d 筆逾期/已確認的 MYAI 初始密碼", n)
    return n


async def provision_user(db: Session, user) -> dict:
    """
    ZH: v3.3 自動開通主流程（首次 SSO 登入後以背景任務呼叫，**不阻塞登入**）。
        狀態機（每一步都可安全重入，重複呼叫不會重複建號）：
          disabled      → 功能旗標關閉
          bound         → 已有綁定，什麼都不用做
          linked_only   → 廠商端已有此 email（先前已註冊）→ 只補綁定，不建號
          created       → 呼叫廠商批次註冊建號 + 綁定 + 暫存初始密碼
          failed        → 廠商端失敗（保留錯誤訊息供 admin 檢視）
    EN: Auto-provision orchestrator; idempotent, runs in background after SSO login.

    @node job-scheduler/app/services/myai_sync.py::provision_user
    """
    from .. import crud
    if not crud.get_setting(db, "myai_autoprovision"):
        return {"status": "disabled"}

    email = (user.email or "").strip()
    if not email or email.endswith("@unknown"):
        return {"status": "skipped", "reason": "no_usable_email"}

    acc = (db.query(models.ExternalAiAccount)
             .filter(models.ExternalAiAccount.user_id == user.id).first())
    if acc and acc.vendor_username:
        return {"status": "bound"}

    # ==========================================================================
    # ZH: v3.4 建號前的雙重防呆（核心目的：**不要在 MYAI 造出綁不到人的垃圾帳號**）
    #   ① 先用「同一 local-part × 所有已知網域」查廠商端有沒有既有帳號 → 有就綁定不建號
    #      （教職員很可能早就有 liangyu@mail.mcu.edu.tw；只查我們推導的那個會漏掉）
    #   ② 推導不可信（規則沒命中，例如 sub 是員編數字）→ **跳過建號**，列入待人工處理
    # EN: v3.4 guards — look up existing vendor accounts across all known domains
    #     first, and refuse to create when the derived address isn't trustworthy.
    # ==========================================================================
    local = email.split("@")[0] if "@" in email else (user.username or "")
    known_domains = {(r.get("domain") or "").strip().lstrip("@")
                     for r in ((SSO_POLICY.get("oidc", {}) or {}).get("email_rules") or [])}
    known_domains.discard("")
    candidates = {email.lower()} | {f"{local}@{d}".lower() for d in known_domains if local}

    exist = None
    for cand in candidates:
        exist = (db.query(models.MyaiAccount)
                   .filter(models.MyaiAccount.email.ilike(cand)).first())
        if exist:
            if exist.email and exist.email.lower() != email.lower():
                logger.info("MYAI 既有帳號網域與推導不同，改用廠商端實際值：%s → %s",
                            email, exist.email)
                email = exist.email      # ZH: 以廠商端實際 email 為準
            break
    if exist:
        if not acc:
            acc = models.ExternalAiAccount(user_id=user.id, vendor_username=email,
                                           status="active", note="auto-provision(linked)")
            db.add(acc)
        acc.vendor_username = email
        acc.myai_vendor_sn = exist.vendor_sn
        db.commit()
        return {"status": "linked_only", "email": email}

    # ZH: ② 唯一的前置條件 —— **有沒有信箱**（事實），不是「信箱像不像真的」（預測）。
    #     沒有地址就無從建號；有地址就照建、照寄，帳號到底存不存在交給退件告訴我們
    #     （EmailLog 會記下寄給誰、被誰拒絕、誰不存在）。這裡刻意不做可信度評分。
    if not classify_email(email)["email"]:
        logger.info("MYAI 自動開通跳過 %s：沒有可用的 email（SSO 未提供且推導不出）。",
                    user.username)
        return {"status": "skipped", "reason": "no_email", "username": user.username}

    # ZH: 真正建號 —— 走廠商管理端官方批次註冊
    password = gen_initial_password()
    rows = [{"email": email, "nickname": _nickname_for(user),
             "password": password, "remark": "auto-provision"}]
    try:
        res = await register_batch(rows)
    except Exception as e:  # noqa: BLE001
        logger.error("MYAI 自動開通失敗 %s: %s", email, e)
        return {"status": "failed", "error": str(e)[:200]}
    # ZH: 🔴 判準是 `created`（第二段確認完成）而不是 `ok`（第一段預覽回 200）。
    #     這裡原本看的是 `ok`，而第二段當時根本沒實作 —— 於是「預覽成功」
    #     被寫成「帳號建好了」，還寄了開通信、存了對應不到任何帳號的初始密碼。
    if not res.get("created"):
        return {"status": "failed", "error": f"廠商回應 {res.get('status')}"}

    if not acc:
        acc = models.ExternalAiAccount(user_id=user.id, vendor_username=email,
                                       status="active", note="auto-provision")
        db.add(acc)
    acc.vendor_username = email
    db.commit()
    db.refresh(acc)
    store_initial_password(db, acc, password)
    logger.info("MYAI 自動開通完成: %s", email)

    # ZH: 開通通知信（每人只寄一次，不含密碼）。同時是**唯一的探針** ——
    #     SSO 路徑本來完全不寄信，不寄就永遠不會有退信，也就永遠不知道
    #     我們替他組出來的信箱到底存不存在。寄了、退了，才是事實。
    #     寄信是同步阻塞的，這裡在 async 流程中 → 丟到執行緒避免卡住事件迴圈。
    #     寄失敗絕不影響已完成的開通。
    # ZH: v3.9 開發階段可以關掉（myai_provision_email=0）—— 反覆拿真帳號測流程時，
    #     每一次成功都會寄一封信給真的學生信箱。
    #     ⚠️ 關掉的期間，上面那段「退信才是事實」的機制等於停擺。
    from .. import crud as _crud
    if not _crud.get_setting(db, "myai_provision_email"):
        logger.info("MYAI 開通通知信已關閉（myai_provision_email=0），未寄給 %s", email)
    else:
        try:
            import asyncio
            from .email_service import send_myai_provisioned
            url = _crud.get_system_config(db, "platform_public_url", "") or ""
            await asyncio.to_thread(send_myai_provisioned, email, _nickname_for(user), url)
        except Exception as e:  # noqa: BLE001
            logger.warning("MYAI 開通通知信寄送失敗（不影響開通）：%s", e)

    granted = await grant_initial_credit(db, acc, email)
    return {"status": "created", "email": email, "credit": granted}


async def grant_initial_credit(db: Session, acc, email: str) -> dict:
    """
    ZH: 發放新帳號的初始點數（`myai_initial_credit`，0 = 不發）。

    ZH: 三件事必須同時成立才會真的送出：設定值 > 0、這個帳號**沒發過**、
        呼叫端明確允許扣點。三者缺一就安靜地不發。

    ZH: 🔴 **冪等靠資料庫的事實，不靠流程。** `acc.credit_granted_at` 有值就直接回。
        自動開通是背景任務、SSO 可能重複觸發，只要有一次重入就會重複發放，
        而發出去的點數收不回來。

    ZH: 🔴 **失敗不重試、也不拋例外。** 發點數失敗不該讓「帳號已經建好」這件事
        看起來像失敗；把原因記在 credit_grant_note 裡，交給管理者人工處理。
        重試才是真正危險的選項 —— 第二段送出後失敗時，點數可能已經轉出了。

    @node job-scheduler/app/services/myai_sync.py::grant_initial_credit
    """
    from .. import crud
    if acc is None:
        return {"granted": False, "reason": "no_account"}
    if acc.credit_granted_at is not None:
        return {"granted": False, "reason": "already_granted",
                "points": acc.credit_granted_pts}
    try:
        points = int(crud.get_setting(db, "myai_initial_credit") or 0)
    except Exception:  # noqa: BLE001
        points = 0
    if points <= 0:
        return {"granted": False, "reason": "disabled"}

    rows = [{"email": email, "points": points, "remark": "auto-provision"}]
    try:
        res = await transfer_credit_batch(rows, confirm_grant=True)
    except Exception as e:  # noqa: BLE001
        # ZH: 中止在預覽階段（對帳不符、上傳失敗…）—— 點數確定沒有轉出。
        acc.credit_grant_note = f"failed: {str(e)[:180]}"
        db.commit()
        logger.error("MYAI 初始點數發放失敗 %s：%s", email, e)
        return {"granted": False, "reason": "error", "error": str(e)[:200]}

    if not res.get("granted"):
        # ZH: 🔴 第二段送出後失敗 —— 點數**可能已經轉出**。標成 unknown 讓人去對帳，
        #     並且**照樣寫 granted_at**，這樣重入時不會再送一次。
        acc.credit_granted_at = datetime.now(timezone.utc)
        acc.credit_granted_pts = 0
        acc.credit_grant_note = (f"unknown: 確認階段回應 {res.get('count')} 筆但未成功，"
                                 f"請到廠商後台對帳（預計 {points} 點）")
        db.commit()
        logger.error("MYAI 初始點數狀態未知 %s（%d 點）—— 不重試，請人工對帳", email, points)
        return {"granted": False, "reason": "unknown", "points": points}

    acc.credit_granted_at = datetime.now(timezone.utc)
    acc.credit_granted_pts = points
    acc.credit_grant_note = "ok"
    db.commit()
    logger.info("MYAI 初始點數發放完成 %s：%d 點", email, points)
    return {"granted": True, "points": points}


def provision_status(db: Session, user) -> dict:
    """
    ZH: 學生端查詢自己的開通狀態。**只在保留期內、且尚未確認修改**時才回傳初始密碼。
        身分一律由 JWT 推導（呼叫端傳 user 物件），查不到別人的。
    EN: Per-user provisioning status; the initial password is returned only while
        within the retention window and not yet acknowledged.

    @node job-scheduler/app/services/myai_sync.py::provision_status
    """
    from .. import crud
    acc = (db.query(models.ExternalAiAccount)
             .filter(models.ExternalAiAccount.user_id == user.id).first())
    if not acc:
        return {"provisioned": False, "email": None, "initial_password": None}
    days = crud.get_setting(db, "myai_init_pwd_days")
    pwd = read_initial_password(db, acc, days)
    return {
        "provisioned": True,
        "email": acc.vendor_username,
        "initial_password": pwd,                     # None = 已確認/逾期/本來就沒有
        "acknowledged": bool(acc.init_pwd_ack),
        "retention_days": days,
    }


def acknowledge_initial_password(db: Session, user) -> bool:
    """ZH: 學生按「我已修改密碼」→ 立即銷毀暫存的初始密碼

    @node job-scheduler/app/services/myai_sync.py::acknowledge_initial_password
    """
    acc = (db.query(models.ExternalAiAccount)
             .filter(models.ExternalAiAccount.user_id == user.id).first())
    if not acc:
        return False
    acc.init_pwd_ack = 1
    clear_initial_password(db, acc)
    return True


# ==============================================================================
# ZH: v3.4 MYAI 即時使用狀態（四象限）—— 交叉比對「MYAI 近期用量」×「平台在線」
# EN: v3.4 live MYAI usage quadrants — recent vendor usage × platform online
# ------------------------------------------------------------------------------
# ZH: 用途有二：
#   (a) 監控：現在有多少人真的在用 MYAI、用哪些模型
#   (b) 稽核：**有用量但人不在平台** → 可能在別台電腦使用，或共用機台沒登出被盜用
#       （呼應共用機台換手問題；這象限刻意標紅）
#   ⚠️ 資料新鮮度受輪詢限制：交易來自每 N 分鐘輪詢廠商，故「當前」實為「近即時」，
#      最壞情況落後一個輪詢週期。UI 需明示上次同步時間，別讓人誤以為是即時推播。
# ==============================================================================
ONLINE_WINDOW_MINUTES = 10      # ZH: 平台在線判定（對齊 admin 的 _ONLINE_THRESHOLD）
USAGE_WINDOW_MINUTES = 15       # ZH: MYAI「近期有用量」判定（含輪詢延遲的緩衝）


def live_usage_quadrants(db: Session, usage_minutes: int = USAGE_WINDOW_MINUTES,
                         online_minutes: int = ONLINE_WINDOW_MINUTES) -> dict:
    """
    ZH: 回傳四象限狀態。象限＝(MYAI 近期有 ai_usage) × (平台近期在線)
          using_active   ✅✅ 正常使用中
          using_offplat  ✅❌ **有用量但人不在平台**（稽核重點）
          online_idle    ❌✅ 在平台但沒用 AI
          （❌❌ 閒置者不列出，避免清單無意義地變長）
        身分關聯：external_ai_accounts 的 myai_vendor_sn 優先、退回 vendor_username(email)。
        未綁定但有用量的廠商帳號另列 unlinked（多半是老師/管理者或尚未綁定的學生）。
    EN: Cross-tab of recent vendor usage × platform presence; unlinked vendor rows listed separately.

    @node job-scheduler/app/services/myai_sync.py::live_usage_quadrants
    """
    now = datetime.now(timezone.utc)
    usage_cut = now - timedelta(minutes=max(1, usage_minutes))
    online_cut = now - timedelta(minutes=max(1, online_minutes))

    # ZH: 近期 ai_usage 逐筆（只取扣點事件；login/transfer 不算「在使用」）
    rows = (
        db.query(models.MyaiTransaction)
        .filter(models.MyaiTransaction.event_type == "ai_usage")
        .filter(models.MyaiTransaction.occurred_at >= usage_cut.replace(tzinfo=None))
        .all()
    )
    # 依廠商帳號彙總：最後一次活動、期間內筆數與消耗點數、用過的模型
    by_vendor: dict = {}
    for t in rows:
        key = (t.vendor_sn or "").strip() or (t.email or "").strip().lower()
        if not key:
            continue
        agg = by_vendor.setdefault(key, {
            "vendor_sn": t.vendor_sn, "email": t.email, "name": t.name,
            "last_at": None, "events": 0, "points": 0, "models": set(),
        })
        agg["events"] += 1
        agg["points"] += abs(t.points_delta or 0)
        if t.model:
            agg["models"].add(t.model)
        occ = t.occurred_at
        if occ is not None and (agg["last_at"] is None or occ > agg["last_at"]):
            agg["last_at"] = occ
            agg["email"] = t.email or agg["email"]
            agg["name"] = t.name or agg["name"]

    # ZH: 綁定表 → 平台使用者（sn 與 email 兩種索引都建，對應查詢順序）
    accs = db.query(models.ExternalAiAccount).all()
    user_ids = {a.user_id for a in accs}
    users = {u.id: u for u in db.query(models.User).filter(models.User.id.in_(user_ids)).all()} if user_ids else {}
    by_sn, by_email = {}, {}
    for a in accs:
        u = users.get(a.user_id)
        if not u:
            continue
        if a.myai_vendor_sn:
            by_sn[str(a.myai_vendor_sn).strip()] = u
        if a.vendor_username:
            by_email[a.vendor_username.strip().lower()] = u

    def _online(u) -> bool:
        """@node job-scheduler/app/services/myai_sync.py::live_usage_quadrants.<nested@870>._online"""
        last = getattr(u, "last_activity", None)
        if not last:
            return False
        if last.tzinfo is None:
            last = last.replace(tzinfo=timezone.utc)
        return last >= online_cut

    used_user_ids = set()
    using_active, using_offplat, unlinked = [], [], []
    for key, agg in by_vendor.items():
        u = by_sn.get(key) or by_email.get(key)
        last_at = agg["last_at"]
        item = {
            "vendor_sn": agg["vendor_sn"],
            "email": agg["email"],
            "vendor_name": agg["name"],
            "last_used_at": last_at.isoformat() if last_at else None,
            "events": agg["events"],
            "points_used": agg["points"],
            "models": sorted(agg["models"]),
            "username": getattr(u, "username", None),
            "user_id": getattr(u, "id", None),
            "role": getattr(u, "role", None),
        }
        if u is None:
            unlinked.append(item)
            continue
        used_user_ids.add(u.id)
        (using_active if _online(u) else using_offplat).append(item)

    # ZH: 在平台但近期沒動 MYAI（只列非 admin，admin 開著後台不算「使用者在用」）
    # ZH: ⚠️ v3.8 身分與權限拆開後這裡**刻意仍看 `role`** ——
    #     它排除的是「純粹在開後台的系統操作者帳號」,不是在做權限判定。
    #     改成 is_admin 的話,一個學生兼管理員的**真實學生用量**也會被排除掉。
    online_idle = []
    for u in db.query(models.User).filter(models.User.role != "admin").all():
        if u.id in used_user_ids or not _online(u):
            continue
        la = u.last_activity
        if la is not None and la.tzinfo is None:
            la = la.replace(tzinfo=timezone.utc)
        online_idle.append({
            "username": u.username, "user_id": u.id, "email": u.email,
            "last_activity": la.isoformat() if la else None,
        })

    latest_sync = db.query(models.MyaiTransaction.synced_at).order_by(
        models.MyaiTransaction.synced_at.desc()).first()
    return {
        "using_active": sorted(using_active, key=lambda x: x["last_used_at"] or "", reverse=True),
        "using_offplat": sorted(using_offplat, key=lambda x: x["last_used_at"] or "", reverse=True),
        "online_idle": sorted(online_idle, key=lambda x: x["last_activity"] or "", reverse=True),
        "unlinked": sorted(unlinked, key=lambda x: x["last_used_at"] or "", reverse=True),
        "usage_window_minutes": usage_minutes,
        "online_window_minutes": online_minutes,
        "last_tx_sync": latest_sync[0].isoformat() if latest_sync and latest_sync[0] else None,
    }


def has_online_users(db: Session, online_minutes: int = ONLINE_WINDOW_MINUTES) -> bool:
    """
    ZH: 平台上是否有「非 admin」使用者在線 —— 給輪詢迴圈判斷要不要跳過這一輪。
        排除 admin 的理由：管理者開著後台不代表有學生在用 MYAI，否則永遠不會休息。
    EN: Whether any non-admin user is active; drives the poll-skip optimization.

    @node job-scheduler/app/services/myai_sync.py::has_online_users
    """
    cut = (datetime.now(timezone.utc) - timedelta(minutes=max(1, online_minutes))).replace(tzinfo=None)
    return db.query(models.User).filter(
        models.User.role != "admin",
        models.User.last_activity.isnot(None),
        models.User.last_activity >= cut,
    ).first() is not None


# ==============================================================================
# ZH: v3.4 email 可信度判定 —— 決定「能不能拿這個 email 去廠商端建帳號」
# EN: v3.4 email trust check — gates vendor account creation
# ------------------------------------------------------------------------------
# ZH: 動機（使用者 2026-08-06 明示）：這整套規則的目的**不是為了寄信**，而是
#     **不要在 MYAI 亂建帳號**——建錯 email 會產生「綁不到人 / 沒人用」的殭屍帳號，
#     而且該老師可能早就有自己的帳號 → 變成兩個帳號、點數分裂。
#     因此判定從嚴：只有 IdP 直接給的、或規則明確命中的，才允許建號。
# ==============================================================================
TRUSTED_EMAIL_SOURCES = ("idp",)      # ZH: 規則命中的另外用 startswith("rule:") 判斷


def account_for_user(db: Session, user):
    """
    ZH: 取某位平台使用者對應的 MYAI 帳號列（綁定 sn 優先 → 廠商帳號 → email）；無則 None。

    ZH: 🔴 **全站唯一的對照邏輯。** 原本住在 routers/external_ai.py,
        但排程寄信也需要同一份 —— 服務層不能反向 import router,
        於是搬到這裡,router 改成呼叫它。

    ZH: 三段退路的順序是有意義的：`vendor_sn` 是廠商那邊的穩定鍵（改名改信箱都不變）,
        email 比對只是最後手段 —— 廠商後台的信箱欄位使用者自己改得動。

    @node job-scheduler/app/services/myai_sync.py::account_for_user
    """
    acc = crud.get_external_account_by_user_id(db, user.id)
    row = None
    if acc:
        if acc.myai_vendor_sn:
            row = db.query(models.MyaiAccount).filter(
                models.MyaiAccount.vendor_sn == acc.myai_vendor_sn).first()
        if not row and acc.vendor_username:
            row = db.query(models.MyaiAccount).filter(
                models.MyaiAccount.email.ilike(acc.vendor_username)).first()
    if not row and user.email:
        row = db.query(models.MyaiAccount).filter(
            models.MyaiAccount.email.ilike(user.email)).first()
    return row


def _alerted_recently(db: Session, user_id: str, stage: str, days: int) -> bool:
    """
    ZH: 這位使用者的這個階段,近 `days` 天內是否已經寄過。

    ZH: 🔴 節流是**分人又分階段**的：
          · 不分人 → 一天只寄得出一封,第二個人永遠收不到。
          · 不分階段 → 昨天寄了「快用完」,今天真的用完了卻被當成重複而不寄 ——
            那正是最需要通知的一刻。
        兩個階段各自獨立計時,所以「快用完 → 用完」一定寄得出兩封。

    ZH: 用 email_log 當節流狀態,不另開表：它本來就記了 kind/user_id/時間,
        而且**寄失敗也會留紀錄** —— 這是刻意的,SMTP 掛掉時不該把人的信箱洗版。

    @node job-scheduler/app/services/myai_sync.py::_alerted_recently
    """
    since = datetime.now(timezone.utc) - timedelta(days=days)
    return db.query(models.EmailLog).filter(
        models.EmailLog.user_id == user_id,
        models.EmailLog.kind == email_service.MYAI_BALANCE_KIND_PREFIX + stage,
        models.EmailLog.created_at >= since,
    ).first() is not None


def notify_balance_alerts(db: Session) -> dict:
    """
    ZH: MYAI 點數的兩段提醒（快用完／已用完）—— 寄信的部分。畫面提示走 /external-ai/my-balance。

    ZH: 只寄給**啟用中、有信箱、且真的綁得到 MYAI 帳號**的人。
        `state == "unknown"`（沒綁帳號）不寄 —— 那個人根本還沒開始用,
        提醒他「額度用完」是錯的。

    ZH: 回傳計數而不是 None,是為了讓排程的日誌看得出「跑了但沒寄」與「根本沒跑」的差別。

    @node job-scheduler/app/services/myai_sync.py::notify_balance_alerts
    """
    days = crud.get_setting(db, "myai_balance_alert_days")
    if not days or int(days) <= 0:
        return {"sent": 0, "skipped": 0, "disabled": True}      # ZH: 0 = 管理員關掉寄信（畫面提示仍在）

    days = int(days)
    threshold = crud.myai_low_balance_threshold(db)
    guide = crud.get_system_config(db, "myai_apply_guide_url", "")
    sent = skipped = 0

    users = db.query(models.User).filter(
        models.User.is_active == True,           # noqa: E712
        models.User.email.isnot(None),
    ).all()

    for u in users:
        row = account_for_user(db, u)
        if row is None:
            continue
        stage = crud.myai_balance_state(row.points, threshold)
        if stage not in ("low", "empty"):
            continue
        if _alerted_recently(db, u.id, stage, days):
            skipped += 1
            continue
        try:
            email_service.send_myai_balance_alert(
                u.email, u.username or u.email, u.id, stage,
                int(row.points or 0), threshold, guide or "")
            sent += 1
        except Exception as e:
            # ZH: 一個人寄失敗不該讓其餘的人收不到 —— 這是批次,不是交易。
            logger.warning(f"MYAI balance alert failed for {u.id}: {e}")

    return {"sent": sent, "skipped": skipped, "disabled": False}


def classify_email(email: str) -> dict:
    """
    ZH: 依 sso_policy.yaml 的 email_rules 判定「這個信箱屬於哪一類」—— 只看**網域**。
        回 {"email":…, "domain":…, "label": "student"|"staff"|None}

        ⚠ 設計原則（2026-08-06 使用者定調，別再改回去）：
          1. **不對信箱做任何預測或評估**。這裡不算信心度、不判斷帳號存不存在，
             只做分類。帳號到底存不存在，由「實際寄送後的退件」告訴我們（EmailLog），
             那是事實；我們自己猜的信心度不是。
          2. **不拿使用者名稱當判定依據**。平台允許使用者自由更改名稱，
             拿可變欄位當身分判準本身就不成立。
    EN: Classify an address by DOMAIN only. No confidence scoring, no existence guessing —
        whether the mailbox exists is answered by real bounces (EmailLog), not by us.
        Never keys off the username: users can freely rename themselves.

    @node job-scheduler/app/services/myai_sync.py::classify_email
    """
    from ..config import SSO_POLICY
    cfg = (SSO_POLICY or {}).get("oidc", {}) or {}
    addr = (email or "").strip()
    # ZH: @unknown 是 SSO 完全取不到信箱時寫入的佔位值，不是地址 → 視為沒有信箱
    if not addr or "@" not in addr or addr.lower().endswith("@unknown"):
        return {"email": "", "domain": "", "label": None}
    domain = addr.split("@")[-1].lower()
    label = next((r.get("label") for r in (cfg.get("email_rules") or [])
                  if (r.get("domain") or "").strip().lstrip("@").lower() == domain), None)
    return {"email": addr, "domain": domain, "label": label}


def provision_candidates(db: Session) -> dict:
    """
    ZH: 給 admin 檢視 —— SSO 使用者中「尚未綁定 MYAI」的人，依**事實**分兩類：
          ready     有 email → 可以自動開通（信箱真假不預判，寄出後看退件）
          no_email  完全沒有 email（SSO 未提供、也推導不出）→ 無從建號，需人工補
        刻意不做「信心度」分類：帳號存不存在由退件紀錄回答，不由我們猜。
    EN: Unbound SSO users split by a FACT — has an address or not. No confidence scoring.

    @node job-scheduler/app/services/myai_sync.py::provision_candidates
    """
    bound = {a.user_id for a in db.query(models.ExternalAiAccount).all()}
    ready, no_email = [], []
    for u in db.query(models.User).filter(models.User.auth_source == "sso_oidc").all():
        if u.id in bound:
            continue
        info = classify_email(u.email)
        row = {"user_id": u.id, "username": u.username, "platform_email": u.email,
               "email": info["email"], "domain": info["domain"], "label": info["label"]}
        (ready if info["email"] else no_email).append(row)
    return {"ready": ready, "no_email": no_email}


def staff_pending(db: Session) -> list[dict]:
    """
    ZH: 疑似教職員清單 —— **信箱網域**屬於教職員域、但平台角色仍是 student 的 SSO 帳號。
        判定只看網域（不看使用者名稱，那是可自由更改的欄位）。
        ⚠️ **v3.8 起這句話變了**：建帳號時已經會依網域自動判角色（crud.role_from_email）,
        所以這張清單現在列的是「**v3.8 之前建立的**,或被管理者改回 student 的」帳號 ——
        也就是需要回填的那一批,不是「等待人工升權」的那一批。
    EN: SSO users whose mail DOMAIN says staff but whose platform role is still student.
        Domain-based only (never the mutable username). Never auto-promoted.

    @node job-scheduler/app/services/myai_sync.py::staff_pending
    """
    out: list[dict] = []
    users = (db.query(models.User)
               .filter(models.User.auth_source != "local",
                       models.User.role == "student")
               .all())
    for u in users:
        info = classify_email(u.email)
        if info["label"] == "staff":
            out.append({"user_id": u.id, "username": u.username,
                        "platform_email": u.email, "email": info["email"],
                        "domain": info["domain"], "role": u.role})
    return out
